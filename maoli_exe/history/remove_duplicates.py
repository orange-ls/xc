'''
功能：毛利表筛重
描述：
    1、【合同类型】列仅看“标准合同”“变更协议”，当【厂商PO号】一致，根据【合同编号】列及【审核时间】列，将“成本与总价”及“利润与分析”sheet中处最新时间数据保留
    2、筛选【合同类型】列中“增补协议”，当【厂商PO号】与【分析成本】一致，保留最新时间数据 其余删除。
'''

import pandas as pd
import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

def remove_duplicates(address, result_file, template_file):
    '''
    :param address: 毛利表的地址
    :param result_file: 输出Excel文件地址
    :param template_file: 模板Excel文件地址
    '''
    # 读取两个Sheet
    cost_df = pd.read_excel(address, sheet_name='成本与总价', engine='openpyxl')
    cost_df = cost_df.dropna(how="all")
    profit_df = pd.read_excel(address, sheet_name='利润与分析', engine='openpyxl')
    profit_df = profit_df.dropna(how="all")

    # 初始化两个空的DataFrame，用于存储保留和删除的数据
    retained_data = pd.DataFrame()
    deleted_data = pd.DataFrame()

    # 将审批时间列转换为datetime格式，方便比较
    cost_df['审核时间'] = pd.to_datetime(cost_df['审核时间'])

    # 先处理“成本与总价”Sheet中的重复数据
    # 处理【合同类型】列为 标准合同 和 变更协议
    standard_and_change = cost_df[cost_df['合同类型'].isin(['标准合同', '变更协议'])]
    grouped = standard_and_change.groupby('厂商PO号')

    for po, group in grouped:
        if len(group) > 1:
            group_del = group
            # 找到审批时间最新的行
            latest_row = group.loc[group['审核时间'].idxmax()]
            # 保留时间最新的数据
            group = group[group.index == latest_row.name]
            # 将删除的行添加到 deleted_data
            deleted_rows = group_del[group_del['审核时间'] != latest_row['审核时间']]
            deleted_data = pd.concat([deleted_data, deleted_rows])
        # 将保留的行添加到retained_data
        retained_data = pd.concat([retained_data, group])

    # 处理【合同类型】列为 增补协议
    supplement = cost_df[cost_df['合同类型'] == '增补协议']
    grouped = supplement.groupby(['厂商PO号', '分析成本'])

    for (po, cost), group in grouped:
        if len(group) > 1:
            group_del = group
            # 找到审批时间最新的行
            latest_row = group.loc[group['审核时间'].idxmax()]
            # 保留时间最新的数据
            group = group[group.index == latest_row.name]
            # 将删除的行添加到 deleted_data
            deleted_rows = group_del[group_del['审核时间'] != latest_row['审核时间']]
            deleted_data = pd.concat([deleted_data, deleted_rows])
        # 将保留的行添加到retained_data
        retained_data = pd.concat([retained_data, group])

    # 将合同类型不为'标准合同', '变更协议', '增补协议'的数据加入保留数据
    retained_data = pd.concat([retained_data, cost_df[~cost_df['合同类型'].isin(['标准合同', '变更协议', '增补协议'])]])

    # 重新把datetime格式的审批时间列转换为str格式
    if not retained_data.empty:
        retained_data['审核时间'] = retained_data['审核时间'].dt.strftime('%Y-%m-%d %H:%M:%S')
        retained_data['合同编号'] = retained_data['合同编号'].astype(str)
        retained_data = retained_data.sort_values(by='合同编号')
    if not deleted_data.empty:
        deleted_data['审核时间'] = deleted_data['审核时间'].dt.strftime('%Y-%m-%d %H:%M:%S')
        deleted_data['合同编号'] = deleted_data['合同编号'].astype(str)
        deleted_data = deleted_data.sort_values(by='合同编号')

    # 开始处理“利润与分析”Sheet中的重复数据
    # 获取“成本与总价” Sheet 中删除数据的合同编码
    if not deleted_data.empty:
        # deleted_serial_numbers = deleted_data[deleted_data['合同编号'].isna()]['序号'].unique()
        # deleted_contract_codes = deleted_data[deleted_data['合同编号'].notna()]['合同编号'].astype(str).unique()
        deleted_contract_codes = deleted_data[deleted_data['合同编号'].notna()][['合同编号', '审核时间']].drop_duplicates()
    else:
        deleted_contract_codes = []
        # deleted_serial_numbers = []

    # 在“利润与分析” Sheet 中删除对应的合同编码数据
    profit_df['合同编号'] = profit_df['合同编号'].astype(str)
    profit_df['审批时间'] = pd.to_datetime(profit_df['审批时间'], errors='coerce')
    profit_df['审批时间'] = profit_df['审批时间'].dt.strftime('%Y-%m-%d %H:%M:%S')
    # profit_df = profit_df[~profit_df['合同编号'].isin(deleted_contract_codes)]

    merged_df = profit_df.merge(deleted_contract_codes.rename(columns={'审核时间': '审批时间'}), on=['合同编号', '审批时间'], how='left', indicator=True)
    profit_df = merged_df[merged_df['_merge'] == 'left_only'].drop(columns='_merge')
    # 删除'合同编号'为空的对应数据
    # profit_df = profit_df[~profit_df['序号'].isin(deleted_serial_numbers)]

    # 复制模板表，并重命名
    finalPath = f'{result_file}/毛利表(筛重).xlsx'
    # 将template_file路径下的模板文件复制到finalPath路径下
    os.makedirs(os.path.dirname(finalPath), exist_ok=True)
    with open(template_file, 'rb') as f:
        data = f.read()
    with open(finalPath, 'wb') as f:
        f.write(data)

    # 加载模板文件
    wb = load_workbook(template_file)
    ws_cost = wb['成本与总价']
    ws_profit = wb['利润与分析']

    # 设置居中对齐样式
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 将retained_data中的"毛利率"列和"标准毛利率"列改为百分数形式,并保留两位小数
    retained_data['毛利率'] = retained_data['毛利率'].apply(lambda x: '{:.2f}%'.format(x * 100))
    retained_data['标准毛利率'] = retained_data['标准毛利率'].apply(lambda x: '{:.2f}%'.format(x * 100))

    # 写入“成本与总价”Sheet的数据
    for r, row in enumerate(retained_data.values, start=3):
        for c, value in enumerate(row, start=1):
            cell = ws_cost.cell(row=r, column=c, value=value)
            cell.alignment = center_alignment

    # 写入“利润与分析”Sheet的数据
    for r, row in enumerate(profit_df.values, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws_profit.cell(row=r, column=c, value=value)
            cell.alignment = center_alignment

    wb.save(finalPath)
    wb.close()

    # 输出删除的数据到文件
    with pd.ExcelWriter(os.path.join(result_file, '删除的数据.xlsx')) as writer:
        deleted_data.to_excel(writer, sheet_name='删除的数据', index=False)


if __name__ == '__main__':
    # # 输入Excel文件地址
    # address = 'C:/Users/user/Desktop/毛利表(1).xlsx'
    # # address = 'C:/Users/user/Desktop/毛利表.xlsx'
    # # 输出Excel文件地址
    # result_file = 'C:/Users/user/Desktop'
    # # 模板Excel文件地址
    # template_file = 'C:/Users/user/Desktop/模板.xlsx'

    # 初始化Tkinter（不显示主窗口）
    root = Tk()
    root.withdraw()

    # 弹出文件选择对话框，选择输入文件
    print("请选择毛利表：")
    address = filedialog.askopenfilename(
        title="选择输入文件",
        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
    )
    if not address:
        print("未选择文件，程序退出。")
        exit()

    # 弹出文件选择对话框，选择模板文件
    print("选择模板文件")
    template_file = filedialog.askopenfilename(
        title="选择输入文件",
        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
    if not template_file:
        print("未选择文件，程序退出。")
        exit()

    # 弹出文件夹选择对话框，选择输出文件夹
    print("选择保存文件夹")
    result_file = filedialog.askdirectory(title="选择输出文件夹")
    if not result_file:
        print("未选择输出文件夹，程序退出。")
        exit()

    remove_duplicates(address, result_file, template_file)
    print("程序执行完毕。")



