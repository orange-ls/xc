import pandas as pd
import cpca
from tkinter import Tk, filedialog
from openpyxl import load_workbook


# 读取Excel文件（请修改文件路径）
# input_file = "C:/Users/user/Desktop/国家管网服务器分布-机器视觉建设项目（二期）硬件标包框架第一批450台发货清单.xlsx"
# result_file = "C:/Users/user/Desktop"

# 初始化Tkinter（不显示主窗口）
root = Tk()
root.withdraw()

# 弹出文件选择对话框，选择输入文件
print("请选择要处理的Excel文件：")
input_file = filedialog.askopenfilename(
    title="选择输入文件",
    filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
)
if not input_file:
    print("未选择文件，程序退出。")
    exit()

# 弹出文件夹选择对话框，选择输出文件夹
print("选择保存文件夹")
result_file = filedialog.askdirectory(title="选择输出文件夹")
if not result_file:
    print("未选择输出文件夹，程序退出。")
    exit()

# 读取数据
df = pd.read_excel(input_file)

# 拼接地址字段（根据实际情况调整分隔符）
combined_address = df["作业区"].astype(str) + df["部署位置"].astype(str)

# 使用cpca解析地址
df_addr = cpca.transform(combined_address)

# 使用openpyxl直接操作Excel文件
wb = load_workbook(input_file)
ws = wb.active

# 定位或创建省份列
province_col = None
for cell in ws[1]:  # 遍历第一行
    if cell.value == "省份":
        province_col = cell.column  # 定位省份列
        break

if province_col is None:
    # 在最后一列新增省份列
    province_col = ws.max_column + 1
    ws.cell(row=1, column=province_col, value="省份")

# 在指定列写入省份数据
for row, province in enumerate(df_addr["省"], start=2):  # 从第2行开始
    ws.cell(row=row, column=province_col, value=province)

# 保存新文件（保留原格式）
wb.save(f"{result_file}/匹配所在省份.xlsx")

print("处理完成！结果已保存至", result_file)