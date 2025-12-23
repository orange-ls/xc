# -*- coding: utf-8 -*-
"""
Excel 数据修复工具 - 服务器整机存量信息导入模板修复
用于修复用户导入Excel时常见的格式问题

功能：
1. 日期格式修复（支持多种日期格式转换为标准格式）
2. 必填字段校验与提示
3. 空值处理
4. 数据格式标准化
5. 生成修复报告

使用方法：
1. 双击运行 excel_fixer.exe
2. 选择需要修复的Excel文件
3. 工具会自动修复并生成新文件
"""

import os
import sys
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading


class ExcelFixer:
    """Excel修复工具核心类"""
    
    # 字段配置：(中文名, 字段名, 是否必填, 字段类型)
    FIELD_CONFIG = [
        ('系统编号', 'unique_id', False, 'char'),
        ('交货单号', 'delivery_no', False, 'char'),
        ('整机SN', 'complete_sn', True, 'char'),
        ('服务开始时间', 'service_start_time', True, 'date'),
        ('维保服务结束时间', 'maintenance_service_end_date', True, 'date'),
        ('服务产品类别(内部)', 'service_product_type', True, 'char'),
        ('服务产品类别(外部)', 'service_product_type_out', False, 'char'),
        ('签约客户名称', 'customer_name', False, 'char'),
        ('项目名称', 'proj_name', False, 'char'),
        ('CRM立项编号', 'crm_no', True, 'char'),
        ('项目所属行业/区域', 'proj_industry', False, 'char'),
        ('ConfigNo.', 'config_no', False, 'char'),
        ('整机型号', 'complete_model', True, 'char'),
        ('整机销售员', 'complete_sale', False, 'char'),
        ('整机采购时间', 'complete_purchase_time', False, 'date'),
        ('所在省份', 'province', False, 'char'),
        ('所在城市', 'city', False, 'char'),
        ('交付地址', 'delivery_address', False, 'char'),
        ('产品交付时间', 'deliver_time', False, 'date'),
        ('机型', 'model', False, 'char'),
        ('备注', 'remark', False, 'char'),
        ('服务方', 'producer', False, 'char'),
        ('项目审批一线工程师', 'sales', False, 'char'),
    ]
    
    # 日期字段列表
    DATE_FIELDS = ['服务开始时间', '维保服务结束时间', '整机采购时间', '产品交付时间']
    
    # 必填字段列表
    REQUIRED_FIELDS = ['整机SN', '服务开始时间', '维保服务结束时间', '服务产品类别(内部)', 'CRM立项编号', '整机型号']
    
    # 特殊服务类型（这些类型不需要日期）
    NO_DATE_SERVICE_TYPES = ['无服务', '华为服务']
    
    def __init__(self, callback=None):
        """
        初始化修复工具
        :param callback: 日志回调函数
        """
        self.callback = callback
        self.errors = []
        self.warnings = []
        self.fixed_count = 0
        self.error_count = 0
        
    def log(self, message, level='INFO'):
        """输出日志"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_msg = f"[{timestamp}] [{level}] {message}"
        if self.callback:
            self.callback(log_msg)
        print(log_msg)
        
    def parse_date(self, value, row_num, field_name):
        """
        解析并修复日期格式
        支持的格式：
        - 2024-01-01, 2024/01/01, 2024.01.01
        - 01-01-2024, 01/01/2024
        - 20240101
        - Excel日期序列号
        - 中文日期：2024年1月1日
        """
        if value is None or str(value).strip() in ['', 'None', 'null', 'NULL', '暂无', '#N/A', 'N/A']:
            return None, None
            
        original_value = value
        
        # 如果已经是date或datetime对象
        if isinstance(value, (date, datetime)):
            return value.strftime('%Y-%m-%d'), None
            
        value_str = str(value).strip()
        
        # 尝试多种日期格式
        date_formats = [
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y.%m.%d',
            '%d-%m-%Y',
            '%d/%m/%Y',
            '%m-%d-%Y',
            '%m/%d/%Y',
            '%Y%m%d',
            '%Y年%m月%d日',
            '%Y-%m-%d %H:%M:%S',
            '%Y/%m/%d %H:%M:%S',
        ]
        
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(value_str, fmt)
                fixed_value = parsed_date.strftime('%Y-%m-%d')
                if value_str != fixed_value:
                    self.fixed_count += 1
                    return fixed_value, f"第{row_num}行 [{field_name}]: '{original_value}' -> '{fixed_value}'"
                return fixed_value, None
            except ValueError:
                continue
        
        # 尝试解析Excel日期序列号
        try:
            if value_str.replace('.', '').isdigit():
                excel_date = float(value_str)
                if 1 < excel_date < 100000:  # 合理的Excel日期范围
                    # Excel日期从1900-01-01开始（但有1900年闰年bug）
                    from datetime import timedelta
                    base_date = datetime(1899, 12, 30)
                    parsed_date = base_date + timedelta(days=excel_date)
                    fixed_value = parsed_date.strftime('%Y-%m-%d')
                    self.fixed_count += 1
                    return fixed_value, f"第{row_num}行 [{field_name}]: Excel序列号 '{original_value}' -> '{fixed_value}'"
        except (ValueError, OverflowError):
            pass
            
        # 无法解析
        self.error_count += 1
        return None, f"第{row_num}行 [{field_name}]: 无法解析日期格式 '{original_value}'，请手动修正"
    
    def clean_string(self, value):
        """清理字符串值"""
        if value is None:
            return None
        value_str = str(value).strip()
        if value_str in ['', 'None', 'null', 'NULL', 'nan', 'NaN']:
            return None
        # 移除不可见字符
        value_str = re.sub(r'[\x00-\x1F\x7F]', '', value_str)
        return value_str
    
    def validate_required_field(self, value, row_num, field_name, service_type=None):
        """验证必填字段"""
        # 特殊处理：如果服务类型是"无服务"或"华为服务"，日期字段可以为空
        if field_name in self.DATE_FIELDS and service_type in self.NO_DATE_SERVICE_TYPES:
            return True, None
            
        if value is None or str(value).strip() in ['', 'None', 'null', 'NULL', '暂无', '#N/A']:
            self.error_count += 1
            return False, f"第{row_num}行 [{field_name}]: 必填字段为空，请补充数据"
        return True, None
    
    def fix_excel(self, input_path, output_path=None):
        """
        修复Excel文件
        :param input_path: 输入文件路径
        :param output_path: 输出文件路径（默认在原文件名后加_fixed）
        :return: (success, output_path, report)
        """
        self.errors = []
        self.warnings = []
        self.fixed_count = 0
        self.error_count = 0
        
        if output_path is None:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_已修复{ext}"
        
        self.log(f"开始处理文件: {input_path}")
        
        try:
            # 加载工作簿
            wb = load_workbook(input_path, data_only=True)
            sheet = wb.active
            
            # 获取表头
            headers = [cell.value for cell in sheet[1]]
            self.log(f"检测到 {len(headers)} 列，{sheet.max_row - 1} 行数据")
            
            # 创建表头到列索引的映射
            header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}
            
            # 检查必要的列是否存在
            missing_required = []
            for field in self.REQUIRED_FIELDS:
                if field not in header_map:
                    missing_required.append(field)
            
            if missing_required:
                self.log(f"警告: 缺少必填列: {', '.join(missing_required)}", 'WARNING')
                self.warnings.append(f"文件缺少必填列: {', '.join(missing_required)}")
            
            # 获取服务产品类别列索引
            service_type_col = header_map.get('服务产品类别(内部)')
            
            # 创建新工作簿
            new_wb = Workbook()
            new_sheet = new_wb.active
            new_sheet.title = "修复后数据"
            
            # 定义样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, name="微软雅黑", size=11)
            error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            fixed_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            normal_font = Font(name="微软雅黑", size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = new_sheet.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                new_sheet.column_dimensions[get_column_letter(col_idx)].width = 18
            
            # 添加"修复说明"列
            remark_col = len(headers) + 1
            remark_cell = new_sheet.cell(row=1, column=remark_col, value="修复说明")
            remark_cell.fill = header_fill
            remark_cell.font = header_font
            remark_cell.alignment = center_alignment
            remark_cell.border = thin_border
            new_sheet.column_dimensions[get_column_letter(remark_col)].width = 50
            
            # 处理数据行
            for row_idx in range(2, sheet.max_row + 1):
                row_data = [cell.value for cell in sheet[row_idx]]
                row_remarks = []
                has_error = False
                has_fix = False
                
                # 获取当前行的服务类型
                service_type = None
                if service_type_col is not None and service_type_col < len(row_data):
                    service_type = self.clean_string(row_data[service_type_col])
                
                # 处理每个单元格
                for col_idx, value in enumerate(row_data):
                    if col_idx >= len(headers) or headers[col_idx] is None:
                        continue
                        
                    field_name = str(headers[col_idx]).strip()
                    new_value = value
                    
                    # 日期字段处理
                    if field_name in self.DATE_FIELDS:
                        # 特殊服务类型不需要日期
                        if service_type in self.NO_DATE_SERVICE_TYPES:
                            new_value = None
                            if value is not None and str(value).strip() != '':
                                row_remarks.append(f"[{field_name}]: 服务类型为'{service_type}'，日期已清空")
                                has_fix = True
                        else:
                            fixed_value, remark = self.parse_date(value, row_idx, field_name)
                            new_value = fixed_value
                            if remark:
                                if '无法解析' in remark:
                                    self.errors.append(remark)
                                    has_error = True
                                else:
                                    row_remarks.append(remark.split(': ', 1)[1] if ': ' in remark else remark)
                                    has_fix = True
                    else:
                        # 字符串字段清理
                        new_value = self.clean_string(value)
                    
                    # 必填字段验证
                    if field_name in self.REQUIRED_FIELDS:
                        is_valid, error_msg = self.validate_required_field(
                            new_value, row_idx, field_name, service_type
                        )
                        if not is_valid:
                            self.errors.append(error_msg)
                            has_error = True
                            row_remarks.append(f"[{field_name}]: 必填字段为空")
                    
                    row_data[col_idx] = new_value
                
                # 写入数据行
                for col_idx, value in enumerate(row_data, 1):
                    cell = new_sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = normal_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    if has_error:
                        cell.fill = error_fill
                    elif has_fix:
                        cell.fill = fixed_fill
                
                # 写入修复说明
                remark_text = "; ".join(row_remarks) if row_remarks else ""
                remark_cell = new_sheet.cell(row=row_idx, column=remark_col, value=remark_text)
                remark_cell.font = normal_font
                remark_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                remark_cell.border = thin_border
                
                if row_idx % 100 == 0:
                    self.log(f"已处理 {row_idx - 1} 行...")
            
            # 保存文件
            new_wb.save(output_path)
            self.log(f"文件已保存: {output_path}")
            
            # 生成报告
            report = self.generate_report()
            
            return True, output_path, report
            
        except Exception as e:
            self.log(f"处理失败: {str(e)}", 'ERROR')
            return False, None, str(e)
    
    def generate_report(self):
        """生成修复报告"""
        report_lines = [
            "=" * 60,
            "Excel 数据修复报告",
            "=" * 60,
            f"修复时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"自动修复: {self.fixed_count} 处",
            f"错误数量: {self.error_count} 处",
            "",
        ]
        
        if self.warnings:
            report_lines.append("【警告信息】")
            for warning in self.warnings:
                report_lines.append(f"  ⚠ {warning}")
            report_lines.append("")
        
        if self.errors:
            report_lines.append("【错误信息】（需要手动修正）")
            for error in self.errors[:50]:  # 最多显示50条
                report_lines.append(f"  ✗ {error}")
            if len(self.errors) > 50:
                report_lines.append(f"  ... 还有 {len(self.errors) - 50} 条错误")
            report_lines.append("")
        
        report_lines.extend([
            "【颜色说明】",
            "  🟢 绿色背景: 已自动修复的数据",
            "  🔴 红色背景: 存在错误需要手动修正",
            "",
            "【修复说明】",
            "  1. 日期格式已统一转换为 YYYY-MM-DD 格式",
            "  2. 空白字符和特殊字符已清理",
            "  3. '无服务'和'华为服务'类型的日期字段已清空",
            "  4. 必填字段为空的行已标记为红色",
            "",
            "=" * 60,
        ])
        
        return "\n".join(report_lines)


class ExcelFixerGUI:
    """Excel修复工具图形界面"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel 数据修复工具 - 服务器整机存量信息 v1.0")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # 设置图标（如果有的话）
        try:
            self.root.iconbitmap("icon.ico")
        except:
            pass
        
        self.setup_ui()
        self.fixer = None
        
    def setup_ui(self):
        """设置界面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 标题
        title_label = ttk.Label(
            main_frame, 
            text="Excel 数据修复工具", 
            font=("微软雅黑", 16, "bold")
        )
        title_label.pack(pady=(0, 5))
        
        subtitle_label = ttk.Label(
            main_frame, 
            text="服务器整机存量信息导入模板修复", 
            font=("微软雅黑", 10)
        )
        subtitle_label.pack(pady=(0, 15))
        
        # 文件选择框架
        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.file_path_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=70)
        file_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        browse_btn = ttk.Button(file_frame, text="浏览...", command=self.browse_file)
        browse_btn.pack(side=tk.LEFT)
        
        # 操��按钮框架
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.fix_btn = ttk.Button(
            btn_frame, 
            text="开始修复", 
            command=self.start_fix,
            width=20
        )
        self.fix_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.open_btn = ttk.Button(
            btn_frame, 
            text="打开输出文件", 
            command=self.open_output,
            width=20,
            state=tk.DISABLED
        )
        self.open_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        clear_btn = ttk.Button(
            btn_frame, 
            text="清空日志", 
            command=self.clear_log,
            width=15
        )
        clear_btn.pack(side=tk.RIGHT)
        
        # 进度条
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            main_frame, 
            variable=self.progress_var, 
            maximum=100
        )
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        
        # 日志框架
        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, 
            wrap=tk.WORD, 
            font=("Consolas", 9),
            height=20
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(
            main_frame, 
            textvariable=self.status_var, 
            relief=tk.SUNKEN, 
            anchor=tk.W
        )
        status_bar.pack(fill=tk.X, pady=(10, 0))
        
        # 输出文件路径
        self.output_path = None
        
        # 显示使用说明
        self.show_help()
        
    def show_help(self):
        """显示使用说明"""
        help_text = """
╔══════════════════════════════════════════════════════════════╗
║                    Excel 数据修复工具 使用说明                    ║
╠══════════════════════════════════════════════════════════════╣
║                                                              ║
║  【功能说明】                                                  ║
║  本工具用于修复"服务器整机存量信息"导入Excel的常见格式问题：          ║
║  ✓ 日期格式自动转换（支持多种格式）                               ║
║  ✓ 必填字段校验与提示                                          ║
║  ✓ 空值和特殊字符清理                                          ║
║  ✓ 生成详细的修复报告                                          ║
║                                                              ║
║  【必填字段】                                                  ║
║  • 整机SN                                                    ║
║  • 服务开始时间                                                ║
║  • 维保服务结束时间                                            ║
║  • 服务产品类别(内部)                                          ║
║  • CRM立项编号                                                ║
║  • 整机型号                                                   ║
║                                                              ║
║  【使用步骤】                                                  ║
║  1. 点击"浏览"选择需要修复的Excel文件                            ║
║  2. 点击"开始修复"进行处理                                      ║
║  3. 查看日志了解修复详情                                        ║
║  4. 点击"打开输出文件"查看修复后的文件                            ║
║                                                              ║
║  【输出说明】                                                  ║
║  • 修复后的文件保存在原文件同目录，文件名添加"_已修复"后缀           ║
║  • 绿色背景：已自动修复的数据                                    ║
║  • 红色背景：存在错误需要手动修正                                 ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝
"""
        self.log_text.insert(tk.END, help_text)
        
    def browse_file(self):
        """浏览文件"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[
                ("Excel文件", "*.xlsx"),
                ("所有文件", "*.*")
            ]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.status_var.set(f"已选择文件: {os.path.basename(file_path)}")
            
    def log_callback(self, message):
        """日志回调"""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def start_fix(self):
        """开始修复"""
        file_path = self.file_path_var.get()
        if not file_path:
            messagebox.showwarning("提示", "请先选择要修复的Excel文件")
            return
            
        if not os.path.exists(file_path):
            messagebox.showerror("错误", "文件不存在，请重新选择")
            return
        
        # 禁用按钮
        self.fix_btn.config(state=tk.DISABLED)
        self.open_btn.config(state=tk.DISABLED)
        self.progress_var.set(0)
        self.status_var.set("正在处理...")
        
        # 清空日志
        self.log_text.delete(1.0, tk.END)
        
        # 在新线程中执行修复
        thread = threading.Thread(target=self.do_fix, args=(file_path,))
        thread.start()
        
    def do_fix(self, file_path):
        """执行修复（在新线程中）"""
        try:
            self.fixer = ExcelFixer(callback=self.log_callback)
            
            self.progress_var.set(10)
            success, output_path, report = self.fixer.fix_excel(file_path)
            
            self.progress_var.set(90)
            
            if success:
                self.output_path = output_path
                self.log_callback("\n" + report)
                self.progress_var.set(100)
                self.status_var.set(f"修复完成！输出文件: {os.path.basename(output_path)}")
                self.open_btn.config(state=tk.NORMAL)
                
                # 显示完成提示
                self.root.after(0, lambda: messagebox.showinfo(
                    "完成", 
                    f"文件修复完成！\n\n"
                    f"自动修复: {self.fixer.fixed_count} 处\n"
                    f"错误数量: {self.fixer.error_count} 处\n\n"
                    f"输出文件:\n{output_path}"
                ))
            else:
                self.status_var.set("修复失败")
                self.root.after(0, lambda: messagebox.showerror("错误", f"修复失败:\n{report}"))
                
        except Exception as e:
            self.log_callback(f"[ERROR] 发生异常: {str(e)}")
            self.status_var.set("处理出错")
            self.root.after(0, lambda: messagebox.showerror("错误", f"处理出错:\n{str(e)}"))
        finally:
            self.fix_btn.config(state=tk.NORMAL)
            
    def open_output(self):
        """打开输出文件"""
        if self.output_path and os.path.exists(self.output_path):
            os.startfile(self.output_path)
        else:
            messagebox.showwarning("提示", "输出文件不存在")
            
    def clear_log(self):
        """清空日志"""
        self.log_text.delete(1.0, tk.END)
        self.show_help()
        
    def run(self):
        """运行程序"""
        self.root.mainloop()


def main():
    """主函数"""
    app = ExcelFixerGUI()
    app.run()


if __name__ == "__main__":
    main()

