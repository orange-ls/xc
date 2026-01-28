import gc
import tkinter as tk

from tkinter import filedialog
import threading
import os
import numpy as np
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
import result_table_processing
from python_calamine import CalamineWorkbook
import openpyxl
from decimal import Decimal
from tkcalendar import DateEntry
from urllib.parse import quote_plus

# 字段映射配置
column_mapping = {
    '业绩ID': 'performance_id',
    '业绩金额(¥)': 'sales_amount',
    '业绩形成时间': 'performance_date',
    '特殊返点类型': 'special_rebate_type',
    '二级经销商名称': 'secondary_dealer',
    '客户名称': 'customer_name',
    '产品类型编码': 'product_code',
    '客户标签': 'customer_tag',
    '销售纵队': 'sales_team',
    '服务产品部': 'service_department',
    '是否流量型产品': 'is_traffic_product',
    '专线产品': 'leased_line_product',
    '企业协同': 'enterprise_coop',
    '销售员': 'salesperson',
    '区域': 'region',
    '季度': 'quarter'
}
selected_columns = list(column_mapping.keys())

class App(object):
    def __init__(self, root):
        self.filePath = {}

        root.title("华为云RPA")
        root.geometry('500x500')

        self.two_six_data = tk.StringVar()
        label02 = tk.Label(root, text="26年业绩表：")
        label02.grid(row=1, column=0)
        entry02 = tk.Entry(root, textvariable=self.two_six_data, width=40)
        entry02.grid(row=1, column=1)
        btn02 = tk.Button(root, text="选择", command=lambda: self.selectPath(self.two_six_data))
        btn02.grid(row=1, column=2)

        self.product_details = tk.StringVar()
        label03 = tk.Label(root, text="2026产品明细：")
        label03.grid(row=2, column=0)
        entry03 = tk.Entry(root, textvariable=self.product_details, width=40)
        entry03.grid(row=2, column=1)
        btn03 = tk.Button(root, text="选择", command=lambda: self.selectPath(self.product_details))
        btn03.grid(row=2, column=2)

        self.customer_cor = tk.StringVar()
        label04 = tk.Label(root, text="客户对应关系表：")
        label04.grid(row=3, column=0)
        entry04 = tk.Entry(root, textvariable=self.customer_cor, width=40)
        entry04.grid(row=3, column=1)
        btn04 = tk.Button(root, text="选择", command=lambda: self.selectPath(self.customer_cor))
        btn04.grid(row=3, column=2)

        self.data_requirements = tk.StringVar()
        label05 = tk.Label(root, text="数据需求表：")
        label05.grid(row=4, column=0)
        entry05 = tk.Entry(root, textvariable=self.data_requirements, width=40)
        entry05.grid(row=4, column=1)
        btn05 = tk.Button(root, text="选择", command=lambda: self.selectPath(self.data_requirements))
        btn05.grid(row=4, column=2)

        self.two_five_data = tk.StringVar()
        label01 = tk.Label(root, text="25年数据：")
        label01.grid(row=5, column=0)
        entry01 = tk.Entry(root, textvariable=self.two_five_data, width=40)
        entry01.grid(row=5, column=1)
        btn01 = tk.Button(root, text="选择", command=lambda: self.selectPath(self.two_five_data))
        btn01.grid(row=5, column=2)

        # 时间选择框
        self.start_time_var = tk.StringVar()
        self.end_time_var = tk.StringVar()
        label06 = tk.Label(root, text="导出时间范围：")
        label06.grid(row=6, column=0, padx=0)
        start_cal = DateEntry(
            root,
            textvariable=self.start_time_var,
            date_pattern='yyyy-mm-dd',
            locale='zh_CN',
            width=12,
            borderwidth=2
        )
        start_cal.grid(row=6, column=1, padx=(0, 1), sticky='w')
        tk.Label(root, text="-").grid(row=6, column=1, padx=(1, 1))
        end_cal = DateEntry(
            root,
            textvariable=self.end_time_var,
            date_pattern='yyyy-mm-dd',
            locale='zh_CN',
            width=12,
            borderwidth=2
        )
        end_cal.grid(row=6, column=1, padx=(1, 0), sticky='e')

        btn001 = tk.Button(root, text="导入25年数据", command=self.import_25_data_to_db)
        btn001.grid(row=8, column=0)

        btn002 = tk.Button(root, text="导出26年数据", command=self.export_26_year_data)
        btn002.grid(row=8, column=1)

        btn005 = tk.Button(root, text="匹配", command=self.start)
        btn005.grid(row=8, column=2)

        # 结果打印框
        self.text = tk.Text(selectbackground="red", insertbackground="blue", spacing2=10, bd=0)
        self.text.grid(row=9, column=0, columnspan=10)

        self.text.insert(tk.END, "25年数据导入一次即可，不用重复导入！\r\n")
        self.text.insert(tk.END, "26年数据增量导入，全量导入会导致执行时间过长\r\n")
        self.text.insert(tk.END, "2026产品明细、客户对应关系表、数据需求表 必须选择\r\n\r\n")

    def start(self):
        self.T = threading.Thread(target=self.data_handling)
        self.T.setDaemon(True)
        self.T.start()

    def selectPath(self, path):
        path_ = filedialog.askopenfilename()
        path.set(path_)

    def data_handling(self):
        try:
            self.text.insert(tk.END, "开始...\r\n")
            two_six_path = self.two_six_data.get()
            product_details_path = self.product_details.get()
            customer_cor_path = self.customer_cor.get()
            data_requirements_path = self.data_requirements.get()
            if not product_details_path or not customer_cor_path or not data_requirements_path:
                self.text.insert(tk.END, "文件路径不能为空！\r\n")
                return
            self.text.insert(tk.END, "检查客户对应关系表是否完整...\r\n")
            # 检查客户对应关系表是否完整
            flag = self.check_customer_cor(customer_cor_path)
            if not flag:
                return

            self.text.insert(tk.END, "数据解析中...\r\n")
            engine = self.connect_db()
            if not engine:
                return
            # 基础表 数据入库
            self.common_excel_to_db(engine, product_details_path, customer_cor_path)

            # 26年业绩表增加BI到BO列
            if two_six_path:
                self.text.insert(tk.END, "26年业绩表增加BI到BO列...\r\n")
                self.add_bi_to_bo(engine, two_six_path)

            # 25年同期时间
            start_date = '2025-01-01'
            max_date = self.get_max_date(engine)
            if not max_date:
                raise Exception("找不到26年最晚日期！")
            self.text.insert(tk.END, f"最晚的‘业绩形成时间’是：{max_date}\r\n")
            # 开始结果数据处理
            self.text.insert(tk.END, "数据第一部分...\r\n")
            result_one = result_table_processing.result_table_one(engine, start_date, max_date)
            self.text.insert(tk.END, "数据第二部分...\r\n")
            result_two = result_table_processing.result_table_two(engine)
            self.text.insert(tk.END, "数据第三部分...\r\n")
            result_three = result_table_processing.result_table_three(engine)
            self.text.insert(tk.END, "数据第四部分...\r\n")
            result_four = result_table_processing.result_table_four(engine, start_date, max_date)
            self.text.insert(tk.END, "数据第五部分...\r\n")
            result_five = result_table_processing.result_table_five(engine)
            self.text.insert(tk.END, "数据第六部分...\r\n")
            result_six = result_table_processing.result_table_six(engine, start_date, max_date)
            self.text.insert(tk.END, "数据第七部分...\r\n")
            result_seven = result_table_processing.result_table_seven(engine, start_date, max_date)
            self.text.insert(tk.END, "数据第八部分...\r\n")
            result_eight = result_table_processing.result_table_eight(engine)
            self.text.insert(tk.END, "数据第九部分...\r\n")
            result_nine = result_table_processing.result_table_nine(engine)
            self.text.insert(tk.END, "数据第十部分...\r\n")
            result_ten = result_table_processing.result_table_ten(engine)
            self.text.insert(tk.END, "数据第十一部分...\r\n")
            result_eleven = result_table_processing.result_table_eleven(engine, start_date, max_date)
            self.text.insert(tk.END, "结果表下载中...\r\n")
            self.generate_result_table(data_requirements_path, result_one, result_two, result_three, result_four, result_five, result_six, result_seven, result_eight, result_nine, result_ten, result_eleven)

            self.text.insert(tk.END, "处理完成！\r\n")
            engine.dispose()
        except BaseException as e:
            self.text.insert(tk.END, "发生错误！\r\n")
            self.text.insert(tk.END, e)

    # 配置数据库连接
    def connect_db(self):
        try:
            # 数据库配置
            DB_HOST = 'localhost'
            DB_PORT = 3306
            DB_USER = 'root'
            DB_PASS = '1234'
            DB_NAME = 'test_sync'

            # # 生产环境
            # DB_HOST = '10.126.64.28'
            # DB_PORT = 3306
            # DB_USER = 'root'
            # DB_PASS = 'root^#123'
            # # DB_PASS = quote_plus("Iwfecats1213@")
            # DB_NAME = 'huawei_cloud_rpa'

            # 创建数据库连接
            engine = create_engine(f'mysql+pymysql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}')
            return engine
        except Exception as e:
            self.text.insert(tk.END, f"数据库连接失败: {str(e)}\r\n")
            return False

    # 验证客户对应关系表是否完整
    def check_customer_cor(self, customer_cor_path):
        # 读取客户对应关系表，判断“客户名称”不为空的行，是否都有“销售员”和“区域”
        try:
            # 读取Excel文件
            df = pd.read_excel(customer_cor_path)

            # 筛选有客户名称但缺失销售员或区域的行
            missing_data = df[
                (df['客户名称'].notnull()) &
                (df[['销售员', '区域']].isnull().any(axis=1))
                ]

            if not missing_data.empty:
                error_info = "\n".join([
                    f"第{row.Index + 2}行 {row['客户名称']}缺少: "
                    f"{'销售员' if pd.isnull(row['销售员']) else ''} "
                    f"{'区域' if pd.isnull(row['区域']) else ''}".strip()
                    for row in missing_data.itertuples()
                ])
                self.text.insert(tk.END, "客户对应关系表不完整！\r\n")
                self.text.insert(tk.END, error_info + "\r\n")
                return False
            return True
        except Exception as e:
            self.text.insert(tk.END, f"验证客户表时发生错误: {str(e)}\r\n")
            return False

    # 客户关系表和2025年产品明细表入库
    def common_excel_to_db(self, engine, product_details_path, customer_cor_path):
        # 先把客户关系表和2025年产品明细表入库
        try:
            # 读取Excel文件
            df_customer_cor = pd.read_excel(customer_cor_path).rename(columns={'客户名称': 'customer_name', '销售员': 'salesperson', '区域': 'region'})

            # 筛选出2025产品明细 需要入库的字段
            cloud_services = pd.read_excel(product_details_path, sheet_name='云服务名称')[['云服务编码', '服务产品部']].rename(columns={'云服务编码': 'cloud_services_code', '服务产品部': 'service_department'})
            details_flow = pd.read_excel(product_details_path, sheet_name='流量产品清单')[['产品类型编码', '产品类型']].rename(columns={'产品类型编码': 'product_code', '产品类型': 'product_type'})
            details_special = pd.read_excel(product_details_path, sheet_name='2025年产品专项', header=1)[['L4层产品编码', '名称']].rename(columns={'L4层产品编码': 'product_code', '名称': 'product_name'})
            details_collaborate = pd.read_excel(product_details_path, sheet_name='企业协同')[['云服务编码', '云服务名称']].rename(columns={'云服务编码': 'cloud_services_code', '云服务名称': 'cloud_services_name'})

            # 写入数据库
            with engine.begin() as conn:  # 自动提交/回滚
                # 阶段1：清空旧数据
                conn.execute(text("DELETE FROM customer_correspondence"))
                conn.execute(text("DELETE FROM two_five_details_cloud_services"))
                conn.execute(text("DELETE FROM two_five_details_flow"))
                conn.execute(text("DELETE FROM two_five_details_special"))
                conn.execute(text("DELETE FROM two_five_details_collaborate"))

                # 阶段2：插入新数据.'append'表示在现有表中追加数据（保留原有数据）、False表示不写入索引、'multi'表示使用多行组合的批量插入语法、批量插入1000条
                df_customer_cor.to_sql(name='customer_correspondence', con=conn, if_exists='append', index=False, method='multi', chunksize=1000)
                cloud_services.to_sql(name='two_five_details_cloud_services', con=conn, if_exists='append', index=False, method='multi', chunksize=1000)
                details_flow.to_sql(name='two_five_details_flow', con=conn, if_exists='append', index=False, method='multi', chunksize=1000)
                details_special.to_sql(name='two_five_details_special', con=conn, if_exists='append', index=False, method='multi', chunksize=1000)
                details_collaborate.to_sql(name='two_five_details_collaborate', con=conn, if_exists='append', index=False, method='multi', chunksize=1000)

        except Exception as e:
            self.text.insert(tk.END, f"写入表时发生错误: {str(e)}\r\n")
            return False

    # 数据入库方法
    def big_data_to_db(self, table, engine, df):
        # 数据库写入优化（使用批量UPSERT）
        chunk_size = 5000  # 根据测试调整最佳值
        columns = list(column_mapping.keys())
        db_columns = [column_mapping[col] for col in columns]

        # 准备SQL模板
        insert_sql = text(f"""
                               INSERT INTO {table} ({','.join(db_columns)})
                               VALUES ({','.join([':%s' % col for col in db_columns])})
                               ON DUPLICATE KEY UPDATE
                                   {','.join([f"{db_col}=VALUES({db_col})"
                                              for db_col in db_columns if db_col != 'performance_id'])}
                           """)

        with engine.begin() as conn:
            # 分块处理数据
            for i in range(0, len(df), chunk_size):
                chunk = df.iloc[i:i + chunk_size]
                # 转换为字典列表（内存优化）
                data = chunk.rename(columns=column_mapping) \
                    .replace({np.nan: None}) \
                    .to_dict('records')
                try:
                    # 批量执行
                    conn.execute(insert_sql, data)
                except Exception as e:
                    self.text.insert(tk.END, f"批量写入失败: {str(e)}\r\n")
                    raise e

    # 25年数据入库
    def import_25_data_to_db(self):
        self.text.insert(tk.END, f"开始读取25年数据...\r\n")
        two_five_path = self.two_five_data.get()
        if not two_five_path:
            self.text.insert(tk.END, "文件路径不能为空！\r\n")
            return
        try:
            engine = self.connect_db()
            if not engine:
                raise ConnectionError("数据库连接失败")

            with CalamineWorkbook.from_path(two_five_path) as wb:
                # 获取第一个工作表
                sheet = wb.get_sheet_by_index(0)
                # 读取所有数据（包含标题行）
                rows = sheet.to_python()
                selected_indices = [rows[0].index(col) for col in selected_columns]
                # 转换为DataFrame
                df = pd.DataFrame([[row[i] for i in selected_indices] for row in rows[1:]], columns=selected_columns)
                self.text.insert(tk.END, f"开始导入25年数据！\r\n")
                self.big_data_to_db('hw_two_five_data', engine, df)
                self.text.insert(tk.END, "数据导入成功！\r\n")
                del df
                # 获取第二个工作表
                sheet = wb.get_sheet_by_name('SMBcore')
                # 读取所有数据（包含标题行）
                rows = sheet.to_python()
                selected_indices = [rows[0].index(col) for col in selected_columns]
                # 转换为DataFrame
                df = pd.DataFrame([[row[i] for i in selected_indices] for row in rows[1:]], columns=selected_columns)
                self.text.insert(tk.END, f"开始导入25年SMBcore数据！\r\n")
                self.big_data_to_db('hw_two_five_data_smbcore', engine, df)
                self.text.insert(tk.END, "数据导入成功！\r\n")
                del df
                # 获取第三个工作表
                sheet = wb.get_sheet_by_name('NA')
                # 读取所有数据（包含标题行）
                rows = sheet.to_python()
                selected_indices = [rows[0].index(col) for col in selected_columns]
                # 转换为DataFrame
                df = pd.DataFrame([[row[i] for i in selected_indices] for row in rows[1:]], columns=selected_columns)
                self.text.insert(tk.END, f"开始导入25年NA数据！\r\n")
                self.big_data_to_db('hw_two_five_data_na', engine, df)
                self.text.insert(tk.END, "数据导入成功！\r\n")
                del df, rows
                gc.collect()
        except Exception as e:
            self.text.insert(tk.END, f"25年数据读取失败: {str(e)}\r\n")
        finally:
            engine.dispose()

    # 26年业绩表增加BI到BO列，含导出26年数据
    def add_bi_to_bo(self, engine, two_six_path):
        # 1. Excel读取优化 使用calamine引擎读取数据
        sel_columns_26 = ['业绩ID', '业绩金额(¥)', '业绩形成时间', '特殊返点类型', '二级经销商名称', '客户名称', '产品类型编码', '客户标签', '销售纵队']
        try:
            wb = CalamineWorkbook.from_path(two_six_path)
            # 获取第一个工作表
            sheet = wb.get_sheet_by_index(0)
            # 读取所有数据（包含标题行）
            rows = sheet.to_python()
            selected_indices = [rows[0].index(col) for col in sel_columns_26]
            # 转换为DataFrame
            two_six_df = pd.DataFrame([[row[i] for i in selected_indices] for row in rows[1:]], columns=sel_columns_26)
            wb.close()
        except Exception as e:
            raise ValueError(f"26年数据读取失败: {str(e)}")

        # 2. 数据库查询
        with engine.connect() as conn:
            sql_list = ['SELECT cloud_services_code, service_department FROM two_five_details_cloud_services;',
                        'SELECT product_code, product_type FROM two_five_details_flow',
                        'SELECT product_code, product_name FROM two_five_details_special',
                        'SELECT cloud_services_code, cloud_services_name FROM two_five_details_collaborate',
                        'SELECT customer_name, salesperson, region FROM customer_correspondence']
            dfs = [
                pd.read_sql_query(sql, conn, chunksize=None)
                for sql in sql_list
            ]

            # 解析查询结果
            cloud_services_map = dfs[0].set_index('cloud_services_code')['service_department']
            flow_products = set(dfs[1]['product_code'])
            special_products_map = dfs[2].set_index('product_code')['product_name']
            collaborate_map = dfs[3].set_index('cloud_services_code')['cloud_services_name']
            customer_relations = dfs[4][['customer_name', 'salesperson', 'region']]

            two_six_df = two_six_df.merge(customer_relations, left_on='客户名称', right_on='customer_name',how='left')

        # 3. 数据加工
        try:
            # 服务产品部映射
            two_six_df['服务产品部'] = two_six_df['产品类型编码'].map(cloud_services_map).fillna('')
            # 流量产品标记
            two_six_df['是否流量型产品'] = np.where(two_six_df['产品类型编码'].isin(flow_products), '是', '否')
            # 专线产品映射
            two_six_df['专线产品'] = two_six_df['产品类型编码'].map(special_products_map).fillna('')
            # 企业协同映射
            two_six_df['企业协同'] = two_six_df['产品类型编码'].map(collaborate_map).fillna('')
            # 客户信息映射
            two_six_df['销售员'] = two_six_df['salesperson'].fillna('').astype('category')
            two_six_df['区域'] = two_six_df['region'].fillna('').astype('category')
            # 清理临时列
            two_six_df.drop(['customer_name', 'salesperson', 'region'], axis=1, inplace=True)
            # 季度
            months = pd.to_datetime(two_six_df['业绩形成时间']).dt.month
            two_six_df['季度'] = 'Q' + ((months - 1) // 3 + 1).astype(str)
        except KeyError as e:
            raise ValueError(f"数据加工异常，缺少关键字段: {str(e)}")
        # 4. 数据入库
        try:
            self.text.insert(tk.END, f"26年数据入库中...\r\n")
            self.big_data_to_db('hw_two_six_data', engine, two_six_df)
        except Exception as e:
            self.text.insert(tk.END, f"26年数据入库失败: {str(e)}\r\n")
        # # 5. 导出25年的数据
        # self.text.insert(tk.END, "正在导出25年数据...\r\n")
        # # 直接把two_five_df中的数据导出到Excel文件
        # try:
        #     # 优化导出配置
        #     excel_path = two_five_path.replace(".xlsx", "_processed.xlsx")
        #     # 使用xlsxwriter引擎并启用优化参数
        #     with pd.ExcelWriter(
        #             excel_path,
        #             engine="xlsxwriter",
        #             engine_kwargs={
        #                 "options": {
        #                     "strings_to_urls": False,  # 禁用超链接检测
        #                     "constant_memory": True,  # 分段写入模式（核心优化）
        #                     "use_zip64": True,  # 支持超大文件
        #                 }
        #             },
        #     ) as writer:
        #         two_five_df.to_excel(
        #             writer,
        #             index=False,  # 不写入索引
        #             header=True,  # 保留标题行
        #             sheet_name="25年数据",
        #         )
        #     self.text.insert(tk.END, f"成功导出至：{excel_path}\r\n")
        # except Exception as e:
        #     self.text.insert(tk.END, f"导出失败: {str(e)}\r\n")

    # 找出25年同期时间
    def get_max_date(self, engine):
        sql_select = "SELECT DATE_SUB(MAX(performance_date), INTERVAL 1 YEAR) FROM hw_two_six_data"
        max_data = engine.connect().execute(text(sql_select)).fetchone()[0]
        return max_data

    # 生成结果表
    def generate_result_table(self, data_requirements_path, one, two, three, four, five, six, seven, eight, nine, ten, eleven):
        # 加载Excel文件
        wb = openpyxl.load_workbook(data_requirements_path)
        try:
            if '一' in wb.sheetnames:
                ws1 = wb['一']
                # Sheet1的填充逻辑
                row_mapping = {'北京': 3, '广州': 4, '深圳': 5, '上海': 6, '南京': 7, '成都': 8, '其他': 9, '总计': 10}
                category_columns = {'整体业绩': 2, 'NA业绩': 5, 'SMB业绩': 8, 'SMBcore业绩': 11}

                for category in ['整体业绩', 'NA业绩', 'SMB业绩', 'SMBcore业绩']:
                    if category not in one: continue
                    for region, values in one[category].items():
                        row = row_mapping.get(region)
                        if not row: continue
                        start_col = category_columns[category]
                        ws1.cell(row, start_col, float(values.get('渠道', 0)))
                        ws1.cell(row, start_col + 1, float(values.get('直客', 0)))
                        ws1.cell(row, start_col + 2, float(values.get('合计', 0)))

                # 处理同期增长率
                growth_data = one.get('同期增长率', {})
                for region, values in growth_data.items():
                    row = row_mapping.get(region)
                    if row:
                        ws1.cell(row, 14, values.get('整体业绩', ''))  # N列
                        ws1.cell(row, 15, values.get('NA业绩', ''))  # O列
                        ws1.cell(row, 16, values.get('SMB业绩', ''))  # P列
                        ws1.cell(row, 17, values.get('SMBcore业绩', ''))

                # 25年同期数据
                data_25 = one.get('25年同期数据', {})
                for region, values in data_25.items():
                    row = row_mapping.get(region)
                    if row:
                        ws1.cell(row, 18, values.get('整体业绩', ''))
                        ws1.cell(row, 19, values.get('NA业绩', ''))
                        ws1.cell(row, 20, values.get('SMB业绩', ''))
                        ws1.cell(row, 21, values.get('SMBcore业绩', ''))

            if '二' in wb.sheetnames:
                ws2 = wb['二']
                # Sheet2的列映射（字典字段 -> Excel列字母）
                column_map = {
                    '全量业绩': 'B',  # 全国业绩
                    '全量H1进度': 'D',  # 全国H2进度
                    '全量全年进度': 'F',  # 全国全年进度
                    'SMBH1进度': 'H',  # SMBH2进度
                    'SMB全年进度': 'J'  # SMB全年进度
                }

                # 遍历Sheet2每一行匹配区域
                for row in ws2.iter_rows(min_row=2):  # 从第2行开始
                    region = row[0].value  # A列为区域名
                    if region in two:
                        data = two[region]
                        # 填充每个字段
                        for field, col in column_map.items():
                            cell = ws2[f"{col}{row[0].row}"]
                            value = data.get(field)
                            cell.value = float(value) if isinstance(value, Decimal) else value

            if '三' in wb.sheetnames:
                ws3 = wb['三']
                # 列名映射
                column_map = {
                    '全量业绩': 'B',
                    '全量H1进度': 'D',
                    '全量全年进度': 'F',
                    'SMBH1进度': 'H',
                    'SMB全年进度': 'J'
                }

                # 遍历Sheet3每一行匹配销售姓名
                for row in ws3.iter_rows(min_row=2):  # 从第2行开始（跳过标题）
                    sales_name = row[0].value  # A列为销售姓名
                    if sales_name in three:
                        data = three[sales_name]
                        # 填充每个字段到对应列
                        for field, col in column_map.items():
                            cell = ws3[f"{col}{row[0].row}"]
                            value = data.get(field)
                            if isinstance(value, Decimal):
                                cell.value = float(value)
                            else:
                                cell.value = value

            if '四' in wb.sheetnames:
                ws4 = wb['四']
                # 列名映射（字典字段 -> Excel列字母）
                column_map = {
                    '1月': 'B', '2月': 'C', '3月': 'D', '4月': 'E', '5月': 'F', '6月': 'G', '7月': 'H', '8月': 'I',
                    '9月': 'J', '10月': 'K', '11月': 'L', '12月': 'M', '合计': 'N', '25年同期': 'O', '增长率': 'P'
                }

                # 填充SMBcore业绩
                # SMBcore业绩数据范围：标题在行2，数据从行3到行10，汇总在行10
                smb_start_row = 3
                smb_regions = ["北京", "广州", "深圳", "上海", "南京", "长春", "其他", "汇总"]
                smb_row_mapping = {region: smb_start_row + idx for idx, region in enumerate(smb_regions)}

                # 遍历SMBcore业绩数据
                smb_data = four.get('SMBcore业绩', {})
                for region, values in smb_data.items():
                    row = smb_row_mapping.get(region)
                    if not row:
                        continue
                    # 填充月份、合计、25年同期、增长率
                    for field, col in column_map.items():
                        cell = ws4[f"{col}{row}"]
                        value = values.get(field)
                        if isinstance(value, Decimal):
                            cell.value = float(value)
                        else:
                            cell.value = value

                # 填充NA业绩
                # NA业绩数据范围：标题在行13，数据从行14到行21，汇总在行21
                na_start_row = 14
                na_regions = ["北京", "广州", "深圳", "上海", "南京", "长春", "其他", "汇总"]
                na_row_mapping = {region: na_start_row + idx for idx, region in enumerate(na_regions)}

                # 遍历NA业绩数据
                na_data = four.get('NA业绩', {})
                for region, values in na_data.items():
                    row = na_row_mapping.get(region)
                    if not row:
                        continue
                    # 填充月份、合计、24年同期、增长率
                    for field, col in column_map.items():
                        cell = ws4[f"{col}{row}"]
                        value = values.get(field)
                        if isinstance(value, Decimal):
                            cell.value = float(value)
                        else:
                            cell.value = value

            if '五' in wb.sheetnames:
                ws5 = wb['五']
                # 列映射（字典字段 -> Excel列字母）
                column_map = {
                    '渠道': 'A', '客户': 'B', '销售员': 'C', '区域': 'D', '1月': 'E', '2月': 'F', '3月': 'G', '4月': 'H', '5月': 'I', '6月': 'J',
                    '7月': 'K', '8月': 'L', '9月': 'M', '10月': 'N', '11月': 'O', '12月': 'P', '合计': 'Q'
                }
                # 起始行（数据从第4行开始填充）
                current_row = 3
                # 遍历每一条数据
                for item in five:
                    # 填充渠道和客户
                    ws5[f"A{current_row}"] = item.get('渠道', '')
                    ws5[f"B{current_row}"] = item.get('客户', '')

                    # 填充月份和合计
                    for field, col in column_map.items():
                        if field in ['渠道', '客户']:
                            continue  # 已单独处理
                        value = item.get(field)
                        if isinstance(value, Decimal):
                            ws5[f"{col}{current_row}"] = float(value)
                        else:
                            ws5[f"{col}{current_row}"] = value
                    current_row += 1  # 移动到下一行

            if '六' in wb.sheetnames:
                ws6 = wb['六']
                # 列映射
                column_map = {
                    '26年截止目前业绩': 'C', '25年同期业绩': 'D', '同期增长率': 'E', '同比25年正负值': 'F'
                }

                # 动态遍历A列所有行
                for row in ws6.iter_rows(min_row=2):  # 从第2行开始遍历
                    company_name = row[0].value  # A列值
                    if company_name in six:  # 当公司名称存在于数据字典时
                        data = six[company_name]
                        # 填充对应列数据
                        for field, col in column_map.items():
                            cell = ws6[f"{col}{row[0].row}"]  # 使用当前行号
                            value = data.get(field)
                            if isinstance(value, Decimal):
                                cell.value = float(value)
                            else:
                                cell.value = value

            if '七' in wb.sheetnames:
                ws7 = wb['七']
                # 定义列映射
                column_map = {
                    '26Q1': 'F', '26Q2': 'G', '26Q3': 'H', '26Q4': 'I', '26年目前业绩': 'J', '25年同期业绩': 'K', '同比增长': 'L',
                }

                # 动态遍历A列所有有效行
                for row in ws7.iter_rows(min_row=2):  # 从第2行开始
                    product_cell = row[0]  # A列单元格
                    product_name = product_cell.value

                    # 严格匹配条件：产品名称存在且在数据字典中
                    if product_name and product_name in seven:
                        product_data = seven[product_name]

                        # 验证数据完整性
                        if all(key in product_data for key in column_map):
                            for field, col in column_map.items():
                                cell = ws7[f"{col}{product_cell.row}"]
                                value = product_data[field]
                                # 特殊处理数值类型
                                if field in ['26Q1', '26Q2', '26Q3', '26Q4', '26年目前业绩', '25年同期业绩']:
                                    cell.value = float(value) if isinstance(value, Decimal) else value
                                else:
                                    cell.value = str(value)

            if '八' in wb.sheetnames:
                ws8 = wb['八']
                # 动态遍历数据列表，填充到Sheet8
                for row_idx, data in enumerate(eight, start=2):  # 从第2行开始，假设第1行为标题
                    # 计算当前行号
                    current_row = row_idx
                    # 填充数据到对应的列
                    ws8.cell(row=current_row, column=1, value=data.get('新增渠道', ''))  # A列
                    ws8.cell(row=current_row, column=2, value=float(data.get('业绩金额', 0)))  # B列
                    ws8.cell(row=current_row, column=3, value=float(data.get('NA业绩', 0)))  # C列
                    ws8.cell(row=current_row, column=4, value=float(data.get('SMB业绩', 0)))  # D列
                    ws8.cell(row=current_row, column=5, value=float(data.get('SMB-CORE', 0)))  # E列
                    ws8.cell(row=current_row, column=6, value=data.get('销售员', ''))  # F列

            if '九' in wb.sheetnames:
                ws9 = wb['九']
                # 动态遍历数据列表，填充到Sheet9
                for row_idx, data in enumerate(nine, start=2):  # 从第2行开始
                    # 计算当前行号
                    current_row = row_idx
                    # 填充数据到对应的列
                    ws9.cell(row=current_row, column=1, value=data.get('新增客户', ''))  # A列
                    ws9.cell(row=current_row, column=2, value=data.get('渠道名称', ''))  # B列
                    ws9.cell(row=current_row, column=3, value=float(data.get('业绩金额', 0)))  # C列
                    ws9.cell(row=current_row, column=4, value=float(data.get('NA业绩', 0)))  # D列
                    ws9.cell(row=current_row, column=5, value=float(data.get('SMB业绩', 0)))  # E列
                    ws9.cell(row=current_row, column=6, value=float(data.get('SMB-CORE', 0)))  # F列
                    ws9.cell(row=current_row, column=7, value=data.get('销售员', ''))  # G列
                    ws9.cell(row=current_row, column=8, value=data.get('客户标签', ''))  # H列

            if '十' in wb.sheetnames:
                ws10 = wb['十']
                # 列映射（字典字段 -> Excel列字母）
                column_map = {
                    '渠道': 'A', '客户': 'B', '销售员': 'C', '区域': 'D', '1月': 'E', '2月': 'F', '3月': 'G', '4月': 'H', '5月': 'I', '6月': 'J',
                    '7月': 'K', '8月': 'L', '9月': 'M', '10月': 'N', '11月': 'O', '12月': 'P', '合计': 'Q'
                }
                # 起始行（数据从第4行开始填充）
                current_row = 3
                # 遍历每一条数据
                for item in ten:
                    # 填充渠道和客户
                    ws10[f"A{current_row}"] = item.get('渠道', '')
                    ws10[f"B{current_row}"] = item.get('客户', '')

                    # 填充月份和合计
                    for field, col in column_map.items():
                        if field in ['渠道', '客户']:
                            continue  # 已单独处理
                        value = item.get(field)
                        if isinstance(value, Decimal):
                            ws10[f"{col}{current_row}"] = float(value)
                        else:
                            ws10[f"{col}{current_row}"] = value
                    current_row += 1  # 移动到下一行

            if '十一' in wb.sheetnames:
                ws11 = wb['十一']
                # 列映射
                column_map = {
                    '26年截止目前业绩': 'C', '25年同期业绩': 'D', '同期增长率': 'E', '同比25年正负值': 'F'
                }

                # 动态遍历A列所有行
                for row in ws11.iter_rows(min_row=2):  # 从第2行开始遍历
                    company_name = row[0].value  # A列值
                    if company_name in eleven:  # 当公司名称存在于数据字典时
                        data = eleven[company_name]
                        # 填充对应列数据
                        for field, col in column_map.items():
                            cell = ws11[f"{col}{row[0].row}"]  # 使用当前行号
                            value = data.get(field)
                            if isinstance(value, Decimal):
                                cell.value = float(value)
                            else:
                                cell.value = value

            wb.save(f"{data_requirements_path.replace('.xlsx', '')}_result.xlsx")
            self.text.insert(tk.END, "结果表生成成功\r\n")
            return True
        except Exception as e:
            self.text.insert(tk.END, f"结果表生成失败：{e}\r\n")
            return False

    # 导出数据库中的26年数据
    def export_26_year_data(self):
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop_path, "26年数据_result.xlsx")
        try:
            self.text.insert(tk.END, "开始导出26年数据...\r\n")
            start_time = self.start_time_var.get().strip()
            end_time = self.end_time_var.get().strip()
            if not start_time or not end_time:
                self.text.insert(tk.END, "请填写时间范围\r\n")
                return False

            # 连接数据库
            engine = self.connect_db()
            if not engine:
                raise ConnectionError("数据库连接失败")
            # 获取当前工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "26年数据"

            # 定义每次读取的数据批量大小
            batch_size = 100000

            # 写入表头
            ws.append(selected_columns)

            # 创建数据库会话
            Session = sessionmaker(bind=engine)
            session = Session()

            # 获取总数据量
            total_count = session.execute(text(f"SELECT COUNT(*) FROM hw_two_six_data WHERE performance_date BETWEEN '{start_time}' AND '{end_time}'")).scalar()
            self.text.insert(tk.END, f"总数据量：{total_count} 条\r\n")
            current_row = 2
            if total_count > 0:
                # 分批读取并写入 Excel
                for offset in range(0, total_count, batch_size):
                    # 构建分页查询
                    batch_query = text(f"SELECT * FROM hw_two_six_data WHERE performance_date BETWEEN '{start_time}' AND '{end_time}' LIMIT {batch_size} OFFSET {offset}")
                    batch_data = [dict(row) for row in session.execute(batch_query).mappings().fetchall()]

                    # 将数据批量写入 Excel
                    for row in batch_data:
                        row_data = [row.get(col) for col in column_mapping.values()]
                        ws.append(row_data)
                        current_row += 1

                    # 定期保存以避免内存占用过多
                    wb.save(file_path)

                # 最终保存文件
                wb.save(file_path)
                self.text.insert(tk.END, f"结果表生成成功:{file_path}\r\n")
            return True
        except Exception as e:
            self.text.insert(tk.END, f"结果表生成失败：{e}\r\n")
            return False
        finally:
            if hasattr(self, 'session'):
                self.session.expunge_all()  # 清除所有ORM对象缓存
                self.session.close()
                del self.session  # 删除会话引用

            if hasattr(self, 'engine'):
                self.engine.dispose()
                del self.engine

            # 强制内存回收
            gc.collect()



if __name__ == '__main__':
    root = tk.Tk()

    App(root)
    root.mainloop()
