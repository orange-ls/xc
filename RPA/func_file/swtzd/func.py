#!/usr/bin/python3
# -*- coding: UTF-8 -*-


import os
import re
import shutil
from datetime import datetime, timedelta

import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter


# 根据输入的文件根目录返回当日文件保存目录
def getSaveDir(rootDir):
    """
    :param rootDir：文件保存根目录，例E:\downDir
    :return: 返回根目录+当日日期的目录，如E:\downDir\2022年\6月\8日
    """
    today = datetime.now()
    year, month, day = today.year, today.month, today.day
    finialDir = os.path.join(rootDir, "%s年\\%s月\\%s日" % (year, month, day))
    if not os.path.exists(finialDir):
        os.makedirs(finialDir)

    # 防止文件过多，只保留大约3个月内的数据：
    today = datetime.today()
    limitDay = today + timedelta(days=-100)
    year, month = limitDay.year, limitDay.month
    for root, dirs, files in os.walk(rootDir):
        if root.endswith("月"):
            searchResult = re.search(".*\\\\(\d+)年\\\\(\d+)月", root)
            if int(searchResult.group(1)) < year or (
                    int(searchResult.group(1)) == year and int(searchResult.group(2)) < month):
                shutil.rmtree(root, ignore_errors=True)

    return finialDir


# 调用xlwing插入大量行，且需要保持单元格格式一致
def excelInserRow(ws, startRow, totalNum):
    """
    :param ws: 需要处理的sheet
    :param startRow: 需要复制的行所在行号
    :param totalNum: 总共需要的行数
    :return: 
    """
    # 初始行为1行开始复制
    baseNum = 1
    calRow = startRow
    for i in range(totalNum):
        if totalNum == 1:
            break
        ws.range(f"{startRow}:{calRow}").insert(copy_origin="format_from_right_or_below")
        baseNum = baseNum * 2
        if totalNum - baseNum <= 0:
            break
        elif 0 < totalNum - baseNum < baseNum:
            calRow = totalNum - baseNum + startRow - 1
        else:
            calRow = baseNum + startRow - 1


# 获取服务对应的厂商物料编码
def getFirmCode(df_, rate, serviceName):
    df_ = df_.loc[(df_["服务名称"] == serviceName.strip()) & (df_["税率"] == rate)]
    if df_.empty:
        return ""
    else:
        code = df_.iloc[0]["厂商物料编码"]
        return code


# 重新设置“服务开票名称参考”字段
def resetServiceField(series):
    # 如果"独立软件/年费"为年费，则不作为“服务”数据
    if series["独立软件/年费"] == "年费":
        return None
    else:
        # 如果有“设备/服务”列为服务，但“服务开票名称参考”列无数据，则返回默认的服务名称 unNamedService
        if series["设备/服务"] == "服务" and pd.isna(series["服务开票名称参考"]):
            return unNamedService
        else:
            return series["服务开票名称参考"]


# 获取并写入通知单数据
def getNoticeFileContent(noticeFile, templatePath, saveDir, relationPath):
    """
    :param noticeFile: 商务通知单文件路径
    :param templatePath: 合同清单模板路径
    :param saveDir: 结果文件保存的文件夹
    :param relationPath: 服务物料对应关系表路径
    :return: 合同清单结果文件
    """
    # 1.读取通知单文件基本信息，获取数据字典
    dfdict = pd.read_excel(noticeFile, dtype=str, sheet_name=None)
    baseDf = dfdict["商务通知单(进供货)基本信息"].fillna("").applymap(lambda x: x.strip())
    columns = baseDf.columns
    dictDf1 = baseDf.drop_duplicates(columns[2], keep="first", ignore_index=True)
    dictDf2 = baseDf.drop_duplicates(columns[6], keep="first", ignore_index=True)
    msgDict = {}
    msgDict.update(dict(zip(dictDf1[columns[2]], dictDf1[columns[3]])))
    msgDict.update(dict(zip(dictDf2[columns[6]], dictDf2[columns[7]])))

    # 2.获取需要写入的基础数据
    if "变更后配置明细-物料" in dfdict.keys():  # 商务通知单已变更
        if "合同取消" in msgDict["变更类型："] or "基本信息变更" in msgDict["变更类型："]:
            return None
        df_default = dfdict["变更后配置明细-物料"]
        df_use = df_default
        # 退货或退换货且有退货赔偿金的情况，需要同步处理sheet《变更后配置明细-累计退货》
        if "退货" in msgDict["变更类型："] or ("退换货" in msgDict["变更类型："] and float(msgDict["累计退货赔偿金额："]) != 0):
            df_use = df_use.append(dfdict["变更后配置明细-累计退货"].rename(columns={"退货赔偿金额": "成交总价"}))
    else:  # 商务通知单未变更
        df_use = dfdict["配置明细-物料"]
    df_use["服务开票名称参考"] = df_use.apply(resetServiceField, axis=1)
    df_use["数量"] = pd.to_numeric(df_use["数量"])
    df_use["成交总价"] = pd.to_numeric(df_use["成交总价"])
    df_use = df_use.loc[(df_use["数量"] != 0) & (~df_use["数量"].isna()) & (df_use["成交总价"] != 0) & (~df_use["成交总价"].isna())]

    df_base = df_use.loc[df_use["服务开票名称参考"].isna()].copy()
    df_base["PO"] = msgDict['华为订单号：']
    inputArray = df_base[cols].values.tolist()

    # 3. 获取商务通知单"税率"和"服务物料对应关系表"数据
    rateDf = df_use.loc[~df_use["税率"].isna()]
    if rateDf.empty:
        rate = ""
    else:
        rate = rateDf.iloc[0]["税率"]
    df_relation = pd.read_excel(relationPath, dtype=str, header=1)
    df_relation["税率"] = pd.to_numeric(df_relation["税率"]).apply(lambda x: str(int(new_round(x) * 100)) + "%")
    df_relation["服务名称"] = df_relation["服务名称"].apply(lambda x: x.strip())
    # print(rate)

    # 4.对不同的“服务开票名称参考”进行分组，获取“附件”数据
    df_group = df_use.loc[df_use["成交总价"] != 0].groupby(["服务开票名称参考"])
    attachDict = {}
    idx = 0
    for key, df in df_group:
        df["序号"] = range(1, df.shape[0] + 1)
        usedData = df[["序号", "物料型号", "物料描述", "数量", "成交总价"]].values.tolist()
        # key_ = f"{key}（附件）" if idx ==0 else f"{key}（附件{idx}）"
        Total_transaction_price = df["成交总价"].sum()
        # 按“成交总价”的和来判断使用哪种公式计算单价和总价
        # 先计算出总价，再计算单价，保持总价和表格生成的一致
        for row in usedData:
            if Total_transaction_price > 1:
                price_server = round(row[4] / row[3] * (float(msgDict["出货指导价："]) / float(msgDict["进货总金额："])))
            else:
                price_server = round(row[4] / (1 - (float(msgDict["出货指导价："]) - float(msgDict["进货总金额："])) / float(msgDict["出货指导价："])) / row[3])
            total_price = price_server * row[3]
            # 去掉"成交总价" 再把单价和总价加入列表
            usedData[usedData.index(row)] = row[:4] + [price_server, total_price]

        attachDict[key] = usedData
        code = getFirmCode(df_relation, rate, key)
        inputArray.append([msgDict['华为订单号：'], code, key, key, 1, df["成交总价"].sum()])
        idx += 1
    dataNum = len(inputArray)
    if dataNum == 0:
        return None
    # else:
    #     if unNamedService in attachDict.keys():
    #         extraData = attachDict.pop(unNamedService)
    #         attachDict[unNamedService] = extraData

    # 5. 将合同清单模板复制到结果路径下并重命名
    if "变更后配置明细-物料" in dfdict.keys():
        # finalPath = os.path.join(saveDir, f"【{msgDict['华为订单号：']}】-{msgDict['变更类型：']}.xlsx")
        finalPath = os.path.join(saveDir, f"【{msgDict['华为订单号：']}】-合同清单.xlsx")
    else:
        finalPath = os.path.join(saveDir, f"【{msgDict['华为订单号：']}】-合同清单.xlsx")
    os.popen(f'copy "{templatePath}" "{finalPath}"')

    # 5. 打开结果文件，写入数据
    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finalPath)
    insertRowFlag = False  # 记录是否进行拆行
    for sheetName in ["合同清单表【模板】", "合同清单-word"]:
        ws = wb.sheets[sheetName]
        if sheetName in ["合同清单表【模板】"]:
            ws.range("Q3").value = msgDict["进货总金额："]
            ws.range("R3").value = msgDict["出货指导价："]
            ws.range("S3").value = msgDict["现金折扣:"]
            ws.range("T3").value = float(msgDict["运费:"]) * -1
            ws.range("B3").value = inputArray
            # Step1 对辅助列【成本单价】<1的数据，需要计算【成本总价 /（1-（评审出货总金额-评审进货总金额）/评审出货总金额）/数量】，生成销售单价，
            # 用 ROUND 函数 取值（保留2位小数），并以 数值形式 写入到【单价调整】对应的单元格处；
            for row in range(3, dataNum + 3):
                judCell = ws.range(f"N{row}").value
                if judCell < 1:
                    ws.range("Q1").formula = f"=ROUND(G{row}/(1-($R$3-$Q$3)/$R$3)/F{row},2)"
                    ws.range(f"L{row}").value = ws.range("Q1").value

            # Step2 定位销售金额【做合同】列金额最大的数据，计算【销售金额【做合同】-（差异值）】/数量
            # 用 TRUNC 函数 保留2位小数，并以 数值形式 写入到【销售单价【做合同】】对应的单元格处替换原销售单价
            saleAmountList = ws.range(f"I3:I{dataNum + 2}").value
            saleAmountList = [saleAmountList] if dataNum == 1 else saleAmountList

            maxRow = saleAmountList.index(max(saleAmountList)) + 3  # 最大【销售金额【做合同】】值所在的行
            ws.range("Q1").formula = f"=TRUNC((I{maxRow}-$R$5)/F{maxRow},2)"
            ws.range(f"H{maxRow}").value = ws.range("Q1").value
            ws.range("Q1").value = None  # 计算结束后，将辅助计算的Q1单元格置空

            # Step3 判断“销售单价【做合同】”、“毛利率”、“销售价核对”列是否有异常，有异常需要进行提示
            univalenceList = ws.range(f"H3:H{dataNum + 2}").value
            rateOfMarginList = ws.range(f"J3:J{dataNum + 2}").value
            salePriceCheckList = ws.range(f"M3:M{dataNum + 2}").value
            if dataNum == 1:
                univalenceList = [univalenceList]
                rateOfMarginList = [rateOfMarginList]
                salePriceCheckList = [salePriceCheckList]
            # 毛利率的值可能会有"-"，需要去除
            rateOfMarginList = [eve for eve in rateOfMarginList if eve != "-"]

            noticKey = ""
            if min(univalenceList) < 0:
                # “销售单价【做合同】”列有<0的值需要进行提示
                noticKey += "销售单价<0"
            if min(rateOfMarginList) < 0:
                # “销毛利率”列有<0的值需要进行提示
                if noticKey == "":
                    noticKey += "毛利率<0"
                else:
                    noticKey += "、毛利率<0"
            if "销售价异常" in salePriceCheckList:
                # “销售价核对”列有"销售价异常"时需要进行提示
                if noticKey == "":
                    noticKey += "销售价异常"
                else:
                    noticKey += "、销售价异常"
            if noticKey != "":
                noticText = f"存在{noticKey}的数据，请人工核对"
                ws.range("D1").value = noticText

        elif sheetName in ["合同清单-word"]:
            insertRow = dataNum + 1 if insertRowFlag else dataNum
            ws.range("B3").value = f"【{msgDict['华为订单号：']}】{msgDict['项目名称：']}"
            excelInserRow(ws, 5, insertRow)
            lastRow = insertRow + 4
            ws.range("B5:M5").copy(destination=ws.range(f"B5:M{lastRow}"))
            # 写入服务（附件）数据
            startCol = ws.range("P4").column

            # 找出“信息技术服务”和“维护保养服务”的“销售单价【做合同】”
            salePriceList = {}
            for row in range(5, dataNum + 5):
                if "服务" in ws.range(f"D{row}").value:
                    salePriceList.update({ws.range(f"D{row}").value: ws.range(f"G{row}").value})

            # 将“信息技术服务”和“维护保养服务”的详细数据根据总价进行处理
            for key, value in attachDict.items():
                sum_total_price = sum(va[5] for va in value)
                target_total_price = salePriceList.get(key)
                if sum_total_price and target_total_price:
                    difference = round(target_total_price - sum_total_price)
                else:
                    difference = 0
                # 如果差异值不为0，则将差异值添加到第一个数量为1的行
                flag = False
                if difference != 0:
                    for va in value:
                        if va[3] == 1:
                            va[4] = round(va[4] + difference)
                            va[5] = va[4] * va[3]
                            flag = True
                            break
                    # 如果没有数量为1的行，就将最后一行拆分一个出来加上差异值
                    if not flag:
                        value[-1][3] = value[-1][3] - 1
                        value[-1][5] = value[-1][3] * value[-1][4]
                        value.append([value[-1][0]+1, value[-1][1], value[-1][2], 1, value[-1][4]+difference, value[-1][4]+difference])

            idx = 0
            for key, writeData in attachDict.items():
                operateLetter_s = get_column_letter(startCol + idx * 7)
                operateLetter_e = get_column_letter(startCol + idx * 7 + 6)
                wb.sheets["附件"].range("A1:F3").copy(destination=ws.range(f"{operateLetter_s}4"))
                ws.range(f"{operateLetter_s}6:{operateLetter_e}6").copy(
                    destination=ws.range(f"{operateLetter_s}6:{operateLetter_e}{len(writeData) + 5}"))
                ws.range(f"{operateLetter_s}4").value = key
                ws.range(f"{operateLetter_s}6").value = writeData
                idx += 1
    else:
        wb.sheets["附件"].delete()
    wb.save(finalPath)
    wb.close()
    app.quit()

    # 商务通知单源文件复制到结果文件夹下，用于备份
    bakPath = os.path.join(saveDir, f"【{msgDict['华为订单号：']}】-商务通知单.xlsm")
    os.popen(f'copy "{noticeFile}" "{bakPath}"')

    return finalPath


# 四舍五入
def new_round(_float, _len=2):
    """
    :param _float: 需要四舍五入的数
    :param _len: 保留小数点位数
    :return: 四舍五入结果
    """
    if isinstance(_float, float):
        if str(_float)[::-1].find('.') <= _len:
            return _float
        if str(_float)[-1] == '5':
            return round(float(str(_float)[:-1] + '6'), _len)
        else:
            return round(_float, _len)
    else:
        return round(_float, _len)


# 对数值类型的数据模拟计算excel的trunc函数，保留两位小数（trunc：直接截取，不进行四舍五入）
def calTrunc(x, len_=2):
    """
    :param x: 传入的金额
    :param len_: 需要保留的小数点位数
    :return:
    """
    value = str(x)
    idx = value.find(".")
    if idx == -1:
        idx = len(value) + 1
    else:
        idx = idx + len_ + 1
    value = float(value[:idx])
    return value


"""
unNamedService: 商务通知单中【设备/服务】列 = 服务，但【服务开票名称参考】列为空时默认的"服务开票名称参考"
cols:结果表需要的列
"""
unNamedService = "其他服务"
cols = ["PO", "物料编码", "物料型号", "物料描述", "数量", "成交总价"]

if __name__ == "__main__":
    getNoticeFileContent(
        r"D:\xc_files\商务通知单\1Y05102306180K.xlsm",
        r"D:\xc_files\商务通知单\【模板】- 合同清单 - V4.0.xlsx",
        r"D:\xc_files\商务通知单\result",
        r"D:\xc_files\商务通知单\服务物料【厂商物料编码】对应关系表.xlsx")

    # import glob
    #
    # filePath = glob.glob(r"E:\My_Project\神州数码\Uibot项目\华为商务通知单需求资料\商务通知单转发流程附加功能需求资料\最新模板和规则\商务通知单原文件\*")
    # for path in filePath:
    #     print(path)
    #     result = getNoticeFileContent(path,
    #                                   r"E:\My_Project\神州数码\Uibot项目\华为商务通知单需求资料\商务通知单转发流程附加功能需求资料\最新模板和规则\【模板】- 合同清单 - V4.0.xlsx",
    #                                   r"E:\My_Project\神州数码\Uibot项目\华为商务通知单需求资料\商务通知单转发流程附加功能需求资料\最新模板和规则\商务通知单结果文件",
    #                                   r"E:\My_Project\神州数码\Uibot项目\华为商务通知单需求资料\商务通知单转发流程附加功能需求资料\服务物料【厂商物料编码】对应关系表.xlsx")
    #     print(result)
    #     print("*" * 50)
