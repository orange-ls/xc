#!/usr/bin/python3
# -*- coding: UTF-8 -*-


import glob
import logging.config
import os
import re
from datetime import datetime, timedelta
from functools import wraps

import numpy as np
import openpyxl
import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter


# 初始化日志
def initWriteLog(rootDir):
    """
    :param rootDir: 日志写入目录
    :return:
    """
    global logger
    log_name = "销售明细"
    simple_format = '[%(levelname)s][%(asctime)s]%(message)s'
    logfile_dir = rootDir  # log文件的目录
    logfile_name = '%s.log' % log_name  # log文件名
    # 如果不存在定义的日志目录就创建一个
    if not os.path.exists(logfile_dir):
        os.makedirs(logfile_dir)

    # log文件的全路径
    logfile_path = os.path.join(logfile_dir, logfile_name)
    LOGGING_DIC = {
        'version': 1,
        'disable_existing_loggers': False,
        'formatters': {
            'simple': {
                'format': simple_format
            },
        },
        'filters': {},
        'handlers': {
            'default': {
                'level': 'INFO',
                'class': 'logging.handlers.RotatingFileHandler',
                'formatter': 'simple',
                'filename': logfile_path,
                'maxBytes': 1024 * 1024 * 5,
                'backupCount': 5,
                'encoding': 'utf-8',
            },
        },
        'loggers': {
            '': {
                'handlers': ['default'],
                'level': 'INFO',
                'propagate': True,
            },
        },
    }
    logging.config.dictConfig(LOGGING_DIC)
    logger = logging.getLogger(__name__)


# 定义一个日志装饰器，每次调用方法前后打印日志
# 注 uibot中无法调用，弃用
def logfun(func):
    @wraps(func)
    def logfunStep(*args, **kwargs):
        logger.info(f"{'-' * 25}{func.__name__}开始{'-' * 25}")
        result = func(*args, **kwargs)
        logger.info(f"{'-' * 25}{func.__name__}结束{'-' * 25}")
        return result

    return logfunStep


# 将MHTML文件转为xlsx
def changeMhtmlToXlsx(path):
    """
    :param path: MHTML文件路径
    :return: 转换后的xlsx文件路径
    """
    outPath = os.path.join(os.path.dirname(path), os.path.basename(path).replace(".MHTML", ".xlsx"))
    if not os.path.exists(outPath):
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = True
        wb = app.books.open(path)
        # ws = wb.sheets["Sheet1"]
        # ws.autofit()
        wb.save(outPath)
        # 关闭工作簿
        wb.close()
        app.quit()

        # 删除MHTML文件
        # os.remove(path)
    return outPath


# 读取配置文件生成配置字典
def getConfigDict(baseConfPath):
    """
    :param baseConfPath：配置文件目录
    :return: 生成配置字典
    """
    df = pd.read_excel(baseConfPath, dtype=str, sheet_name="Sheet1").fillna("")
    resultDict = dict(zip(df["配置名称"], df["配置内容"]))
    # 华为账户密码有多个，处理为列表格式
    for key in ["华为网站账号", "华为网站密码", "销售明细排除的销售员"]:
        resultDict[key] = resultDict[key].replace("；", ";").strip(";").split(";")
    return resultDict


# 根据输入的文件根目录返回当日文件保存目录
def getSaveDir(rootDir):
    """
    :param rootDir：文件保存根目录，例E:\downDir
    :return: 返回根目录+当日日期的目录，如E:\downDir\2022年\6月\8日
    """
    today = datetime.now()
    year, month, day = today.year, today.month, today.day
    finialDir = os.path.join(rootDir, "%s年\\%s月\\%s日" % (year, month, day))
    return finialDir


# 四舍五入
def new_round(_float, _len=2):
    """
    :param _float: 需要四舍五入的数
    :param _len: 保留小数点位数
    :return: 四舍五入结果
    """
    if isinstance(_float, float):
        if str(_float)[::-1].find('.') <= _len:
            return _float
        if str(_float)[-1] == '5':
            return round(float(str(_float)[:-1] + '6'), _len)
        else:
            return round(_float, _len)
    else:
        return round(_float, _len)


# 判断值是否在字符串、字典、列表中
def jud_in(val, targetObj):
    return val in targetObj


# 获取"欠款明细表"和"销售明细表"的操作菜单
def getOperateDate(sTime, eTime):
    """
    :param sTime: 用户输入的开始日期
    :param eTime: 用户输入的结束日期
    :return: {"欠款明细": [{"菜单"：菜单, "标题关键字":标题关键字}], "销售明细": [{"菜单"：菜单, "标题关键字":标题关键字}]}
    """
    fmt = "%Y-%m-%d"
    qkmxList = []
    for tStr in [sTime, eTime]:
        qkmxList.append({"菜单": f"{tStr[:4]}年{int(tStr[5:7])}月", "标题关键字": f"{tStr[5:7]}{tStr[8:]}"})

    xsmxList = []
    sTime = datetime.strptime(sTime, fmt) + timedelta(days=1)
    timeRange = pd.period_range(sTime, eTime, freq='M')
    timeList = [t_.strftime("%Y-%m") for t_ in timeRange]
    for tStr in timeList:
        xsmxList.append({"菜单": f"{tStr[:4]}年", "标题关键字": f"{int(tStr[5:])}月"})

    return {"欠款明细": qkmxList, "销售明细": xsmxList}


# 获取"销售明细表"中需要操作的日期
def getValidDateList(sTime, eTime):
    """
    :param sTime: 用户输入的开始日期
    :param eTime: 用户输入的结束日期
    :return: 销售明细表需要操作的日期列表
    """
    dateRange = pd.period_range(sTime, eTime, freq='D')
    dateList = [t_.strftime("%Y-%m-%d") for t_ in dateRange]
    return dateList[1:]


# 获取"回款明细表"的下载菜单
def getReceivableOperateDate(saveDir, searchTIme):
    """
    :param searchTIme: 用户输入的“未清项目明细表”处理日期，如2022-07-18
    :return: 返回回款明细的下载目录（菜单、标题关键词）
    """

    """
    回款明细汇总表：文件名:回款明细_20220728.xlsx
    1.日期表示文件保存的截止日期，回款明细表会下载月数据，当月数据会重新下载
    2.下载时的日期即为2022/07-当天所在年月
    """
    # 获取回款明细汇总表文件名及日期
    fileList = glob.glob(f"{saveDir}\\回款明细_*.xlsx")
    filePath = fileList[0]
    matchObj = re.search("回款明细_(\d{8}).xlsx", os.path.basename(filePath))

    # 计算文件日期和本次操作日期的差值判断文件是否为最新
    lastDate = datetime.strptime(matchObj.group(1), "%Y%m%d")
    dayDelta = (datetime.strptime(searchTIme, "%Y-%m-%d") - lastDate).days
    if dayDelta <= 0:
        lastFlag = True
    else:
        lastFlag = False

    # 获取下载的年月列表并生成下载的菜单
    pi = pd.period_range(lastDate, searchTIme, freq="M")
    totalDownloadList = [str(i) for i in pi]
    xsmxList = []
    for tStr in totalDownloadList:
        menu = f"{tStr[:4]}年"
        keyWord = f"回款明细-{tStr[5:7]}"
        xsmxList.append({"菜单": menu, "标题关键字": keyWord})

    # return 回款明细汇总表路径，回款明细表下载菜单， 文件最新标识
    return filePath, xsmxList, lastFlag


# 合并回款明细表
# @logfun
def updateHKMXFile(addfilePathList, finalPath, searchTIme):
    """
    :param addfilePathList: 下载的回款明细表路径列表
    :param finalPath: 回款明细汇总表路径
    :param searchTIme: 用户输入的“未清项目明细表”处理日期，如2022-07-18
    :return: 返回回款明细汇总表路径
    """
    # 读取下载的回款明细表并汇总到回款明细汇总表中
    hkmxDf = pd.DataFrame()
    for path_ in [finalPath] + addfilePathList:
        df_temp = pd.read_excel(path_, sheet_name="回款明细", dtype=str)
        if not df_temp.empty:
            oldColumns = df_temp.columns
            newColumns = [re.sub("[(（].*[)）]", "", col) for col in oldColumns]
            df_temp.columns = newColumns
            hkmxDf = hkmxDf.append(df_temp[receivableTableCol])
    # 按日期排序后去除重复项，保留最新【输入日期】的数据，删除2年前的数据
    exportDate = datetime.strptime(searchTIme, "%Y-%m-%d")
    minDate = datetime(year=exportDate.year - 2, month=exportDate.month, day=exportDate.day).strftime("%Y-%m-%d")
    hkmxDf = hkmxDf[(hkmxDf["输入日期"].str[:10] >= minDate) & (~hkmxDf["销售员"].isna())]
    hkmxDf.sort_values(by=["输入日期"], ascending=[False], inplace=True)
    hkmxDf["财务凭证号FI"] = hkmxDf["财务凭证号FI"].str.zfill(10)
    hkmxDf["客户代码"] = hkmxDf["客户代码"].str.zfill(10)
    hkmxDf.drop_duplicates(["财务凭证号FI", "客户代码"], keep="first", inplace=True, ignore_index=True)

    app = xw.App(visible=True, add_book=False)
    app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finalPath)

    ws = wb.sheets["回款明细"]
    # 表中新增”回款明细“+"当前汇总表日期"的sheet，用于备份之前的数据
    sheetNames = [sheet.name for sheet in wb.sheets]
    bakSheetName = "回款明细" + finalPath[-13:-5]
    if bakSheetName not in sheetNames:
        wb.sheets.add(bakSheetName, after=sheetNames[0])
        ws_ = wb.sheets[bakSheetName]
        ws.used_range.copy(destination=ws_.range("A1"))

    # 删除多余的sheet页
    for sheetName in sheetNames:
        if sheetName not in ["回款明细", bakSheetName]:
            wb.sheets[sheetName].delete()

    ws.activate()
    ws.clear()

    colIndex = get_column_letter(len(receivableTableCol))
    # 全表格式改成常规格式、内容字体为 Arial 10号、表头加粗，字体为 Arial 11号，设置背景色
    ws.range(f"A:{colIndex}").number_format = "G/通用格式"
    ws.range(f"A:{colIndex}").font.size = 10
    ws.range(f"A:{colIndex}").font.name = "Arial"
    ws.range(f"A1:{colIndex}1").font.size = 11
    ws.range(f"A1:{colIndex}1").font.bold = True
    ws.range(f"A1:{colIndex}1").color = (218, 238, 243)

    # "财务凭证号FI"、"客户代码"、"销售员代码"列为文本格式
    for col in ["财务凭证号FI", "客户代码", "销售员代码"]:
        colLetter = get_column_letter(receivableTableCol.index(col) + 1)
        ws.range(f"{colLetter}:{colLetter}").number_format = "@"

    # 各日期列设置为只显示年月日
    for col in ["记帐日期", "输入日期"]:
        colLetter = get_column_letter(receivableTableCol.index(col) + 1)
        ws.range(f"{colLetter}:{colLetter}").number_format = "yyyy/m/d"
    # 金额列设置为千分位显示两位小数

    for col in ["利润中心本位币金额"]:
        colLetter = get_column_letter(receivableTableCol.index(col) + 1)
        ws.range(f"{colLetter}:{colLetter}").number_format = "#,##0.00"

    # 设置边框
    for borderIdx in [7, 8, 9, 10, 11, 12]:
        """
        7: 左边框
        8: 上边框
        9: 下边框
        10: 右边框
        11: 区域单元格垂直边框
        12: 区域单元格水平边框
        """
        if borderIdx in [7, 9, 10, 11, 12]:
            ws.range(f"A2:{colIndex}{hkmxDf.shape[0] + 1}").api.Borders(borderIdx).LineStyle = 1  # 实线
            ws.range(f"A2:{colIndex}{hkmxDf.shape[0] + 1}").api.Borders(borderIdx).ColorIndex = 3  # 红色
        if borderIdx in [7, 8, 9, 10, 11]:
            ws.range(f"A1:{colIndex}1").api.Borders(borderIdx).LineStyle = 1  # 实线

    # 重新载入数据以正确显示
    ws.range("A1").value = receivableTableCol
    ws.range("A2").value = hkmxDf.values

    # 进行冻结操作
    active_window = wb.app.api.ActiveWindow
    active_window.FreezePanes = False
    # wb.app.range("A2").select()  # 选"A2"冻结首行
    # active_window.SplitColumn = resultCol.index("客户名称") + 1  # 冻结至哪一列
    active_window.SplitRow = 1  # 冻结至哪一行
    active_window.FreezePanes = True

    # 设置列宽
    ws.used_range.column_width = 10

    # ws.autofit()
    wb.save(finalPath)
    # 关闭工作簿
    wb.close()
    app.quit()

    # 更新文件名
    newFileName = "回款明细_" + searchTIme.replace("-", "") + ".xlsx"
    newFilePath = os.path.join(os.path.dirname(finalPath), newFileName)
    os.rename(finalPath, newFilePath)

    return newFilePath


# 判断二维列表文本中是否包含某个字符串
def findLocInAllTitle(keyValue, titleList):
    """
    :param keyValue: 输入的字符串
    :param titleList: 输入的二维列表 例如 [["*a*"],["*b*"]]
    :return: 若不在二维列表中返回False，否则返回所在的index
    """
    for idx, arrList in enumerate(titleList):
        if keyValue in arrList[0]:
            return idx
    else:
        return False


# 匹配扣款时间
def matchDeductTime(inputTime):
    """
    :param inputTime: 输入的时间 "YYYY-mm-dd"
    :return: FY + 年份后两位 + -Q + 所在季度，例如 FY22-Q1
    """
    year_ = inputTime[2:4]
    month_ = int(inputTime[5:7])
    quarter_ = (month_ + 2) // 3
    return f"FY{year_}-Q{quarter_}"


# 初始化读取的欠款明细表，修改列名、筛选数据
def initDebtDf(filepath):
    """
    :param filepath: 输入欠款明细表路径
    :return: 处理后的df
    """
    df_ = pd.read_excel(filepath, dtype=str).fillna("")
    # df_中列名中有括号，（）或()的内容删除
    allCol = df_.columns
    allCol = [re.sub("[(（].*[)）]", "", col) for col in allCol]
    df_.columns = allCol
    # 筛选需要的列，修改列名，并将“欠款金额”列作为number
    df_ = df_[debtfilterCol]
    df_.rename(columns=debtRenameColDict, inplace=True)
    df_["欠款金额"] = pd.to_numeric(df_["欠款金额"])
    # 筛选业务范围代码
    df_ = df_[df_["业务范围代码"].str.contains(businessCodeStr)]
    # 生成新索引
    df_["索引temp"] = df_["销售单代码"] + df_["索引"] + df_["收付基准日期"]

    return df_


# 计算欠款明细表的账期费用
def caldebtCost(series, stime, etime, indexDict):
    """
    :param series: df的行数据
    :param sTime: 用户输入的开始日期
    :param eTime: 用户输入的结束日期
    :param indexDict: 计算公式时需要的列所在列标识字典，如 {"欠款金额":"M","贷款利率"："K"}
    :return: series[["账期财务费用", "超期财务费用", "备注"]]
    """
    fmt = "%Y-%m-%d"

    if pd.isna(series["实际回款日"]):  # 未回款
        # 判断【收付基准日期】与欠款明细表开始日期的差值
        jugtimeDiff1 = (datetime.strptime(series["收付基准日期"][:10], fmt) - datetime.strptime(stime, fmt)).days
        if jugtimeDiff1 > 0:
            # 判断【收付基准日期】与欠款明细表结束日期的差值
            jugtimeDiff2 = (datetime.strptime(series["收付基准日期"][:10], fmt) - datetime.strptime(etime, fmt)).days
            if jugtimeDiff2 <= 0:
                if series["采购类型"] != "商业分销":
                    series[
                        "账期财务费用"] = '=(DAYS(INDIRECT("{应还款日期}"&ROW()),INDIRECT("{收付基准日期}"&ROW())) - 30) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
                else:
                    series[
                        "账期财务费用"] = '=(DAYS(INDIRECT("{应还款日期}"&ROW()),INDIRECT("{收付基准日期}"&ROW())) - 60) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
                # 判断【应还款日期】与【统计时间】差值（判断有无超期）
                jugtimeDiff3 = (
                        datetime.strptime(series["应还款日期"][:10], fmt) - datetime.strptime(series["统计时间"][:10],
                                                                                         fmt)).days
                if jugtimeDiff3 < 0:
                    series[
                        "超期财务费用"] = '=DAYS(INDIRECT("{统计时间}"&ROW()),INDIRECT("{应还款日期}"&ROW())) * INDIRECT("{超额贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
            else:  # 【收付基准日期】> 欠款明细表结束日期
                if series["采购类型"] != "商业分销":
                    series[
                        "账期财务费用"] = '=(DAYS(INDIRECT("{应还款日期}"&ROW()),INDIRECT("{统计时间}"&ROW())) - 30) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
                else:
                    series[
                        "账期财务费用"] = '=(DAYS(INDIRECT("{应还款日期}"&ROW()),INDIRECT("{统计时间}"&ROW())) - 60) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)

        else:
            if series["索引新增标识"] == "新增":
                series[
                    "账期财务费用"] = '=DAYS(INDIRECT("{应还款日期}"&ROW()),INDIRECT("{凭证记帐日期}"&ROW())) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                    **indexDict)
            else:  # 索引不是新增
                # 判断第二张欠款明细表日期与【应还款日期】差值
                jugtimeDiff3 = (datetime.strptime(etime, fmt) - datetime.strptime(series["应还款日期"][:10], fmt)).days
                if jugtimeDiff3 <= 0:
                    series["备注"] = "去除"
                else:
                    # 判断【应还款日期】与【上次统计时间】的差值
                    jugtimeDiff4 = (
                            datetime.strptime(series["应还款日期"][:10], fmt) - datetime.strptime(series["上次统计时间"][:10],
                                                                                             fmt)).days
                    if jugtimeDiff4 >= 0:
                        series[
                            "超期财务费用"] = '=DAYS(INDIRECT("{统计时间}"&ROW()),INDIRECT("{应还款日期}"&ROW())) * INDIRECT("{超额贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                            **indexDict)
                    else:
                        series[
                            "超期财务费用"] = '=DAYS(INDIRECT("{统计时间}"&ROW()),INDIRECT("{上次统计时间}"&ROW())) * INDIRECT("{超额贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                            **indexDict)
    else:
        # 已回款，判断【收付基准日期】与欠款明细表开始日期的差值
        jugtimeDiff1 = (datetime.strptime(series["收付基准日期"][:10], fmt) - datetime.strptime(stime, fmt)).days
        if jugtimeDiff1 > 0:
            # 判断【收付基准日期】与欠款明细表结束日期的差值
            jugtimeDiff2 = (datetime.strptime(series["收付基准日期"][:10], fmt) - datetime.strptime(etime, fmt)).days
            if jugtimeDiff2 <= 0:
                if series["采购类型"] != "商业分销":
                    series[
                        "账期财务费用"] = '=(DAYS(INDIRECT("{实际回款日}"&ROW()),INDIRECT("{收付基准日期}"&ROW())) - 30) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
                else:
                    series[
                        "账期财务费用"] = '=(DAYS(INDIRECT("{实际回款日}"&ROW()),INDIRECT("{收付基准日期}"&ROW())) - 60) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
            else:  # 【收付基准日期】> 欠款明细表结束日期
                # 判断【应还款日期】与【统计时间】差值
                jugtimeDiff3 = (
                        datetime.strptime(series["应还款日期"][:10], fmt) - datetime.strptime(series["统计时间"][:10], fmt)).days
                if jugtimeDiff3 >= 0:
                    series[
                        "账期财务费用"] = '=DAYS(INDIRECT("{实际回款日}"&ROW()),INDIRECT("{应还款日期}"&ROW())) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
                else:
                    series[
                        "超期财务费用"] = '=DAYS(INDIRECT("{实际回款日}"&ROW()),INDIRECT("{上次统计时间}"&ROW())) * INDIRECT("{超额贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)

        else:

            # 再判断第二张欠款明细表日期与【应还款日期】差值
            jugtimeDiff2 = (datetime.strptime(etime, fmt) - datetime.strptime(series["应还款日期"][:10], fmt)).days
            if jugtimeDiff2 <= 0:
                series[
                    "账期财务费用"] = '=DAYS(INDIRECT("{实际回款日}"&ROW()),INDIRECT("{应还款日期}"&ROW())) * INDIRECT("{正常贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                    **indexDict)

            else:
                # 超期回款，判断【上次统计时间】与【应还款日期】的差值
                jugtimeDiff3 = (datetime.strptime(series["上次统计时间"][:10], fmt) - datetime.strptime(series["应还款日期"][:10],
                                                                                                  fmt)).days
                if jugtimeDiff3 <= 0:
                    series[
                        "超期财务费用"] = '=DAYS(INDIRECT("{实际回款日}"&ROW()),INDIRECT("{应还款日期}"&ROW())) * INDIRECT("{超额贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)
                else:
                    series[
                        "超期财务费用"] = '=DAYS(INDIRECT("{实际回款日}"&ROW()),INDIRECT("{上次统计时间}"&ROW())) * INDIRECT("{超额贷款费率}"&ROW()) / 365 * INDIRECT("{欠款金额}"&ROW())'.format(
                        **indexDict)

    return series[["账期财务费用", "超期财务费用", "备注"]]


# 计算预收补费用sheet的预收利息
def calAdvanceCost(series, lastCalDate, indexDict):
    """
    :param series: df的行数据
    :param lastCalDate: 上次统计时间（基础表中“统计时间”列最新的日期）
    :param indexDict: 计算公式时需要的列所在列标识字典，如 {"欠款金额":"M","贷款利率"："K"}
    :return: "预收利息"
    """
    fmt = "%Y-%m-%d"
    # 如果是第一次统计，无上次统计时间
    if lastCalDate == "":
        jugtimeDiff = 1  # 任一大于0的数均可
    else:
        # 判断"过帐日期"和"上次统计时间"的日期差
        jugtimeDiff = (datetime.strptime(series["过帐日期"][:10], fmt) - datetime.strptime(lastCalDate, fmt)).days
    if jugtimeDiff >= 0:
        resultCost = '=DAYS(INDIRECT("{统计时间}"&ROW()),INDIRECT("{过帐日期}"&ROW())) * INDIRECT("{利率}"&ROW()) / 365 * INDIRECT("{本币金额}"&ROW())'.format(
            **indexDict)
    else:
        resultCost = '=DAYS(INDIRECT("{统计时间}"&ROW()),INDIRECT("{上次统计时间}"&ROW())) * INDIRECT("{利率}"&ROW()) / 365 * INDIRECT("{本币金额}"&ROW())'.format(
            **indexDict)

    return resultCost


# 获取账期费用基础表数据
def getBaseTableData(baseTablePath):
    """
    :param baseTablePath: 账期费用基础表路径
    :return: [{"欠款明细":欠款明细df, "银票":银票df, "预收补费用":预收补费用df},
                {"欠款明细": 最新的统计时间, "预收补费用":最新的统计时间}]
    """
    if baseTablePath != "":
        # 获取需要的df字典
        dfDict = pd.read_excel(baseTablePath, sheet_name=None, dtype=str)
        baseDfDict = {}
        for sheetName, df in dfDict.items():
            # df中列名可能前后有空格，进行删除
            allCol = df.columns
            allCol = [col.strip() for col in allCol]
            df.columns = allCol
            if "欠款明细" in sheetName:
                baseDfDict["欠款明细"] = df
            elif "银票" in sheetName:
                baseDfDict["银票"] = df
            elif "预收补费用" in sheetName:
                baseDfDict["预收补费用"] = df
            else:
                continue
    else:
        # 没有基础表则默认基础表数据为空
        baseDfDict = {"欠款明细": pd.DataFrame(columns=debtResultCol),
                      "银票": pd.DataFrame(columns=bankNotesResultCol),
                      "预收补费用": pd.DataFrame(columns=advanceResultCol)}

    # 获取"欠款明细"和"预收补费用"的最新统计时间
    lastCalTimeDict = {}
    for key in ["欠款明细", "预收补费用"]:
        df_temp = baseDfDict[key].sort_values(by="统计时间", ascending=False)
        if not df_temp.empty:
            lastCalTimeDict[key] = df_temp.iloc[0]["统计时间"][:10]
        else:
            # 如果为空表或者为第一次运行无基础数据的情况，默认“统计时间”=去年最后一天
            today = datetime.today()
            thisYearFirstDay = datetime(year=today.year, month=1, day=1)
            defaultDay = thisYearFirstDay + timedelta(days=-1)
            lastCalTimeDict[key] = defaultDay.strftime("%Y-%m-%d")
    return baseDfDict, lastCalTimeDict


# 操作《欠款明细》sheet, 匹配实际回款日、实际回款金额、扣款时间、月份、统计时间、上次统计时间
def debtSheetOperate(sTime, eTime, debtFileList, saleDetailPathList, lastCalTimeDict):
    """
    :param sTime: 用户输入的开始日期
    :param eTime: 用户输入的结束日期
    :param debtFileList: 下载的欠款明细表路径列表
    :param saleDetailPathList: 下载的销售明细表路径列表
    :param lastCalTimeDict: 统计时间字典{"欠款明细": 最新的统计时间, "预收补费用":最新的统计时间}
    :return: 欠款明细匹配数据后的df
    """
    # Step1: 欠款明细表匹配
    # 读取欠款明细表并做初始化处理
    df_s = initDebtDf(debtFileList[0])
    df_e = initDebtDf(debtFileList[1])

    addDebtDf = pd.DataFrame(columns=df_s.columns)
    allIndexList = set(df_s["索引temp"].tolist())
    for idx in allIndexList:
        df_match_s = df_s.loc[df_s["索引temp"] == idx].copy()
        df_match_e = df_e.loc[df_e["索引temp"] == idx]
        if df_match_e.empty:
            # 索引不在第二张欠款明细表中，说明已被核销
            df_match_s["实际回款日"] = eTime
            df_match_s["实际回款金额"] = df_match_s["欠款金额"]
            addDebtDf = addDebtDf.append(df_match_s)
        else:
            # 索引在第二张欠款明细表中，比较欠款金额，判断是否为部分核销
            diffAmount = new_round(df_match_s["欠款金额"].sum() - df_match_e["欠款金额"].sum())
            series_e = df_match_e.iloc[0].copy()  # 数据以第二张欠款明细表匹配出的任意一条数据为准
            if diffAmount > 0:
                series_e["实际回款日"] = eTime
                series_e["实际回款金额"] = diffAmount
                series_e["欠款金额"] = diffAmount
                addDebtDf = addDebtDf.append(series_e)
            else:
                # 无回款，不进行处理
                continue
    else:
        addDebtDf["扣款时间"] = matchDeductTime(eTime)
        addDebtDf["月份"] = f"{int(eTime[5:7])}月"

    # Step: 销售明细表匹配
    # 读取销售明细表
    df_detail = pd.DataFrame()
    for path in saleDetailPathList:
        df_temp = pd.read_excel(path, dtype=str)
        df_detail = df_detail.append(df_temp)
    # 筛选需要的列，修改列名，并将“欠款金额”列作为number
    df_detail = df_detail[saleDetailfilterCol]
    df_detail.rename(columns=saleDetailRenameColDict, inplace=True)
    df_detail["欠款金额"] = pd.to_numeric(df_detail["欠款金额"])
    df_detail["收付基准日期"] = df_detail["凭证记帐日期"]
    # 筛选产品组（业务范围代码）、出具发票日（凭证记帐日期）、折扣后合同金额（欠款金额）
    validDateList = getValidDateList(sTime, eTime)
    df_detail = df_detail.loc[
        (df_detail["业务范围代码"].str.contains(businessCodeStr)) & (df_detail["凭证记帐日期"].str[:10].isin(validDateList)) & (
                df_detail["欠款金额"] > 0)]
    # 做透视，若为全部核销的数据需要将透视结果均写入结果表
    df_detail[pivotIdxCol] = df_detail[pivotIdxCol].fillna("")
    df_detail = df_detail.pivot_table(index=pivotIdxCol, values="欠款金额", aggfunc="sum")
    df_detail = df_detail.reset_index()

    addSaleDetailDf = pd.DataFrame(columns=df_detail.columns)
    df_e = df_e.sort_values(by="应还款日期", ascending=True).reset_index(drop=True)
    allOrderNum = set(df_detail["销售单代码"].tolist())
    for num in allOrderNum:
        df_match = df_e.loc[df_e["销售单代码"] == num.zfill(10)]
        df_base = df_detail.loc[df_detail["销售单代码"] == num].copy()
        if df_match.empty:
            # 销售订单号（销售单代码）不在第二张欠款明细表中，说明已被核销
            # 核销日期为出具发票日（凭证记帐日期）
            df_base["实际回款日"] = df_base["凭证记帐日期"]
            df_base["实际回款金额"] = df_base["欠款金额"]
            addSaleDetailDf = addSaleDetailDf.append(df_base)
        else:
            # 销售订单号（销售单代码）在第二张欠款明细表中，比较欠款金额，判断是否为部分核销
            diffAmount = new_round(df_base["欠款金额"].sum() - df_match["欠款金额"].sum())
            series_e = df_match.iloc[0].copy()  # 数据以欠款明细表该销售订单号（销售单代码）应还款日期最早的为准
            if diffAmount > 0:
                # todo:销售明细表匹配为部分核销的取值情况
                # series_e["索引"] = ""
                series_e["实际回款日"] = df_base.iloc[0]["凭证记帐日期"]
                series_e["实际回款金额"] = diffAmount
                series_e["欠款金额"] = diffAmount
                addSaleDetailDf = addSaleDetailDf.append(series_e)
            else:
                # 无回款，不进行处理
                continue
    else:
        addSaleDetailDf["扣款时间"] = addSaleDetailDf["凭证记帐日期"].apply(matchDeductTime)
        addSaleDetailDf["月份"] = addSaleDetailDf["凭证记帐日期"].apply(lambda x: f"{int(x[5:7])}月")

    df_e["扣款时间"] = matchDeductTime(eTime)
    df_e["月份"] = f"{int(eTime[5:7])}月"

    df_debt = pd.concat([df_e, addDebtDf, addSaleDetailDf]).reset_index(drop=True)
    df_debt["统计时间"] = eTime
    df_debt["上次统计时间"] = lastCalTimeDict["欠款明细"]

    # 新增“索引新增标识”列，标记需要计算的欠款数据索引与上次欠款明细表相比是否为新增索引
    lastIndexList = df_s.loc[df_s["索引"] != "", "索引"].tolist()
    df_debt["索引新增标识"] = df_debt["索引"].apply(lambda x: "" if x in lastIndexList else "新增")
    return df_debt


# 操作《欠款明细》sheet, 计算总财务费用
def debtSheetCal(stime, etime, df_debt, personConfigPath):
    """
    :param sTime: 用户输入的开始日期
    :param eTime: 用户输入的结束日期
    :param df_debt: 需要计算的欠款明细df
    :param personConfigPath: 销售员信息配置表
    :return: 欠款明细计算结果df
    """
    # 初始化没有的列，默认为空, 下面同时计算多列数据时要求每列都存在
    for col in debtResultCol:
        if col not in df_debt.columns:
            df_debt[col] = ""

    # 去除排除计算的客户名称、公司代码、索引数据
    df_debt = df_debt[(~df_debt["客户名称"].isin(delCompanyName)) & (~df_debt["索引"].isin(delIndex))]
    for code in delCompnyCode:
        df_debt = df_debt[~df_debt["客户代码"].str.startswith(code)]

    # 匹配一级部门、二级部门、三级部门
    global departmentMatchDict, personUseCol
    personfigDict = {"事业部": "三级部门", "区域": "二级部门", "平台": "一级部门"}
    personUseCol = list(personfigDict.values())
    df_person = pd.read_excel(personConfigPath, dtype=str)
    df_person.rename(columns=personfigDict, inplace=True)
    departmentMatchDict = dict(zip(df_person["人员代码"], df_person[personUseCol].values.tolist()))
    df_debt.loc[:, personUseCol] = df_debt.apply(
        lambda x: pd.Series(departmentMatchDict.get(x["人员编号"], [""] * 3), index=personUseCol), axis=1)

    # df_debt["采购类型"] = df_debt["三级部门"].apply(lambda x: "商业分销" if x == "商业业务部" else "")
    df_debt["采购类型"] = ""
    df_debt.loc[df_debt["采购类型"] == "", "采购类型"] = df_debt.loc[df_debt["采购类型"] == "", "业务范围代码"].apply(
        lambda x: "商业分销" if x.startswith("HV") else "")

    # 初始化利率
    df_debt["正常贷款费率"] = 0.055
    df_debt["超额贷款费率"] = 0.09
    df_debt["贴现利率"] = 0.055

    # 获取设置公式时需要的列所在列标识
    indexDict = {}
    for colName in ["应还款日期", "收付基准日期", "正常贷款费率", "欠款金额", "超期财务费用", "超额贷款费率", "上次统计时间", "统计时间", "实际回款日", "账期财务费用",
                    "贴现利息", "凭证记帐日期"]:
        colLetter = get_column_letter(debtResultCol.index(colName) + 1)
        indexDict[colName] = colLetter
    # 计算财务费用(包含已还款和未还款)
    df_debt[["账期财务费用", "超期财务费用", "备注"]] = df_debt.apply(caldebtCost, args=(stime, etime, indexDict), axis=1)
    df_debt = df_debt.loc[df_debt["备注"] != "去除"]
    # 计算总财务费用
    df_debt["总财务费用"] = '=INDIRECT("{账期财务费用}"&ROW())+INDIRECT("{超期财务费用}"&ROW())+INDIRECT("{贴现利息}"&ROW())'.format(
        **indexDict)

    # 筛选出需要的列
    df_debt = df_debt[debtResultCol]

    return df_debt


# 操作《银票》sheet, 计算贴现利息
def bankNotesOperateAndCal(baseDfDict, bankNotesPath):
    """
    :param baseDfDict: 基础数据df字典{"欠款明细":欠款明细df, "银票":银票df, "预收补费用":预收补费用df}
    :param bankNotesPath: 银票记录表路径
    :return: 银票计算结果df，本次处理的银票年份
    """
    # 在"银票记录表"中筛选出不在基础表中的数据
    bankNotesBaseDf = baseDfDict["银票"]
    existNumerList = bankNotesBaseDf["票号"].tolist()
    bankNotesDf = pd.read_excel(bankNotesPath, dtype=str)

    # 获取当前处理银票记录表中的数据所在年份
    calYear = bankNotesDf[~bankNotesDf["日期"].isna()].iloc[0]["日期"][:4]

    # 初始化没有的列，默认为空，现阶段银票记录表字段不全，先将没有的列补充方便计算
    for col in bankNotesResultCol:
        if col not in bankNotesDf.columns:
            bankNotesDf[col] = ""
    bankNotesDf = bankNotesDf.loc[~bankNotesDf["票号"].isin(existNumerList), bankNotesResultCol]

    bankNotesDf["票面金额"] = pd.to_numeric(bankNotesDf["票面金额"])
    bankNotesDf["贴现利率"] = bankNotesDf["是否可贴"].apply(lambda x: 0.025 if x == "是" else 0.055 if x == "否" else np.nan)

    # 获取设置公式时需要的列所在列标识
    indexDict = {}
    for colName in ["到期日", "日期", "贴现利率", "票面金额"]:
        colLetter = get_column_letter(bankNotesResultCol.index(colName) + 1)
        indexDict[colName] = colLetter
    bankNotesDf[
        "贴现利息"] = '=DAYS(INDIRECT("{到期日}"&ROW()),INDIRECT("{日期}"&ROW())) * INDIRECT("{贴现利率}"&ROW()) / 365 * INDIRECT("{票面金额}"&ROW())'.format(
        **indexDict)

    # 匹配一级部门、二级部门、三级部门
    bankNotesDf.loc[:, personUseCol] = bankNotesDf.apply(
        lambda x: pd.Series(departmentMatchDict.get(x["人员编号"], [""] * 3), index=personUseCol), axis=1)

    # 对列排序
    bankNotesDf = bankNotesDf[bankNotesResultCol]

    return bankNotesDf, calYear


# 操作《预收补费用》sheet, 计算预收利息
def advanceOperateAndCal(sapFilePath, receivableFilePath, exportDate, lastCalTimeDict):
    """
    :param sapFilePath: sap下载的未清项目明细表路径
    :param receivableFilePath: O回款明细汇总表路径
    :param exportDate: 用户输入导出未清项目明细表的日期
    :param lastCalTimeDict: 统计时间字典{"欠款明细": 最新的统计时间, "预收补费用":最新的统计时间}
    :return: 预收补费用计算结果df，回款汇总表df（用于匹配之前预收数据的"销售员", "销售员代码"）
    """
    if exportDate != "":
        # sapDf = pd.read_html(sapFilePath, header=0)[0].fillna("").astype(str)
        # 用read_html直接读取导致部分列变为科学计数法
        sapFileXlsxPath = changeMhtmlToXlsx(sapFilePath)
        sapDf = pd.read_excel(sapFileXlsxPath, dtype=str)
        sapDf.rename(columns={"客户": "客户代码"}, inplace=True)

        # 筛选出计算费用的数据
        exportDate_ = datetime.strptime(exportDate, "%Y-%m-%d")
        minDate = datetime(year=exportDate_.year - 1, month=exportDate_.month, day=exportDate_.day).strftime("%Y-%m-%d")
        sapDf = sapDf[~sapDf["公司代码"].isna()]
        sapDf = sapDf[
            (sapDf["业务范围"].str.contains(businessCodeStr)) & (~sapDf["文本"].fillna("").str.contains(advanceDelStr)) & (
                    sapDf["过帐日期"].str[:10] >= minDate)]

        # 凭证编号、客户代码列需要补充至10位
        sapDf["索引temp"] = sapDf["凭证编号"].str.zfill(10) + sapDf["客户代码"].str.zfill(10)
        # 财务凭证号、客户代码列需要补充至10位
        receivableTotalDf = pd.read_excel(receivableFilePath, dtype=str)[["财务凭证号FI", "客户代码", "销售员", "销售员代码"]]
        receivableTotalDf.rename(columns={"销售员": "销售员姓名", "销售员代码": "人员编号"}, inplace=True)
        receivableTotalDf["索引temp"] = receivableTotalDf["财务凭证号FI"].str.zfill(10) + receivableTotalDf["客户代码"].str.zfill(
            10)
        advanceDf = pd.merge(sapDf, receivableTotalDf, how="left", on="索引temp", suffixes=("", "_re"))

        advanceDf["本币金额"] = pd.to_numeric(advanceDf["本币金额"])
        advanceDf["利率"] = 0.055
        advanceDf["统计时间"] = exportDate
        advanceDf["上次统计时间"] = lastCalTimeDict["预收补费用"]

        # 获取设置公式时需要的列所在列标识
        indexDict = {}
        for colName in ["统计时间", "上次统计时间", "过帐日期", "利率", "本币金额"]:
            colLetter = get_column_letter(advanceResultCol.index(colName) + 1)
            indexDict[colName] = colLetter
        advanceDf["预收利息"] = advanceDf.apply(calAdvanceCost, args=(lastCalTimeDict["预收补费用"], indexDict), axis=1)

        # 初始化没有的列，默认为空，结果表中的列可能比计算出的数据列多，这些列默认为空即可
        for col in advanceResultCol:
            if col not in advanceDf.columns:
                advanceDf[col] = ""

        # 匹配一级部门、二级部门、三级部门
        advanceDf.loc[:, personUseCol] = advanceDf.apply(
            lambda x: pd.Series(departmentMatchDict.get(x["人员编号"], [""] * 3), index=personUseCol), axis=1)

        # 去除多余的列（包括参与计算的临时列）
        advanceDf = advanceDf[advanceResultCol]

    else:
        advanceDf = pd.DataFrame(columns=advanceResultCol)
        receivableTotalDf = pd.DataFrame()

    return advanceDf, receivableTotalDf


# 将各sheet的df合并，保存到结果表中
def saveDataToFile(baseDfDict, df_debt, bankNotesDf, advanceDf, resultPath, eTime, bankNotesCalYear, receivableTotalDf,
                   baseTablePath):
    """
    :param baseDfDict: 基础数据df字典{"欠款明细":欠款明细df, "银票":银票df, "预收补费用":预收补费用df}
    :param df_debt: 银票计算结果df
    :param bankNotesDf: 银票计算结果df
    :param advanceDf: 预收补费用计算结果df
    :param resultPath: 结果表路径
    :param eTime: 用户输入的情况明细结束日期
    :param bankNotesCalYear: 本次处理的银票年份
    :param receivableTotalDf: 回款汇总表df（用于匹配之前预收数据的"销售员", "销售员代码"）
    :param baseTablePath: 账期费用基础表路径
    :return:
    """
    global recordDict
    debtSheetName = f"FY{eTime[2:4]}账期明细-欠款明细"
    bankNotesSheetName = f"FY{bankNotesCalYear[2:4]}账期费用-银票"
    advanceSheetName = "预收补费用"

    addDataDict = {"欠款明细": df_debt, "银票": bankNotesDf, "预收补费用": advanceDf}
    newSheetNameDict = {"欠款明细": debtSheetName, "银票": bankNotesSheetName, "预收补费用": advanceSheetName}
    with pd.ExcelWriter(resultPath) as writer:
        book = openpyxl.load_workbook(baseTablePath)
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        for nameKeyWord, df_cal in baseDfDict.items():
            # 获取处理的数据需要写入的sheet的名称
            targetSheetName = ""
            for sheetName in book.sheetnames:
                if nameKeyWord in sheetName:
                    targetSheetName = sheetName
                    break
            # 写入数据
            startRow = df_cal.shape[0] + 1
            addDataDict[nameKeyWord].to_excel(excel_writer=writer, sheet_name=targetSheetName, index=False,
                                              header=False, startrow=startRow)

            # 记录数据写入的起始行和结束行
            recordDict[targetSheetName] = [df_cal.shape[0] + 2, df_cal.shape[0] + addDataDict[nameKeyWord].shape[0] + 1]
            # 预收补费用sheet需要完善基础表中无销售员信息的数据
            """
            if nameKeyWord == "预收补费用":
                if not receivableTotalDf.empty:  # receivableTotalDf为空说明本次《预收补费用》sheet无需操作
                    advanceSalesDict = dict(
                        zip(receivableTotalDf["索引temp"], receivableTotalDf[["销售员姓名", "人员编号"]].values.tolist()))
                    indexList = df_cal.loc[df_cal["销售员姓名"].isna()].index.tolist()
                    print(indexList)
                    letterDict = {}
                    for colName in ["凭证编号", "客户代码", "销售员姓名", "人员编号"]:
                        letterDict[colName] = get_column_letter(advanceResultCol.index(colName) + 1)
                    ws = writer.sheets[targetSheetName]
                    # 遍历需要处理的每行数据，补充销售员信息
                    for idx in indexList:
                        rowNum = idx + 2
                        colValue1 = ws[f'{letterDict["凭证编号"]}{rowNum}'].value
                        if not colValue1:
                            colValue1 = ""
                        colValue2 = ws[f'{letterDict["客户代码"]}{rowNum}'].value
                        if not colValue2:
                            colValue2 = ""
                        indexFlag = str(colValue1).zfill(10) + str(colValue2).zfill(10)
                        saleMsgList = advanceSalesDict.get(indexFlag, None)
                        if saleMsgList:
                            ws[f'{letterDict["销售员姓名"]}{rowNum}'].value = saleMsgList[0]
                            ws[f'{letterDict["人员编号"]}{rowNum}'].value = saleMsgList[1]
             """

            # 可能存在跨年的情况，需要修改sheet名称
            writer.sheets[targetSheetName].title = newSheetNameDict[nameKeyWord]

        book.close()
        # writer.save()
        # writer.close()


# 设置结果表各sheet样式
def setStyle(resultPath):
    global recordDict

    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(resultPath)

    sheetNames = [sheet.name for sheet in wb.sheets]
    for sheetName in sheetNames:
        ws = wb.sheets[sheetName]
        s_row, e_row = recordDict[sheetName]
        if "欠款明细" in sheetName:
            colIndex = get_column_letter(len(debtResultCol))
            # 全表格式改成常规格式、表头字体为微软雅黑9号、内容字体为微软雅黑9号，取消默认加粗
            ws.range(f"A:{colIndex}").number_format = "G/通用格式"
            ws.range(f"A:{colIndex}").font.size = 9
            ws.range(f"A:{colIndex}").font.name = "微软雅黑"
            ws.range(f"A1:{colIndex}1").font.bold = False

            # "公司代码"、"客户代码"、"系统发票代码"、"索引"、"人员编号"列为文本格式
            for col in ["公司代码", "客户代码", "系统发票代码", "索引", "人员编号"]:
                colLetter = get_column_letter(debtResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "@"
            # 各金额列设置为千分位显示两位小数，负数用()括起来，部分列设置颜色
            for col in ["欠款金额", "实际回款金额", "账期财务费用", "超期财务费用", "贴现利息", "总财务费用"]:
                colLetter = get_column_letter(debtResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "#,##0.00_);[红色](#,##0.00)"
                if col in ["账期财务费用", "超期财务费用", "贴现利息", "总财务费用"]:
                    ws.range(f"{colLetter}1:{colLetter}1").color = (255, 255, 0)
            # 各利率列设置为百分比显示两位小数，设置颜色
            for col in ["正常贷款费率", "超额贷款费率", "贴现利率"]:
                colLetter = get_column_letter(debtResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "0.00%"
                ws.range(f"{colLetter}1:{colLetter}1").color = (146, 208, 80)
            # 各日期列设置为只显示年月日
            for col in ["凭证记帐日期", "应还款日期", "收付基准日期", "实际回款日", "统计时间", "上次统计时间"]:
                colLetter = get_column_letter(debtResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "yyyy/m/d"

            # "业务范围代码"到"实际回款金额"列加粗
            colIndex1 = get_column_letter(debtResultCol.index("业务范围代码") + 1)
            colIndex2 = get_column_letter(debtResultCol.index("实际回款金额") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").font.bold = True
            # 设置"采购类型"到"备注"列背景颜色
            colIndex1 = get_column_letter(debtResultCol.index("采购类型") + 1)
            colIndex2 = get_column_letter(debtResultCol.index("备注") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").color = (0, 176, 240)

            # 重新载入数据以正确显示(数据有公式，使用value会导致公式丢失)
            ws.range(f"A{s_row}").value = ws.range(f"A{s_row}:{colIndex}{e_row}").formula

            # 进行冻结操作
            active_window = wb.app.api.ActiveWindow
            active_window.FreezePanes = False
            # wb.app.range("A2").select()  # 选"A2"冻结首行
            # active_window.SplitColumn = resultCol.index("客户名称") + 1  # 冻结至哪一列
            active_window.SplitRow = 1  # 冻结至哪一行
            active_window.FreezePanes = True

            # 设置列宽
            ws.used_range.column_width = 10
            for colName in ["客户名称", "索引", "项目名称"]:
                col = get_column_letter(debtResultCol.index(colName) + 1)
                ws.range(f"{col}:{col}").column_width = 15

            # 设置当前sheet为活动sheet
            ws.activate()

        elif "银票" in sheetName:
            colIndex = get_column_letter(len(bankNotesResultCol))
            # 全表格式改成常规格式、表头字体为微软雅黑9号、内容字体为微软雅黑9号，取消默认加粗
            ws.range(f"A:{colIndex}").number_format = "G/通用格式"
            ws.range(f"A:{colIndex}").font.size = 9
            ws.range(f"A:{colIndex}").font.name = "微软雅黑"
            ws.range(f"A1:{colIndex}1").font.bold = False

            # "票号"、"人员编号"列为文本格式
            for col in ["票号", "人员编号"]:
                colLetter = get_column_letter(bankNotesResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "@"
            # 各金额列设置为千分位显示两位小数，负数用()括起来，部分列设置颜色
            for col in ["票面金额", "贴现利息"]:
                colLetter = get_column_letter(bankNotesResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "#,##0.00_);[红色](#,##0.00)"
                if col in ["贴现利息"]:
                    ws.range(f"{colLetter}1:{colLetter}1").color = (255, 255, 0)
            # 各利率列设置为百分比显示两位小数，设置颜色
            for col in ["贴现利率"]:
                colLetter = get_column_letter(bankNotesResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "0.00%"
                ws.range(f"{colLetter}1:{colLetter}1").color = (146, 208, 80)
            # 各日期列设置为只显示年月日
            for col in ["日期", "出票日期", "到期日"]:
                colLetter = get_column_letter(bankNotesResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "yyyy/m/d"

            # "销售员姓名"到"人员编号"列加粗
            colIndex1 = get_column_letter(bankNotesResultCol.index("销售员姓名") + 1)
            colIndex2 = get_column_letter(bankNotesResultCol.index("人员编号") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").font.bold = True
            # 设置"业务范围"到"是否可贴"列背景颜色
            colIndex1 = get_column_letter(bankNotesResultCol.index("业务范围") + 1)
            colIndex2 = get_column_letter(bankNotesResultCol.index("是否可贴") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").color = (196, 189, 151)
            # 设置"一级部门"到"备注"列背景颜色
            colIndex1 = get_column_letter(bankNotesResultCol.index("一级部门") + 1)
            colIndex2 = get_column_letter(bankNotesResultCol.index("备注") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").color = (0, 176, 240)

            # 重新载入数据以正确显示
            ws.range(f"A{s_row}").value = ws.range(f"A{s_row}:{colIndex}{e_row}").formula

            # 进行冻结操作
            active_window = wb.app.api.ActiveWindow
            active_window.FreezePanes = False
            # wb.app.range("A2").select()  # 选"A2"冻结首行
            # active_window.SplitColumn = resultCol.index("客户名称") + 1  # 冻结至哪一列
            active_window.SplitRow = 1  # 冻结至哪一行
            active_window.FreezePanes = True

            # 设置列宽
            ws.used_range.column_width = 10
            for colName in ["票号", "交来客户", "出票人", "承兑银行"]:
                col = get_column_letter(bankNotesResultCol.index(colName) + 1)
                if colName in ["票号"]:
                    ws.range(f"{col}:{col}").column_width = 33
                else:
                    ws.range(f"{col}:{col}").column_width = 28

        elif "预收补费用" in sheetName:
            colIndex = get_column_letter(len(advanceResultCol))
            # 全表格式改成常规格式、表头字体为微软雅黑9号、内容字体为微软雅黑9号，取消默认加粗
            ws.range(f"A:{colIndex}").number_format = "G/通用格式"
            ws.range(f"A:{colIndex}").font.size = 9
            ws.range(f"A:{colIndex}").font.name = "微软雅黑"
            ws.range(f"A1:{colIndex}1").font.bold = False

            # "科目"、 "客户代码"、"分配"、"人员编号"列为文本格式
            for col in ["科目", "客户代码", "分配", "人员编号"]:
                colLetter = get_column_letter(advanceResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "@"
            # 各金额列设置为千分位显示两位小数，负数用()括起来，部分列设置颜色
            for col in ["本币金额", "预收利息"]:
                colLetter = get_column_letter(advanceResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "#,##0.00_);[红色](#,##0.00)"
                if col in ["预收利息"]:
                    ws.range(f"{colLetter}1:{colLetter}1").color = (255, 255, 0)
            # 各利率列设置为百分比显示两位小数，设置颜色
            for col in ["利率"]:
                colLetter = get_column_letter(advanceResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "0.00%"
                ws.range(f"{colLetter}1:{colLetter}1").color = (146, 208, 80)
            # 各日期列设置为只显示年月日
            for col in ["凭证日期", "过帐日期", "统计时间", "上次统计时间"]:
                colLetter = get_column_letter(advanceResultCol.index(col) + 1)
                ws.range(f"{colLetter}:{colLetter}").number_format = "yyyy/m/d"

            # "销售员姓名"到"上次统计时间"列加粗
            colIndex1 = get_column_letter(advanceResultCol.index("销售员姓名") + 1)
            colIndex2 = get_column_letter(advanceResultCol.index("上次统计时间") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").font.bold = True
            # 设置"公司代码"到"文本"列背景颜色
            colIndex1 = get_column_letter(advanceResultCol.index("公司代码") + 1)
            colIndex2 = get_column_letter(advanceResultCol.index("文本") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").color = (221, 221, 221)
            # 设置"一级部门"到"备注"列背景颜色
            colIndex1 = get_column_letter(advanceResultCol.index("一级部门") + 1)
            colIndex2 = get_column_letter(advanceResultCol.index("备注") + 1)
            ws.range(f"{colIndex1}1:{colIndex2}1").color = (0, 176, 240)

            # 重新载入数据以正确显示
            ws.range(f"A{s_row}").value = ws.range(f"A{s_row}:{colIndex}{e_row}").formula

            # 进行冻结操作
            active_window = wb.app.api.ActiveWindow
            active_window.FreezePanes = False
            # wb.app.range("A2").select()  # 选"A2"冻结首行
            # active_window.SplitColumn = resultCol.index("客户名称") + 1  # 冻结至哪一列
            active_window.SplitRow = 1  # 冻结至哪一行
            active_window.FreezePanes = True

            # 设置列宽
            ws.used_range.column_width = 10
            for colName in ["分配", "文本"]:
                col = get_column_letter(advanceResultCol.index(colName) + 1)
                if colName in ["分配"]:
                    ws.range(f"{col}:{col}").column_width = 20
                else:
                    ws.range(f"{col}:{col}").column_width = 35
        else:
            continue

    wb.save(resultPath)
    wb.close()
    app.quit()


"""
delCompanyName:《欠款明细》数据需要去除的【客户名称】
delCompnyCode:《欠款明细》数据需要去除的【公司代码】
delIndex:《欠款明细》数据需要去除的【索引】
businessCode：《欠款明细》（包括销售明细表）数据需要计算的【业务范围代码】/【产品组】
businessCodeStr：《欠款明细》（包括销售明细表）数据需要计算的【业务范围代码】/【产品组】的pandas解析字符串
debtfilterCol：《欠款明细》数据筛选出的列名
debtRenameColDict:《欠款明细》数据需要修改的列名字典
saleDetailfilterCol:《销售明细》数据筛选出的列名
saleDetailRenameColDict:《销售明细》数据需要修改的列名字典
pivotIdxCol：《销售明细》数据透视的行列表
debtResultCol：《欠款明细》sheet的结果列名列表
bankNotesBaseCol：《银票记录表》数据筛选出的列名
bankNotesResultCol：《银票》sheet的结果列名列表
advanceResultCol：《预收补费用》sheet的结果列名列表
advanceDelTextList：《未清项目明细表》数据需要删除的关键词列表
advanceDelStr：《未清项目明细表》数据需要删除的关键词列表的pandas解析字符串
receivableTableCol:下载的回款明细表需要筛选的列（即回款明细汇总表中的列）
matchFlag：用于防止df.apply对第一条数据重复操作
logger：用于打印日志
recordDict: 记录各sheet页写入数据的起始行和结束行，例如：{"sheet1":[18,20], {"sheet2":[100,208]}
"""
delCompanyName = ["华为技术服务有限公司", "华为技术有限公司", "华为软件技术有限公司", "华为数字技术（成都）有限公司的数据"]
delCompnyCode = ["02", "2"]
delIndex = ["090152583742252"]
businessCode = ["PU", "HI", "HV", "QJ"]
businessCodeStr = "|".join(businessCode)
debtfilterCol = ["业务范围代码", "公司代码", "客户代码", "客户名称", "销售单代码", "系统发票代码", "索引", "销售员姓名", "销售员代码", "凭证记帐日期", "应还款日期",
                 "收付基准日期", "欠款金额"]
debtRenameColDict = {"销售员代码": "人员编号"}
saleDetailfilterCol = ["产品组", "销售组织", "客户编号", "客户名称", "销售订单号", "系统发票", "销售员", "销售员编码", "出具发票日", "预计还款日", "折扣后合同金额"]
saleDetailRenameColDict = {"产品组": "业务范围代码", "销售组织": "公司代码", "客户编号": "客户代码", "销售订单号": "销售单代码", "系统发票": "系统发票代码",
                           "销售员": "销售员姓名", "销售员编码": "人员编号", "出具发票日": "凭证记帐日期", "预计还款日": "应还款日期", "折扣后合同金额": "欠款金额"}
pivotIdxCol = ["业务范围代码", "公司代码", "客户代码", "客户名称", "销售单代码", "系统发票代码", "销售员姓名", "人员编号", "凭证记帐日期", "应还款日期", "收付基准日期"]
debtResultCol = ["业务范围代码", "公司代码", "客户代码", "客户名称", "销售单代码", "系统发票代码", "索引", "销售员姓名", "人员编号", "凭证记帐日期", "应还款日期",
                 "收付基准日期", "欠款金额", "实际回款日", "实际回款金额", "备注1", "扣款时间", "月份", "统计时间", "上次统计时间", "正常贷款费率", "账期财务费用",
                 "超额贷款费率", "超期财务费用", "贴现利率", "贴现利息", "总财务费用", "采购类型", "下单合同号", "项目名称", "一级部门", "二级部门", "三级部门", "产品",
                 "产品线", "服务产品线", "百分比", "备注"]
bankNotesBaseCol = ["业务范围", "日期", "票号", "票面金额", "交来客户", "出票人", "出票日期", "到期日", "承兑银行", "交票人", "公司", "是否可贴", "销售员姓名",
                    "人员编号"]
bankNotesResultCol = bankNotesBaseCol + ["贴现利率", "贴现利息", "一级部门", "二级部门", "三级部门", "备注"]

advanceResultCol = ["公司代码", "用户名称", "科目", "客户代码", "分配", "凭证编号", "业务范围", "凭证类型", "凭证日期", "过帐日期", "本币金额", "清帐凭证", "文本",
                    "销售员姓名", "人员编号", "统计时间", "上次统计时间", "利率", "预收利息", "一级部门", "二级部门", "三级部门", "备注"]

advanceDelTextList = ["样机资产转售到款", "暂挂不核", "负销售"]
advanceDelStr = "|".join(advanceDelTextList)
receivableTableCol = ["业务范围代码", "公司代码", "财务凭证号FI", "说明文本", "记帐日期", "输入日期", "客户代码", "客户名称", "利润中心本位币金额", "销售员", "销售员代码"]
matchFlag = False
logger = None
recordDict = {}

if __name__ == "__main__":
    g_dictGlobal = {"文件保存路径": r"D:\xc_files\账期费用",
                    "销售员大区对应表路径": r"D:\xc_files\账期费用\外挂-华为SBU销售员大区对应表-21.xlsx",
                    "银票记录表": r"D:\xc_files\账期费用\账期费用RPA模板-银票-23.xlsx",
                    }
    g_inputDict = {"账期费用基础表路径": r"D:\xc_files\账期费用\3.6\账期费用_20250304.xlsx",
                   "欠款明细开始日期": "2025-03-04", "欠款明细结束日期": "2025-03-05", "未清项目明细表下载日期": "2025-03-05"}

    debtFileList = [r"D:\xc_files\账期费用\3.6\FY25欠款明细-0304.xlsx",
                    r"D:\xc_files\账期费用\3.6\FY25欠款明细-0305.xlsx"]
    saleFileList = [r"D:\xc_files\账期费用\3.6\FY25销售明细(3月)_0305.xlsx"]
    sapFile = r"D:\xc_files\账期费用\3.6\未清项目明细20250305.MHTML"
    receivableFileList = r"D:\xc_files\账期费用\3.6\FY25回款明细-0305.xlsx"
    resultPath = os.path.join(g_dictGlobal["文件保存路径"], "账期费用.xlsx")

    # g_dictGlobal = {"文件保存路径": r"C:\Users\11598\Desktop",
    #                 "销售员大区对应表路径": r"E:\Uibot项目文件\配置表\外挂-华为SBU销售员大区对应表-21.xlsx",
    #                 "银票记录表": r"E:\My_Project\神州数码\Uibot项目\华为账期费用需求资料\账期费用规则\FY22-银票记录表.xlsx",
    #                 }
    # g_inputDict = {"账期费用基础表路径": r"E:\My_Project\神州数码\Uibot项目\华为账期费用需求资料\账期费用测试文件\账期费用RPA基础表-空表.xlsx",
    #                "欠款明细开始日期": "2022-11-29", "欠款明细结束日期": "2022-12-06", "未清项目明细表下载日期": "2022-09-30"}
    #
    # debtFileList = [r"E:\Uibot项目文件\download\FY22欠款明细-1129.xlsx", r"E:\Uibot项目文件\download\FY22欠款明细-1206.xlsx"]
    # saleFileList = [r"E:\Uibot项目文件\download\FY22销售明细(11月)_1130.xlsx",
    #                 r"E:\Uibot项目文件\download\FY22销售明细(12月)_1231 (1).xlsx"]
    # sapFile = r"E:\Uibot项目文件\download\未清项目明细20220930.MHTML"
    # receivableFileList = [r"E:\Uibot项目文件\download\FY22回款明细-0930.xlsx"]
    # resultPath = os.path.join(g_dictGlobal["文件保存路径"], "test账期费用.xlsx")

    baseDfDict, lastCalTimeDict = getBaseTableData(g_inputDict["账期费用基础表路径"])
    df_debt = debtSheetOperate(g_inputDict["欠款明细开始日期"], g_inputDict["欠款明细结束日期"], debtFileList, saleFileList,
                               lastCalTimeDict)
    df_debt = debtSheetCal(g_inputDict["欠款明细开始日期"], g_inputDict["欠款明细结束日期"], df_debt, g_dictGlobal["销售员大区对应表路径"])
    bankNotesDf, calYear = bankNotesOperateAndCal(baseDfDict, g_dictGlobal["银票记录表"])
    advanceDf, receivableTotalDf = advanceOperateAndCal(sapFile, receivableFileList, g_inputDict["未清项目明细表下载日期"],
                                                        lastCalTimeDict)
    saveDataToFile(baseDfDict, df_debt, bankNotesDf, advanceDf, resultPath, g_inputDict["欠款明细结束日期"],
                   calYear, receivableTotalDf, g_inputDict["账期费用基础表路径"])
    setStyle(resultPath)
