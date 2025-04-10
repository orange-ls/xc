#!/usr/bin/python3
# -*- coding: UTF-8 -*-


import calendar
import gc
import glob
import logging.config
import os
import re
from datetime import datetime, timedelta
from functools import wraps

import numpy as np
import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter


def judFile(filePath, qryOrderNum):
    df = pd.read_excel(filePath, dtype=str, header=None)
    orderNum = df.iloc[4, 2].strip()
    if orderNum != qryOrderNum:
        os.remove(filePath)
        raise Exception(f"需要下载{qryOrderNum}的里程碑付款表，但实际为{orderNum}")


# 初始化日志
def initWriteLog(rootDir):
    """
    :param rootDir: 日志写入目录
    :return:
    """
    global logger
    log_name = "销售明细"
    simple_format = '[%(levelname)s][%(asctime)s]%(message)s'
    logfile_dir = rootDir  # log文件的目录
    logfile_name = '%s.log' % log_name  # log文件名
    # 如果不存在定义的日志目录就创建一个
    if not os.path.exists(logfile_dir):
        os.makedirs(logfile_dir)

    # log文件的全路径
    logfile_path = os.path.join(logfile_dir, logfile_name)
    LOGGING_DIC = {
        'version': 1,
        'disable_existing_loggers': False,
        'formatters': {
            'simple': {
                'format': simple_format
            },
        },
        'filters': {},
        'handlers': {
            'default': {
                'level': 'INFO',
                'class': 'logging.handlers.RotatingFileHandler',
                'formatter': 'simple',
                'filename': logfile_path,
                'maxBytes': 1024 * 1024 * 5,
                'backupCount': 5,
                'encoding': 'utf-8',
            },
        },
        'loggers': {
            '': {
                'handlers': ['default'],
                'level': 'INFO',
                'propagate': True,
            },
        },
    }
    logging.config.dictConfig(LOGGING_DIC)
    logger = logging.getLogger(__name__)


# 定义一个日志装饰器，每次调用方法前后打印日志
# 注 uibot中无法调用，弃用
def logfun(func):
    @wraps(func)
    def logfunStep(*args, **kwargs):
        logger.info(f"{'-' * 25}{func.__name__}开始{'-' * 25}")
        result = func(*args, **kwargs)
        logger.info(f"{'-' * 25}{func.__name__}结束{'-' * 25}")
        return result

    return logfunStep


# 读取配置文件生成配置字典
def getConfigDict(baseConfPath):
    """
    :param baseConfPath：配置文件目录
    :return: 生成配置字典
    """
    df = pd.read_excel(baseConfPath, dtype=str, sheet_name="Sheet1").fillna("")
    resultDict = dict(zip(df["配置名称"], df["配置内容"]))
    # 华为账户密码有多个，处理为列表格式
    for key in ["华为网站账号", "华为网站密码", "销售明细排除的销售员"]:
        resultDict[key] = resultDict[key].replace("；", ";").strip(";").split(";")
    return resultDict


# 根据输入的文件根目录返回当日文件保存目录
def getSaveDir(rootDir):
    """
    :param rootDir：文件保存根目录，例E:\downDir
    :return: 返回根目录+当日日期的目录，如E:\downDir\2022年\6月\8日
    """
    today = datetime.now()
    year, month, day = today.year, today.month, today.day
    finialDir = os.path.join(rootDir, "%s年\\%s月\\%s日" % (year, month, day))
    return finialDir


# 获取查询下载的日期范围
def getQryTimeRange(saveDir, keyWord, timeFmtStr, user=""):
    """
    :param saveDir: 汇总表保存的路径
    :param keyWord: 文件名关键词标识
    :param timeFmtStr: 时间日期格式转换类型，例 %Y/%m/%d
    :param user: 下载华为订单表和业绩表的账号
    :return: 查找的文件路径，本次查询开始日期，本次查询结束日期, 文件最新标志等
    """
    nowday = datetime.now()
    yesterday = nowday - timedelta(days=1)

    if keyWord == "物料移动明细汇总":
        """
        物料移动汇总：文件名:物料移动明细汇总_20210728.xlsx
        1.日期表示文件保存的截止日期，物料移动汇总会下载截止到昨天的数据，再次下载时无需下载重复日期的数据
        2.下载时的日期即为2021/07/29-昨天
        """
        fileList = glob.glob(f"{saveDir}\\*{keyWord}_*.xlsx")
        filePath = fileList[0]
        lastDateStr = re.search(".*_(\d{8}).xlsx", os.path.basename(filePath)).group(1)
        # lastDate： 下载的起始日期
        lastDate = datetime.strptime(lastDateStr, "%Y%m%d") + timedelta(days=1)
        dayDelta = (yesterday - lastDate).days
        if dayDelta <= -1:
            lastFlag = True
        else:
            lastFlag = False
        startDate = lastDate.strftime(timeFmtStr)
        endDate = yesterday.strftime(timeFmtStr)
        # return 物料移动汇总表路径，下载开始日期，下载结束日期， 文件最新标识
        return filePath, startDate, endDate, lastFlag
    elif keyWord == "预提表":
        """
        预提表：文件名:预提表_20210728.xlsx
        1.日期表示文件保存的截止日期，预提表会下载本年的数据，当年数据会重新下载
        2.下载时的日期即为2021/01/01-当天
        """
        fileList = glob.glob(f"{saveDir}\\*{keyWord}_*.xlsx")
        filePath = fileList[0]
        lastDateStr = re.search(".*_(\d{8}).xlsx", os.path.basename(filePath)).group(1)
        lastDate = datetime.strptime(lastDateStr, "%Y%m%d")
        startDate = datetime(lastDate.year, 1, 1)
        endDate = nowday
        dayDelta = (endDate - lastDate).days
        if dayDelta <= 0:
            lastFlag = True
        else:
            lastFlag = False
        startDate = startDate.strftime(timeFmtStr)
        endDate = endDate.strftime(timeFmtStr)
        # yearRange：下载数据的年份跨度，用于替换汇总表该年份数据
        yearRange = list(pd.period_range(startDate, endDate, freq="Y").year)
        # return 预提汇总表路径，下载开始日期，下载结束日期， 文件最新标识，下载日期范围年份跨度
        return filePath, startDate, endDate, lastFlag, yearRange
    elif keyWord == "订单表":  # todo：1.配置表中的每个账号必须已经有汇总数据，新账号先自行下载一部分数据按指定规则命名即可，否则不会下载改新账号数据 2.如需兼容新账号无数据的情况，此方法及合并表的方法需要修改
        # """
        # 华为订单表：文件名:账号_订单表_20210728.xlsx
        # 1.日期表示文件保存的截止日期，华为订单表会下载截止到昨天的数据，再次下载时无需下载重复日期的数据
        # 2.下载时的日期即为2021/07/29-昨天
        # """
        # fileList = glob.glob(f"{saveDir}\\{user}_{keyWord}_*.xlsx")
        # filePath = fileList[0]
        # matchObj = re.search("订单表_(\d{8}).xlsx", os.path.basename(filePath))
        #
        # lastDate = datetime.strptime(matchObj.group(1), "%Y%m%d") + timedelta(days=1)
        # dayDelta = (yesterday - lastDate).days
        # if dayDelta <= -1:
        #     lastFlag = True
        # else:
        #     lastFlag = False
        # startDate = lastDate.strftime(timeFmtStr)
        # endDate = yesterday.strftime(timeFmtStr)
        # # return 华为订单汇总表路径，下载开始日期，下载结束日期， 文件最新标识
        # return filePath, startDate, endDate, lastFlag

        """ 变更：订单表不再是不断汇总的形式，每次直接下载近2年的表"""
        startDay = datetime(year=nowday.year - 2, month=1, day=1)
        startDate = startDay.strftime(timeFmtStr)
        endDate = nowday.strftime(timeFmtStr)
        return startDate, endDate

    elif keyWord == "业绩表":
        """
        华为业绩表：文件名:账号_2020业绩表（20210203）.xlsx
        1.每个账号华为业绩表均会下载2020-当年的数据，重新下载的年份的数据均需要替换
        2.n年的数据需要在>n年下载才无需进行替换
        3.如有跨年情况，上次为20211228下载的2021年数据，本次需要下载到2022年数据，则2021年数据需要重新下载，在下次执行时即不需要
        """
        fileList = glob.glob(f"{saveDir}\\{user}_*{keyWord}（*）.xlsx")
        # totalYearList：从2020年到当年的所有年
        totalYearList = list(pd.period_range(2020, nowday.year, freq="Y").year)
        totalYearList = [str(i) for i in totalYearList]
        # 汇总表目录中的业绩表已经存在的数据年份
        existYear = []
        # 需要下载的年份
        downYearList = []
        # 数据非最新的文件，需要删除
        deleteFileList = []

        for filePath in fileList:
            matchObj = re.search("(\d{4})业绩表（(\d{8})）.xlsx", os.path.basename(filePath))
            existYear.append(matchObj.group(1))
            # 如果非当年数据不是由后几年下载的，则需要重新下载且删除
            if matchObj.group(1) != str(nowday.year):
                if matchObj.group(1) >= matchObj.group(2)[:4]:
                    downYearList.append(matchObj.group(1))
                    deleteFileList.append(filePath)
            else:  # 当年但非当天下载的数据需要重新下载且删除
                if matchObj.group(2) < nowday.strftime("%Y%m%d"):
                    downYearList.append(matchObj.group(1))
                    deleteFileList.append(filePath)
        # 汇总表目录中从2020-当年的数据，出去已经存在并校验的数据，缺失的年份需重新下载
        for i in existYear:
            totalYearList.remove(i)
        downYearList.extend(totalYearList)
        # return 下载的年份列表，需要删除的文件列表
        return downYearList, deleteFileList
    elif keyWord == "订单全字段报表":  # todo：1.配置表中的每个账号必须已经有汇总数据，新账号先自行下载一部分数据按指定规则命名即可，否则不会下载改新账号数据 2.默认下载数据的开始截止时间为 01:00:00
        """
        华为订单全字段报表：文件名:账号_订单全字段报表_20210728.xlsx
        1.日期表示文件保存的截止日期，华为订单全字段报表会下载截止到当天01:00:00的数据，再次下载时无需下载重复日期的数据
        2.下载时的日期即为2021/07/28-当天凌晨1点
        """
        timeSuffix = " 01:00:00"
        fileList = glob.glob(f"{saveDir}\\{user}_{keyWord}_*.xlsx")
        filePath = fileList[0]
        matchObj = re.search("订单全字段报表_(\d{8}).xlsx", os.path.basename(filePath))

        lastDate = datetime.strptime(matchObj.group(1), "%Y%m%d")
        dayDelta = (nowday - lastDate).days
        if dayDelta <= 0:
            lastFlag = True
        else:
            lastFlag = False
        startDate = lastDate.strftime(timeFmtStr) + timeSuffix
        endDate = nowday.strftime(timeFmtStr) + timeSuffix
        # return 华为订单汇总表路径，下载开始日期，下载结束日期， 文件最新标识
        return filePath, startDate, endDate, lastFlag


# 四舍五入
def new_round(_float, _len=2):
    """
    :param _float: 需要四舍五入的数
    :param _len: 保留小数点位数
    :return: 四舍五入结果
    """
    if isinstance(_float, float):
        if str(_float)[::-1].find('.') <= _len:
            return _float
        if str(_float)[-1] == '5':
            return round(float(str(_float)[:-1] + '6'), _len)
        else:
            return round(_float, _len)
    else:
        return round(_float, _len)


# 获取目录下相同文件格式的文件列表
def getSameFormatFile(rootDir, keyWord):
    """
    :param rootDir: 需要查找的文件目录
    :param keyWord: 查找的关键词
    :return:
    """
    fileList = glob.glob(f"{rootDir}\\*{keyWord}*.xlsx")
    return fileList


# 预处理毛利结果表的“备注”列（通过订单表补充“备注”列）
def initAnalyzeNoteText(HWOrderPathList, analyzePath):
    """
    :param HWOrderPathList: 华为订单表路径列表
    :param analyzePath: 毛利分析结果表
    :return:
    """
    # 判断处理后的结果表是否存在，存在直接返回
    resltPath = os.path.join(os.path.dirname(analyzePath),
                             os.path.basename(analyzePath).replace(".xlsx", "_预处理备注.xlsx"))
    if os.path.exists(resltPath):
        return resltPath
    # 读取华为订单表，获取“合神”、“北神”、“城投”的下单合同号列表
    hs, bs, ct = [], [], []
    for path in HWOrderPathList:
        df_ = pd.read_excel(path, dtype=str)
        nameFlag = os.path.basename(path).split("_")[0]
        if nameFlag == "13544480167":
            ct = df_["华为订单号"].tolist()
        elif nameFlag == "hfszsm":
            hs = df_["华为订单号"].tolist()
        elif nameFlag == "szshbj":
            bs = df_["华为订单号"].tolist()
        else:
            pass

    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(analyzePath)
    ws = wb.sheets["账面毛利分析"]
    rows = ws.used_range.shape[0]
    for row in range(2, rows + 1):
        orderNum = ws[f"H{row}"].value  # 下单合同号
        saleType = ws[f"V{row}"].value  # 销售类型
        buyType = ws[f"W{row}"].value  # 采购类型
        noteText = ws[f"U{row}"].value  # 备注
        amount = ws[f"O{row}"].value  # 成本总价
        # 如果下单合同号是None，说明是空行，跳过(可以break)
        if orderNum is None:
            continue

        # 备注不为"合神"、"北神"、"城投"就重新进行备注
        noteTextFlag = True if noteText in companySimpleDict.keys() else False
        # 补充符合条件但备注为空的单元格
        if amount > 0 and saleType == "正常销售" and orderNum.startswith("1Y") \
                and buyType in ['服务', '原厂下单'] and not noteTextFlag:
            ws[
                f"U{row}"].value = "合神" if orderNum in hs else "北神" if orderNum in bs else "城投" if orderNum in ct else ""

    wb.save(resltPath)
    wb.close()
    app.quit()
    return resltPath


# 合并华为订单全字段报表
def updateAllFieldFile(addfilePath, finalPath, dateFlag):
    """
    :param addfilePath: 读取未处理的华为订单全字段报表
    :param finalPath: 华为订单全字段报表汇总表路径
    :param dateFlag: 文件名的更新日期 %Y-%m-%d %H:%M:%S
    :return: 返回华为订单全字段报表汇总表路径
    """
    # 华为订单全字段报表在指定日期内有数据（无数据不下载，addfilePath为None）
    if addfilePath:
        # 初始化App
        app = xw.App(visible=True, add_book=False)
        app.display_alerts = False
        app.screen_updating = True
        # 打开下载表
        wb_ = app.books.open(addfilePath)
        ws_ = wb_.sheets[0]
        # 删除第一行标题行
        newCols = ws_.used_range.rows[0].value
        ws_.range("1:1").delete()

        # 打开汇总表
        wb = app.books.open(finalPath)
        ws = wb.sheets[0]
        oldCols = ws.used_range.rows[0].value
        if oldCols != newCols:
            raise Exception(f"汇总表{finalPath}和下载表{addfilePath}列不同")
        startRow = ws.used_range.shape[0] + 1

        # 将需要添加的数据复制到汇总表中
        ws_.used_range.copy(destination=ws.range(f"A{startRow}"))

        # ws.autofit()
        wb.save(finalPath)
        # 关闭工作簿
        wb.close()
        wb_.close()
        app.quit()

    # 更新文件名
    fileUser = os.path.basename(finalPath).split("_")[0]
    newFileName = f"{fileUser}_订单全字段报表_" + dateFlag.replace("-", "")[:8] + ".xlsx"
    newFilePath = os.path.join(os.path.dirname(finalPath), newFileName)
    os.rename(finalPath, newFilePath)

    # 清理内存
    gc.collect()
    return newFilePath


# 获取“激励、授信记录文件”路径，若不存在则进行创建
def getIncentiveRecordPath(inputPath):
    """
    :param inputPath: 查找的文件夹
    :return: “激励、授信记录文件”路径
    """
    targetPath = os.path.join(inputPath, "激励、授信记录文件.xlsx")
    if not os.path.exists(targetPath):
        df = pd.DataFrame(columns=["下单合同号", "使用激励金额", "是否使用授信"])
        df.to_excel(targetPath, index=False)
    return targetPath


# 读取激励、授信记录文件
def readIncentiveRecord(incentiveRecordPath):
    """
    :param incentiveRecordPath: "激励、授信记录文件"路径
    :return:字典{下单合同号：[使用激励金额, 是否使用授信]}
    """
    df_record = pd.read_excel(incentiveRecordPath, dtype=str)
    incentiveDict = dict(zip(df_record["下单合同号"], df_record[["使用激励金额", "是否使用授信"]].values.tolist()))
    return incentiveDict


# 更新激励、授信记录表
def updateIncentiveRecord(incentiveRecordPath, incentiveDict):
    """
    :param incentiveRecordPath: "激励金额记录文件"路径
    :param incentiveDict: 需要写入的字典{下单合同号：[使用激励金额, 是否使用授信]}}
    :return:
    """
    dfDict = {key: value for key, value in incentiveDict.items()}
    recordDf = pd.DataFrame(data=dfDict).T.reset_index()
    recordDf.columns = ["下单合同号", "使用激励金额", "是否使用授信"]
    recordDf.to_excel(incentiveRecordPath, index=False)


# 判断值是否在字符串、字典、列表中
def jud_in(val, targetObj):
    return val in targetObj


# 更新字典
def updateDict(oldDict, addDict):
    oldDict.update(addDict)
    return oldDict


# 初始化"备注"
def setRemark(series, creditDict):
    """
    :param series: df数据行
    :param creditDict: 授信字典：{华为合同号：[订单激活时间, 订单变更前金额]}
    :return: 备注
    """

    if series["供应商名称"] == companySimpleDict["城投"]:
        return "使用城投"
    # elif series["下单合同号"] in creditDict and series["开单日期"] >= creditDict[series["下单合同号"]][0][:10]:
    #     # 开单日期在授信时间后的标记为“使用授信”
    #     return "使用授信"
    else:
        return ""


# 初始化华为原厂付款数据，做汇总处理
def handlePayInfoYC(df_payO, df_pay):
    """
    :param df_payO: 某华为合同号的全部付款数据df
    :param df_pay: 某华为合同号新增的付款数据df
    :return: 字典{"已被核销付款df": 已被核销付款df, "付款df": 付款df}
    """
    # payRecordDf：已被核销的付款df（付款金额因被完全核销变为0）
    # finalPayDf：汇总后的付款df
    baseCols = ["收据编号", "付款日期", "付款金额", "下单合同号", "最新付款日期"]
    payRecordDf, finalPayDf = pd.DataFrame(columns=baseCols), pd.DataFrame(columns=baseCols)
    # 将全部付款数据df按照"付款日期"升序，新增的付款df进行透视
    df_payO = df_payO.sort_values(by="付款日期", ascending=True)
    df_pay["付款金额"] = pd.to_numeric(df_pay["付款金额"])
    df_pay["收据编号"] = df_pay["收据编号"].fillna("")
    df_payPT = df_pay.pivot_table(index="收据编号", values=["付款日期", "付款金额", "下单合同号"],
                                  aggfunc={"付款日期": "first", "付款金额": "sum", "下单合同号": "first"})
    # 无新增付款数据，则payRecordDf和finalPayDf均为空df
    if df_payPT.empty:
        return {"已被核销付款df": payRecordDf, "付款df": finalPayDf}

    # 判断新增付款透视数据的“付款金额”，如无<=0的数据，则汇总后的付款df即为df_payPT
    df_payPT = df_payPT.sort_values(by="付款日期", ascending=True).reset_index()
    if df_payPT.loc[df_payPT["付款金额"] <= 0].empty:
        finalPayDf = finalPayDf.append(df_payPT)
    else:
        # 对透视的付款数据“付款金额”=0的数据，存入payRecordDf；
        df_temp = df_payPT.loc[df_payPT["付款金额"] == 0]
        for idx, seriesTemp in df_temp.iterrows():
            # 被核销的付款数据，需要获取其"付款日期"
            firstDate = df_payO.loc[df_payO["收据编号"] == seriesTemp["收据编号"], "付款日期"].tolist()[0]
            seriesTemp["付款日期"] = firstDate
            payRecordDf = payRecordDf.append(seriesTemp)

        # 对透视的付款数据有“付款金额”<0的，需要对“付款金额”>0的数据处理：按照"付款日期"由远到近对<0的数据进行核销；
        df_negative = df_payPT.loc[df_payPT["付款金额"] < 0].copy()
        df_positive = df_payPT.loc[df_payPT["付款金额"] > 0].copy()
        for i, series_n in df_negative.iterrows():
            amount_n = series_n["付款金额"] * -1
            for j, series_p in df_positive.iterrows():
                amount_p = series_p["付款金额"]
                if amount_p == 0:
                    continue
                # 某"收据编号"付款金额>退款金额，将该笔退款清空，并扣除对应的付款金额
                if amount_p > amount_n:
                    df_positive.loc[j, "付款金额"] = new_round(amount_p - amount_n)
                    break
                else:
                    # 某"收据编号"付款金额<退款金额，扣除部分退款，对应的付款金额清0并存入payRecordDf
                    amount_n = new_round(amount_n - amount_p)
                    # 获取某笔付款的时间时：①如果该收据编号可以在付款表中查询到，取付款表中该收据编号第一次出现的日期；
                    # ②查询不到说明该笔付款是在下单费用基础表中（一般收据编号为空），取下单费用基础表的日期
                    dateList = df_payO.loc[df_payO["收据编号"] == series_p["收据编号"], "付款日期"].tolist()
                    if len(dateList) != 0:
                        firstDate = dateList[0]
                    else:
                        firstDate = series_p["付款日期"]
                    df_positive.loc[j, ["付款日期", "付款金额"]] = [firstDate, 0]
                    payRecordDf = payRecordDf.append(df_positive.loc[j])
            else:
                # 匹配所有付款数据后，若仍有剩余退款，说明之前部分退款已被开单处理，将这部分退款当做剩余付款存入finalPayDf
                if amount_n > 0:
                    series_n["付款金额"] = amount_n * -1
                    finalPayDf = finalPayDf.append(series_n)
        else:
            # 匹配完成后，将剩余付款存入finalPayDf
            finalPayDf = finalPayDf.append(df_positive.loc[df_positive["付款金额"] > 0])

    # 判断处理后数据日期是否为最新日期，非最新日期的情况需要记录
    def checkAndUpdateDate(df_):
        if not df_.empty and not df_payO.empty:
            lastDate = df_payO.loc[~df_payO["付款日期"].isna()].tail(1).iloc[0]["付款日期"]
            if df_.tail(1).iloc[0]["付款日期"] < lastDate:
                df_.iloc[-1, -1] = lastDate
        return df_

    payRecordDf = checkAndUpdateDate(payRecordDf)
    finalPayDf = checkAndUpdateDate(finalPayDf)

    return {"已被核销付款df": payRecordDf, "付款df": finalPayDf}


# 初始化 超聚变/授信后 的付款数据，做汇总处理
def handlePayInfo_CJB_Credit(df_pay):
    """
    :param df_pay: 某华为合同号新增的付款数据df
    :return: 字典{"已被核销付款df": 已被核销付款df, "付款df": 付款df}
    """
    # payRecordDf：已被核销的付款df（付款金额因被完全核销变为0）
    # finalPayDf：汇总后的付款df
    baseCols = ["付款日期", "付款金额", "下单合同号", "最新付款日期"]
    payRecordDf, finalPayDf = pd.DataFrame(columns=baseCols), pd.DataFrame(columns=baseCols)

    # 将付款数据df按照"付款日期"升序
    df_pay = df_pay.sort_values(by="付款日期", ascending=True)
    df_pay["付款金额"] = pd.to_numeric(df_pay["付款金额"])
    df_pay["付款日期"] = df_pay["付款日期"].fillna("")

    # 如有“付款金额”<0的数据，需要对“付款金额”>0的数据处理：按照"付款日期"由远到近对<0的数据进行核销；
    df_negative = df_pay.loc[df_pay["付款金额"] < 0].copy()
    df_positive = df_pay.loc[df_pay["付款金额"] > 0].copy()
    df_positive = df_positive.pivot_table(index=["付款日期"], values=["付款金额", "下单合同号"],
                                          aggfunc={"付款金额": "sum", "下单合同号": "first"}).reset_index()
    if df_positive.empty:
        df_positive = pd.DataFrame(columns=baseCols)

    for i, series_n in df_negative.iterrows():
        amount_n = series_n["付款金额"] * -1
        for j, series_p in df_positive.iterrows():
            amount_p = series_p["付款金额"]
            if amount_p == 0:
                continue
            # 若付款金额>退款金额，将该笔退款清空，并扣除对应的付款金额
            if amount_p > amount_n:
                df_positive.loc[j, "付款金额"] = new_round(amount_p - amount_n)
                break
            else:
                # 若付款金额<退款金额，扣除部分退款，对应的付款金额清0并存入payRecordDf
                amount_n = new_round(amount_n - amount_p)
                df_positive.loc[j, "付款金额"] = 0
                payRecordDf = payRecordDf.append(df_positive.loc[j])
        else:
            # 匹配所有付款数据后，若仍有剩余退款，说明之前部分退款已被开单处理，将这部分退款当做剩余付款存入finalPayDf
            if amount_n > 0:
                series_n["付款金额"] = amount_n * -1
                finalPayDf = finalPayDf.append(series_n)
    else:
        # 匹配完成后，将剩余付款存入finalPayDf
        finalPayDf = finalPayDf.append(df_positive.loc[df_positive["付款金额"] > 0])

    return {"已被核销付款df": payRecordDf, "付款df": finalPayDf}


# 通过下单数据匹配付款数据
def matchPayInfo(df_order, df_calPay, incentiveDict, payRenameDict, creditDate):
    """
    :param df_order: 下单df
    :param df_calPay: 付款df
    :param incentiveDict: 使用激励字典{下单合同号: 使用激励金额}
    :param payRenameDict: 付款表列名重命名字典(同时key为需要操作的列名)
    :param creditDate: 该合同号的授信日期
    :return: 匹配结果df(包含：正常匹配 + 未付款 + 剩余付款（包括退款）)
    """
    resultDf = pd.DataFrame()
    # 设置授信付款备注
    if not df_calPay.empty:
        df_calPay.loc[df_calPay["付款日期"].str[:10] == creditDate, "备注"] = "使用授信"
    # 匹配时，由于“开单金额”不含税，“付款金额”含税，需要将"开单金额"*(1+"实际税率")再进行匹配
    for i, series_o in df_order.iterrows():
        orderAmount = series_o["开单金额"] * (1 + series_o["实际税率"])
        for j, series_p in df_calPay.iterrows():
            payAmount = series_p["付款金额"]
            creditMsg = series_p["备注"]
            newlyDate = series_p["最新付款日期"]
            if payAmount < 0:
                # 如果payAmount有负值，则此次所有下单均为“未付款”
                series_o["付款日期"] = "未付款"
                series_o["开单金额（含税）"] = orderAmount
                resultDf = resultDf.append(series_o)
                break
            if payAmount == 0:
                continue
            # 若付款金额和开单金额差额在20内，则将这组数据完全匹配并存入resultDf
            if abs(payAmount - orderAmount) <= 20:
                # FIXME:不存在差额在20以上的，这种情况会通过剩余付款和未付款体现
                df_calPay.loc[j, "付款金额"] = 0
                series_o[list(payRenameDict.values())] = series_p[list(payRenameDict.values())]
                series_o["开单金额"] = new_round(orderAmount / (1 + series_o["实际税率"]))
                series_o["付款金额"] = payAmount
                series_o["备注"] = creditMsg
                series_o["最新付款日期"] = newlyDate
                resultDf = resultDf.append(series_o)
                break
            # 若付款金额 >= 开单金额，对应的付款金额减少，匹配的数据存入resultDf
            if payAmount >= orderAmount:
                df_calPay.loc[j, "付款金额"] = new_round(payAmount - orderAmount)
                series_o[list(payRenameDict.values())] = series_p[list(payRenameDict.values())]
                series_o["开单金额"] = new_round(orderAmount / (1 + series_o["实际税率"]))
                series_o["付款金额"] = orderAmount
                series_o["备注"] = creditMsg
                series_o["最新付款日期"] = newlyDate
                resultDf = resultDf.append(series_o)
                break
            else:
                # 若付款金额 < 开单金额:
                # 1.非授信开单：对应的开单金额减少，付款金额清空，匹配的数据存入resultDf
                # 2.授信开单：付款将开单全部核销，开单无需拆行，匹配的数据存入resultDf
                df_calPay.loc[j, "付款金额"] = 0
                series_o[list(payRenameDict.values())] = series_p[list(payRenameDict.values())]
                series_o["付款金额"] = payAmount
                series_o["备注"] = creditMsg
                series_o["最新付款日期"] = newlyDate
                if creditMsg == "使用授信":
                    series_o["开单金额"] = new_round(orderAmount / (1 + series_o["实际税率"]))
                    resultDf = resultDf.append(series_o)
                    break
                else:
                    orderAmount = new_round(orderAmount - payAmount)
                    series_o["开单金额"] = new_round(payAmount / (1 + series_o["实际税率"]))
                    resultDf = resultDf.append(series_o)

        else:
            # 匹配所有付款数据后，若仍有下单数据，先判断激励情况，再作为"未付款"的数据存入resultDf
            if orderAmount > 0:
                # FIXME:剩余付款和 激励金额进行匹配拆分，现仅将激励写入
                # 剩余下单信息从df_order重新获取而不直接使用series_o，以防series_o匹配过导致需要匹配的列有值
                surplusOrderSe = df_order.loc[i].copy()
                surplusOrderSe["开单金额（含税）"] = orderAmount
                surplusOrderSe["开单金额"] = new_round(orderAmount / (1 + surplusOrderSe["实际税率"]))
                # 通过激励字典获取授信金额，由于原厂数据是个列表，列表第一项为授信金额，而超聚变直接是str或float
                incentiveMatch = incentiveDict.get(surplusOrderSe["下单合同号"], [0, ""])
                if isinstance(incentiveMatch, list):
                    incentiveAmount = float(incentiveMatch[0])
                else:
                    incentiveAmount = float(incentiveMatch)
                if incentiveAmount > 0:
                    surplusOrderSe["备注"] = "使用激励"
                    surplusOrderSe["使用激励金额"] = incentiveAmount
                else:
                    surplusOrderSe["付款日期"] = "未付款"
                resultDf = resultDf.append(surplusOrderSe)
    else:
        # 匹配所有下单数据后，将未使用的付款（包括退款）存入resultDf
        resultDf = resultDf.append(df_calPay.loc[df_calPay["付款金额"] != 0])

    return resultDf


# 对某个下单合同号的付款数据初始化，汇总处理后匹配付款数据存入结果df
def matchPayDetail(orderNum, df_operate, df_payO, df_noUsePay, df_orderCredit, creditDict, payRenameDict, incentiveDict,
                   lastPayDateDict, flag):
    """
    :param orderNum: 处理的下单合同号
    :param df_operate: 操作的数据df（原厂下单、超聚变）
    :param df_payO: 该下单合同号的所有付款数据
    :param df_noUsePay: 下单费用表中的剩余付款df
    :param df_orderCredit: 该下单合同号的授信数据df（授信付款外挂表中获取）
    :param creditDict: 华为原厂授信数据的授信字典{合同号：[付款日期, 付款金额]}
    :param payRenameDict: 付款表列名重命名字典(同时key为需要操作的列名)
    :param incentiveDict: 使用激励字典{下单合同号: 使用激励金额}
    :param lastPayDateDict: 已有的最新付款时间字典{下单合同号: 付款时间}
    :param flag: 处理的数据类型（华为原厂、鲲泰、超聚变）
    :return: 某个下单合同号匹配付款后的结果df（包含：已被核销付款 + 正常匹配 + 未付款 + 剩余付款（包括退款））
    """

    resultDf = pd.DataFrame()
    creditDate = creditDict.get(orderNum, ["9999-12-31", ""])[0][:10]
    # 该合同号下单df
    df_order = df_operate.loc[df_operate["下单合同号"] == orderNum].reset_index(drop=True)
    df_order = df_order.sort_values(by="开单日期", ascending=True)
    # type1：授信前开单
    df_order_sx_before = df_order.loc[df_order["开单日期"].str[:10] < creditDate]
    # type2: 授信后开单
    df_order_sx_after = df_order.loc[df_order["开单日期"].str[:10] >= creditDate]

    # 该合同号剩余付款df
    df_orderNoUsePay = df_noUsePay.loc[df_noUsePay["下单合同号"] == orderNum]
    # type1:授信前剩余付款
    df_noUsePay_sx_before = df_orderNoUsePay.loc[df_orderNoUsePay["付款日期"].str[:10] < creditDate]
    # type2:授信后剩余付款
    df_noUsePay_sx_after = df_orderNoUsePay.loc[df_orderNoUsePay["付款日期"].str[:10] >= creditDate]

    # 获取新增的付款数据
    # type1：里程碑付款表的付款数据取：基础表该合同号最新付款数据后的以及授信前的数据
    df_payO = df_payO.copy()
    df_payO["付款金额"] = pd.to_numeric(df_payO["付款金额"])
    df_pay_sx_before = df_payO.loc[
        (lastPayDateDict.get(orderNum, "")[:10] < df_payO["付款日期"].str[:10]) & (
                    df_payO["付款日期"].str[:10] < creditDate)]
    # type2：授信付款外挂表的付款数据取：基础表该合同号最新付款数据后的数据
    df_credit_after = df_orderCredit.loc[lastPayDateDict.get(orderNum, "")[:10] < df_orderCredit["付款日期"].str[:10]]

    # type1：授信前的总付款
    df_sxPay_before = df_noUsePay_sx_before.append(df_pay_sx_before)
    # type2：授信后的总付款(需分为授信付款和授信后付款，授信后付款需要做汇总处理)
    df_sxPay_afterAll = df_noUsePay_sx_after.append(df_credit_after)
    df_sxPay_inDay = df_sxPay_afterAll.loc[df_sxPay_afterAll["付款日期"].str[:10] == creditDate]
    df_sxPay_after = df_sxPay_afterAll.loc[df_sxPay_afterAll["付款日期"].str[:10] > creditDate]

    # 先将非授信付款数据汇总处理，被核销的付款需要体现在下单费用结果表中
    if flag == "华为原厂":
        resultHandle = handlePayInfoYC(df_payO, df_sxPay_before)
    else:
        # 超聚变
        resultHandle = handlePayInfo_CJB_Credit(df_sxPay_before)
    # 授信后的付款数据汇总处理
    resultCredit = handlePayInfo_CJB_Credit(df_sxPay_after)

    resultDf = resultDf.append(resultHandle["已被核销付款df"]).append(resultCredit["已被核销付款df"])

    df_calPayCreditBefore = resultHandle["付款df"].reset_index(drop=True)
    df_calPayCreditAfter = df_sxPay_inDay.append(resultCredit["付款df"])
    df_calPayCreditAfter["付款金额"] = pd.to_numeric(df_calPayCreditAfter["付款金额"])

    # 授信前开单数据df_order_sx_before匹配授信前付款数据df_calPayCreditBefore
    matchResultCreditBeforeDf = matchPayInfo(df_order_sx_before, df_calPayCreditBefore, incentiveDict, payRenameDict,
                                             creditDate)
    resultDf = resultDf.append(matchResultCreditBeforeDf)
    # 授信后开单数据df_order_sx_before匹配授信后付款数据df_calPayCreditAfter
    matchResultCreditAfterDf = matchPayInfo(df_order_sx_after, df_calPayCreditAfter, incentiveDict, payRenameDict,
                                            creditDate)
    resultDf = resultDf.append(matchResultCreditAfterDf)

    return resultDf


# 筛选出超聚变付款外挂表中“付款时间”正常的数据
def filterValidPayInfoCJB(x):
    """
    :param x: 付款时间
    :return: 是否为时间的标识
    """
    fmt = "%Y-%m-%d %H:%M:%S"
    try:
        fmtTime = datetime.strptime(x, fmt)
        return True
    except Exception as e:
        return False


# 匹配"贷款利率"
def matchRate(series, flag):
    if flag != "鲲泰":
        # if series["事业部"] in ["区域事业部", "行业事业部", "超聚变业务部", "商业分销及电商业务部"]:
        if series["事业部"] in ["北区", "南区", "超聚变及商业分销", "新业务"]:
            if series["付款天数差"] <= 15:
                return 0.055
            else:
                if series["运输方式"] == "自提":
                    return 0.055
                elif series["运输方式"] in ["汽运", "空运"]:
                    return 0.09
                else:
                    return f"事业部已配置，付款天数差>15，运输方式为{series['运输方式']}"
        elif series["事业部"] == "服务事业部":
            if series["付款天数差"] <= 8:
                return 0.055
            else:
                return 0.09
        else:
            return "事业部未配置"
    else:
        return 0.055


# 匹配"下单费用"
def matchCost(series, flag):
    if not isinstance(series["贷款利率"], float):
        return series["贷款利率"]
    if flag != "鲲泰":
        # if series["事业部"] in ["区域事业部", "行业事业部", "超聚变业务部", "商业分销及电商业务部"]:
        if series["事业部"] in ["北区", "南区", "超聚变及商业分销", "新业务"]:
            return (series["付款天数差"] - 15) * series["付款金额"] * series["贷款利率"] / 365
        elif series["事业部"] == "服务事业部":
            return (series["付款天数差"] - 8) * series["付款金额"] * series["贷款利率"] / 365
        else:
            # 该步一般不会执行，因为匹配"贷款利率"时已经处理
            return "事业部未配置"
    else:
        return series["付款天数差"] * series["付款金额"] * series["贷款利率"] / 365

# 匹配扣款时间
def matchDeductTime(orderTime):
    year_ = orderTime[2:4]
    month_ = int(orderTime[5:7])
    quarter_ = (month_ + 2) // 3
    return f"FY{year_}-Q{quarter_}"


# 匹配扣款月份
def matchDeductMonth(orderTime):
    month_ = int(orderTime[5:7])
    return f"{month_}月"


# 鲲泰数据匹配“付款日期”
def setPayTimeKT(orderTimeStr):
    """
    :param orderTimeStr: 开单日期字符串
    :return: 付款日期字符串
    """
    fmt = "%Y-%m-%d %H:%M:%S"
    orderTime = datetime.strptime(orderTimeStr, fmt)
    # 对于日期 <=24 的数据，返回开单日期所在年月的25日
    if orderTime.day <= 24:
        payTime = datetime(year=orderTime.year, month=orderTime.month, day=25)
        return payTime.strftime(fmt)
    else:
        # 对于日期 >=25 的数据，返回开单日期所在年月的下月25日
        # 计算开单日期所在年月的天数
        _, monthDays = calendar.monthrange(orderTime.year, orderTime.month)
        # 计算开单日期所在年月的最后一天
        monthLastDate = datetime(year=orderTime.year, month=orderTime.month, day=monthDays)
        # 计算开单日期所在年月的下月第一天
        nextMonthDate = monthLastDate + timedelta(days=1)
        # 计算开单日期所在年月的下月的25日
        payTime = datetime(year=nextMonthDate.year, month=nextMonthDate.month, day=25)
        return payTime.strftime(fmt)


# 筛选出本次下单的有效数据
def filterOrderInfo(orderDf, lastOrderDateDict):
    """
    :param orderDf: 毛利分析结果表中筛选出的下单df
    :param lastOrderDateDict: 下单费用表中各合同号的最新开单日期字典{下单合同号: 开单日期}
    :return: 有效的下单df
    """
    allOrderNum = set(orderDf["下单合同号"])
    resultDf = pd.DataFrame(columns=orderDf.columns)
    for o_ in allOrderNum:
        df_temp = orderDf.loc[
            (orderDf["下单合同号"] == o_) & (orderDf["开单日期"].str[:10] > lastOrderDateDict.get(o_, "")[:10])]
        resultDf = resultDf.append(df_temp)
    return resultDf


# 读取下单费用表和毛利分析结果表，返回字典{"下单df": 下单数据df, "已处理df": 已处理的下单数据df, "剩余付款df": 未处理的付款数据df, "最新付款时间dict":最新付款时间dict}
def getOriginOrderData(analyzePath, orderCostPath):
    """
    :param analyzePath: 毛利分析结果表路径
    :param orderCostPath: 下单费用表路径
    :return:字典：{"下单df": 下单数据df, "已处理df": 已处理的下单数据df, "剩余付款df": 未处理的付款数据df, "最新付款时间dict":最新付款时间dict}
    """
    # 读取毛利分析结果表
    df_analyze = pd.read_excel(analyzePath, sheet_name="账面毛利分析", dtype=str, na_values=[''], keep_default_na=False)
    # 筛选出需要的列+["销售类型", "采购类型"]并重命名（"销售类型"和"采购类型"作为判断条件，后续会进行删除）
    df_analyze = df_analyze[filterCol + ["销售类型", "采购类型"]+ ["市场类型"]].rename(columns=renameColDict)
    df_analyze["开单金额"] = pd.to_numeric(df_analyze["开单金额"])
    # 筛选出符合条件的数据
    df_analyze = df_analyze.loc[
        (df_analyze["销售类型"] == "正常销售") & (
            df_analyze["采购类型"].isin(['服务', '原厂下单', '鲲泰', '超聚变'])) & (
                df_analyze["开单金额"] > 0)]
    # 将毛利分析结果表中的备注信息->供应商名称
    df_analyze["供应商名称"] = df_analyze["供应商名称"].apply(lambda x: companySimpleDict.get(x, x))

    # 读取下单费用基础表
    df_all = pd.read_excel(orderCostPath, sheet_name="下单费用", dtype=str, na_values=[''], keep_default_na=False)
    # 对下单费用表透视，获取每个订单号的付款最新日期
    df_all = df_all.sort_values(by=["付款日期"], ascending=False)
    df_pivot = df_all.loc[df_all["付款日期"] != "未付款"].pivot_table(index="下单合同号", values=["付款日期"],
                                                                      aggfunc={"付款日期": "first"})
    # df_pivot.drop_duplicates("下单合同号", keep="first", inplace=True, ignore_index=True)
    if df_pivot.empty:
        lastPayDateDict = {}
    else:
        lastPayDateDict = dict(zip(df_pivot.index, df_pivot["付款日期"]))

    # 对下单费用表透视，获取每个订单号的下单最新日期
    df_all = df_all.sort_values(by=["开单日期"], ascending=False)
    df_pivot = df_all.pivot_table(index="下单合同号", values=["开单日期"], aggfunc={"开单日期": "first"})
    # df_pivot.drop_duplicates("下单合同号", keep="first", inplace=True, ignore_index=True)
    if df_pivot.empty:
        lastOrderDateDict = {}
    else:
        lastOrderDateDict = dict(zip(df_pivot.index, df_pivot["开单日期"]))

    """---------------------测试---------------------"""
    lastCreditDict = {}
    df_all["备注"] = df_all["备注"].fillna("")
    a = 0
    b = 0
    df_credit = df_all.loc[(df_all["备注"].str.contains("授信")) & (df_all["使用激励金额"].isna())]
    creditOrderList = list(set(df_credit["下单合同号"].tolist()))
    for order in creditOrderList:
        df_match = df_credit.loc[(df_credit["下单合同号"] == order) & (~df_credit["付款日期"].isna())]
        if order == "1Y01302205010F":
            print(df_match)
        allCreditTime = df_match["付款日期"].tolist()
        allCreditTime = list(set([t_[:10] for t_ in allCreditTime]))
        if len(allCreditTime) != 1:
            a += 1
            continue
            # raise Exception(f"下单费用基础表中合同号为{order}的授信时间获取到{len(allCreditTime)}个")
        creditTime = allCreditTime[0]
        df_creditPay = df_all.loc[
            (df_all["下单合同号"] == order) & (df_all["付款日期"].str[:10] == creditTime) & (df_all["开单日期"].isna())]
        if order == "1Y01302205010F":
            print(df_creditPay)
        if df_creditPay.shape[0] > 1:
            b += 1
            print(order)
            continue
            # raise Exception(f"下单费用基础表中合同号为{order}的授信剩余付款获取到{df_creditPay.shape[0]}条")
        elif df_creditPay.shape[0] == 1:
            creditAmount = df_creditPay.iloc[0]["付款金额"]
            lastCreditDict[order] = [creditTime, creditAmount]
        else:
            # 数据量为0说明授信额度被使用完
            lastCreditDict[order] = [creditTime, 0]
    print(lastCreditDict)
    print(lastCreditDict["1Y01302205010F"])
    print(a, b)
    """---------------------测试---------------------"""

    # 筛选出匹配完成，无需操作的数据
    df_base = df_all.query(
        "(~付款日期.isna() and 付款日期 != '未付款' and ~开单日期.isna()) or 付款金额 =='0' or 备注 == '使用激励'")
    # 筛选出仅有开单信息无付款信息的数据（未付款数据）
    df_nopay = df_all.query("付款日期 == '未付款'")
    # 筛选出仅有付款信息无开单信息的数据（剩余付款）
    df_noUsePay = df_all.query("~付款日期.isna() and 开单日期.isna() and 付款金额 !='0' and 备注 != '使用激励'")
    # 筛选出无开单日期和付款日期的数据，如有说明源数据有异常
    df_error = df_all.query("付款日期.isna() and 开单日期.isna()")
    if not df_error.empty:
        raise Exception(f"下单费用表存在即无付款日期也无开单日期的数据{df_error.shape[0]}条，请检查")
    # 若获取的三种数据量和 ！= 总数据量，抛出异常查询原因
    if len(df_all) != len(df_base) + len(df_nopay) + len(df_noUsePay):
        df_dump = df_base.append(df_nopay).append(df_noUsePay)
        df_dump = df_dump[df_dump.duplicated()]
        saveDumpPath = os.path.join(os.path.dirname(orderCostPath), "下单费用基础表异常数据参考.xlsx")
        df_dump.to_excel(saveDumpPath, index=False)
        raise Exception(
            f"下单费用表获取的数据量总和不等于总数据量，请检查筛选规则及表内容格式，可参考{saveDumpPath}进行排查")

    # 本次需要计算的下单数据 = “毛利分析结果表”中需要计算的数据 + "下单费用表"中标记为未付款的数据：
    # fixme：新增下单数据的筛选
    df_analyze = filterOrderInfo(df_analyze, lastOrderDateDict)
    df_analyze = df_analyze.append(df_nopay).reset_index(drop=True)
    df_analyze["开单金额"] = pd.to_numeric(df_analyze["开单金额"])
    df_analyze["实际税率"] = pd.to_numeric(df_analyze["实际税率"])

    # 1.对剩余付款的数据处理（可能存在某合同号，其付款表中最新的扣款时间>付款时间，且数据汇总后有剩余付款，导致剩余付款的日期并非最新日期）
    # 2.将“最新付款日期”有值的数据筛选，更新lastPayDateDict（每个订单号的付款最新日期）
    df_updatePay = df_all.loc[~df_all["最新付款日期"].isna()]
    # 对筛选的数据透视，获取每个订单号的实际最新日期
    df_updatePay = df_updatePay.sort_values(by=["最新付款日期"], ascending=False)
    if not df_updatePay.empty:
        df_pivot = df_updatePay.pivot_table(index="下单合同号", values=["最新付款日期"],
                                            aggfunc={"最新付款日期": "first"})
        lastPayDateDict.update(dict(zip(df_pivot.index, df_pivot["最新付款日期"])))

    return {"下单df": df_analyze, "已处理df": df_base, "剩余付款df": df_noUsePay, "最新付款时间dict": lastPayDateDict}


# 初始化需要下载“里程碑付款&调整台帐”表的合同号字典及授信字典
def initDownLoadOrder(df_analyze):
    """
    :param df_analyze: 需要处理的下单df
    :param allFieldfpList:订单全字段报表列表
    :return:字典：{"华为原厂df": 华为原厂df, "需要下载的合同号Dict": {公司：[下单合同号]}}, "授信字典": {下单合同号: [订单激活时间, 订单变更前金额]}}
    """
    # 筛选出服务、原厂下单的数据（需为1Y号）
    df_origin = df_analyze.loc[
        (df_analyze["采购类型"].isin(['服务', '原厂下单'])) & (df_analyze["下单合同号"].str.startswith("1Y"))]

    # 获取需要处理的数据中，需要下载“里程碑付款&调整台帐”及“使用激励金额”的数据
    AllQryOrderSet = set(df_origin["下单合同号"])
    initDownloadOrderList = list(AllQryOrderSet)

    # 将需要下载的下单合同号按照“供应商名称”分类
    needDownloadDict = {}
    for simpleName, fullName in companySimpleDict.items():
        companyOrder = df_origin.query("供应商名称 == @fullName and 下单合同号.isin(@initDownloadOrderList)")[
            "下单合同号"].tolist()
        needDownloadDict[simpleName] = list(set(companyOrder))

    return {"华为原厂df": df_origin, "需要下载的合同号Dict": needDownloadDict}


# 校验授信付款外挂表数据完整性，若存在授信的合同但外挂表中无对应数据，将这些数据保存并抛出异常
def validCreditData(incentiveDict, creditPayTable, saveDir):
    """
    :param incentiveDict: 使用激励/授信字典{下单合同号: [使用激励金额, 是否使用授信]}
    :param creditPayTable: 授信付款外挂表（华为原厂数据）
    :param saveDir: 结果文件保存目录
    :return: 华为原厂授信数据df
    """
    # 获取本次处理的数据中标记为授信的合同号列表
    allCreditOrder = []
    for orderNum, value in incentiveDict.items():
        if value[1] == "Y":
            allCreditOrder.append(orderNum)

    # 读取授信付款外挂表（华为原厂数据），判断是否有授信数据不在外挂表中
    df_credit = pd.read_excel(creditPayTable, dtype=str)
    existOrder = set(df_credit["合同号"])
    noExistOrder = []
    for orderNum in allCreditOrder:
        if orderNum not in existOrder:
            noExistOrder.append(orderNum)

    # 看是否有需要处理的授信数据但授信付款外挂表无该数据
    recordCreditPath = os.path.join(saveDir, "无授信付款合同号记录表.xlsx")
    if len(noExistOrder) != 0:
        se = pd.Series(data=noExistOrder, name="合同号")
        se.to_excel(recordCreditPath, index=False)
        raise Exception("存在无授信付款的合同号，已保存在结果目录中")
    else:
        if os.path.exists(recordCreditPath):
            os.remove(recordCreditPath)

    df_credit = df_credit.sort_values(by="付款时间", ascending=True)
    df_credit.rename(columns=creditPayTableRenameDict, inplace=True)

    return df_credit


# 计算华为原厂数据的匹配结果
def calDataStep_YC(df_origin, df_credit, incentiveDict, orderFileDict, df_noUsePay, HWOrderPathList, lastPayDateDict):
    """
    :param df_origin: 华为原厂df
    :param df_credit: 华为原厂授信数据df
    :param incentiveDict: 使用激励字典{下单合同号: 使用激励金额}
    :param orderFileDict: 下载的“里程碑付款&调整台帐表”路径列表
    :param df_noUsePay: 下单费用表中的剩余付款df
    :param HWOrderPathList: 华为订单表路径列表
    :param lastPayDateDict: 已有的最新付款时间字典{下单合同号: 付款时间}
    :return: 华为原厂df结果数据，华为原厂数据的下单合同号列表
    """

    # 先获取授信字典，初始化"备注"
    creditDf_temp = df_credit.drop_duplicates("下单合同号", keep="first", ignore_index=True)
    creditDict = dict(zip(creditDf_temp["下单合同号"], creditDf_temp[["付款日期", "付款金额"]].values.tolist()))
    df_origin["备注"] = df_origin.apply(setRemark, args=(creditDict,), axis=1)

    df_originFinal = pd.DataFrame()  # 华为原厂数据处理后的df

    # # 先处理授信数据
    # df_origin_sx = df_origin.loc[df_origin["备注"] == "使用授信"].copy()
    # if not df_origin_sx.empty:
    #     df_origin_sx[["付款日期", "付款金额"]] = df_origin_sx["下单合同号"].apply(lambda x: pd.Series(creditDict.get(x, ["", ""])))
    #     df_originFinal = df_originFinal.append(df_origin_sx)

    # 处理数据
    allOrder = set(df_origin["下单合同号"])
    for orderNum in allOrder:
        # 读取付款表
        df_payO = pd.read_excel(orderFileDict[orderNum], header=8, dtype=str)[payTableRenameDict.keys()]
        df_payO.rename(columns=payTableRenameDict, inplace=True)
        df_payO["付款日期"] = df_payO["付款日期"].fillna(method="ffill")
        # 获取该合同号的授信数据
        df_orderCredit = df_credit.loc[df_credit["下单合同号"] == orderNum]
        # 下单数据匹配付款数据
        matchResultDf = matchPayDetail(orderNum, df_origin, df_payO, df_noUsePay, df_orderCredit, creditDict,
                                       payTableRenameDict, incentiveDict, lastPayDateDict, flag="华为原厂")
        df_originFinal = df_originFinal.append(matchResultDf)

    # 对匹配到付款的下单数据（开单日期、付款日期有值）计算“下单费用等字段”
    df_temp = df_originFinal.loc[
        (~df_originFinal["开单日期"].isna()) & (~df_originFinal["付款日期"].isna()) & (
                    df_originFinal["付款日期"] != "未付款")].copy()
    df_temp = calDataStep2(df_temp, HWOrderPathList, "", flag="华为原厂")
    # 将df_temp的数据同步到df_originFinal中
    df_originFinal.loc[(~df_originFinal["开单日期"].isna()) & (~df_originFinal["付款日期"].isna()) & (
            df_originFinal["付款日期"] != "未付款")] = df_temp

    # 清理内存
    gc.collect()
    return df_originFinal, list(allOrder)


# 匹配"开单金额（含税）"、"付款天数差"、"下单费用"等字段
def calDataStep2(df_temp, HWOrderPathList, analyzePath, flag):
    """
    :param df_temp: 需要匹配"下单费用"等字段的df
    :param HWOrderPathList: 华为订单表路径列表
    :param analyzePath: 毛利分析结果表路径
    :param flag: 处理的数据类型（华为原厂、鲲泰、超聚变）
    :return: 匹配后的df
    """
    # 计算“开单金额（含税）”列
    df_temp["开单金额（含税）"] = df_temp.apply(lambda x: new_round(x["开单金额"] * (1 + x["实际税率"])), axis=1)
    # 计算“核查”列
    df_temp["付款金额"] = pd.to_numeric(df_temp["付款金额"])
    df_temp["核查"] = df_temp["开单金额（含税）"] - df_temp["付款金额"]
    # 计算"付款天数差"
    df_temp["付款天数差"] = (pd.to_datetime(df_temp["开单日期"].astype(str).str[:10]) - pd.to_datetime(
        df_temp["付款日期"].astype(str).str[:10])).dt.days

    if flag == "华为原厂":
        # 超聚变数据传入的订单表路径字典为[]，“运输方式”从外挂表中去而不是华为订单表
        # 匹配"运输方式"
        df_HWOrder = pd.DataFrame()
        for path in HWOrderPathList:
            orderTempDf = pd.read_excel(path, dtype=str).fillna("")
            df_HWOrder = df_HWOrder.append(orderTempDf)
        HWOrderDict = dict(zip(df_HWOrder["华为订单号"], df_HWOrder["运输方式"]))
        # todo: 未找到的运输方式默认值
        df_temp["运输方式"] = df_temp["下单合同号"].apply(lambda x: HWOrderDict.get(x, "自提"))
    elif flag == "超聚变":
        df_saleDetail = pd.read_excel(analyzePath, sheet_name="销售明细", dtype=str)
        transportDict = dict(zip(df_saleDetail["下单合同号"], df_saleDetail["运输方式"]))
        # fixme: 实际上需要匹配运输方式的下单合同号就来源于销售明细中的数据，所以当期数据不会存在未匹配到的情况
        df_temp["运输方式"] = df_temp["下单合同号"].apply(lambda x: transportDict.get(x, "自提"))

    # 匹配"贷款利率"、"下单费用"、"扣款时间"、"扣款月份"
    df_temp["贷款利率"] = df_temp.apply(matchRate, args=(flag,), axis=1)
    df_temp["下单费用"] = df_temp.apply(matchCost, args=(flag,), axis=1)
    df_temp["扣款时间"] = df_temp["开单日期"].apply(matchDeductTime)
    df_temp["扣款月份"] = df_temp["开单日期"].apply(matchDeductMonth)

    return df_temp


# 计算鲲泰数据的匹配结果
def calDataStep_KT(df_analyze):
    """
    :param df_analyze: 需要处理的下单df
    :return: 鲲泰df结果数据
    """
    # 筛选出“鲲泰”的数据
    df_KT = df_analyze.loc[df_analyze["采购类型"].isin(['鲲泰'])].copy()

    # 计算"备注"、"收据编号"、"使用激励金额"、"付款日期"、"付款金额"、"运输方式"
    df_KT["备注"] = ""
    df_KT["收据编号"] = ""
    df_KT["使用激励金额"] = ""
    df_KT["付款日期"] = df_KT["开单日期"].apply(setPayTimeKT)
    df_KT["付款金额"] = df_KT.apply(lambda x: new_round(x["开单金额"] * (1 + x["实际税率"])), axis=1)
    df_KT["运输方式"] = "汽运"

    # 对所有鲲泰数据计算“下单费用等字段”
    df_KT = calDataStep2(df_KT, [], "", "鲲泰")

    return df_KT

# 处理鲲泰运输方式
def transport_KT(df_ktFinal, ktOrderPath):
    df_saleDetail = pd.read_excel(ktOrderPath, dtype=str)
    transportDict = dict(zip(df_saleDetail["供货方编号"], df_saleDetail["运输方式"]))
    df_ktFinal["运输方式"] = df_ktFinal["下单合同号"].apply(lambda x: transportDict.get(x, "汽运"))
    return df_ktFinal

# 计算超聚变数据的匹配结果
def calDataStep_CJB(df_analyze, df_noUsePay, cjbPayDetailPath, lastPayDateDict, analyzePath):
    """
    :param df_analyze: 需要处理的下单df
    :param df_noUsePay: 下单费用表中的剩余付款df
    :param cjbPayDetailPath: 超聚变付款明细外挂表
    :param lastPayDateDict: 已有的最新付款时间字典{下单合同号: 付款时间}
    :param analyzePath: 毛利分析结果表路径
    :return: 超聚变df结果数据，超聚变数据的下单合同号列表
    """
    # 筛选出“超聚变”的数据
    df_CJB = df_analyze.loc[df_analyze["采购类型"].isin(['超聚变'])].copy()
    # 读取“超聚变付款明细外挂表”，并筛选出“付款时间”有效的数据
    df_cjbAllPay = pd.read_excel(cjbPayDetailPath, dtype=str)
    df_cjbAllPay["付款时间"] = df_cjbAllPay.apply(
        lambda x: x["更改授信时间"] if "授信" in x["付款时间"].strip() else x["付款时间"], axis=1)
    # df_cjbAllPay = df_cjbAllPay[df_cjbAllPay["付款时间"].apply(filterValidPayInfoCJB)]
    # 获取授信字典
    creditDf = df_cjbAllPay.loc[~df_cjbAllPay["更改授信时间"].isna()]
    creditDictCJB = dict(zip(creditDf["华为合同号"], creditDf[["更改授信时间", "付款金额"]].fillna(0).values.tolist()))
    # 获取激励字典
    df_cjbAllPay["激励金额"] = pd.to_numeric(df_cjbAllPay["激励金额"], errors="coerce")
    incentiveDf = df_cjbAllPay.loc[~df_cjbAllPay["激励金额"].isna()].sort_values(by="付款时间", ascending=True)
    incentiveDictCJB = dict(zip(incentiveDf["华为合同号"], incentiveDf["激励金额"]))

    # 初始化"备注"
    df_CJB["备注"] = df_CJB.apply(setRemark, args=(creditDictCJB,), axis=1)
    df_cjbFinal = pd.DataFrame()  # 超聚变数据处理后的df

    # # 先处理授信数据
    # df_sjb_sx = df_CJB.loc[df_CJB["备注"] == "使用授信"].copy()
    # if not df_sjb_sx.empty:
    #     df_sjb_sx[["付款日期", "付款金额"]] = df_sjb_sx["下单合同号"].apply(lambda x: pd.Series(creditDictCJB.get(x, ["", ""])))
    #     df_cjbFinal = df_cjbFinal.append(df_sjb_sx)

    # 处理数据
    allOrder = set(df_CJB["下单合同号"])
    df_cjbAllPay = df_cjbAllPay[payTableRenameDictCJB.keys()].rename(columns=payTableRenameDictCJB)
    for orderNum in allOrder:
        # 筛选出该下单合同号的付款数据
        df_payO = df_cjbAllPay.loc[(df_cjbAllPay["下单合同号"] == orderNum)]
        # 获取该合同号的授信数据（授信数据需要保证不含授信前付款）
        df_orderCredit = df_cjbAllPay.loc[(df_cjbAllPay["下单合同号"] == orderNum) & (
                df_cjbAllPay["付款日期"].str[:10] >= creditDictCJB.get(orderNum, ["9999-12-31", ""])[0][:10])]
        # 下单数据匹配付款数据
        matchResultDf = matchPayDetail(orderNum, df_CJB, df_payO, df_noUsePay, df_orderCredit, creditDictCJB,
                                       payTableRenameDictCJB, incentiveDictCJB, lastPayDateDict, flag="超聚变")
        df_cjbFinal = df_cjbFinal.append(matchResultDf)

    # 对匹配到付款的下单数据（开单日期、付款日期有值）计算“下单费用等字段”
    df_temp = df_cjbFinal.loc[
        (~df_cjbFinal["开单日期"].isna()) & (~df_cjbFinal["付款日期"].isna()) & (
                    df_cjbFinal["付款日期"] != "未付款")].copy()
    df_temp = calDataStep2(df_temp, [], analyzePath, flag="超聚变")
    # 将df_temp的数据同步到df_cjbFinal中
    df_cjbFinal.loc[
        (~df_cjbFinal["开单日期"].isna()) & (~df_cjbFinal["付款日期"].isna()) & (
                    df_cjbFinal["付款日期"] != "未付款")] = df_temp

    # 清理内存
    gc.collect()
    return df_cjbFinal, list(allOrder)


# 合并基础数据和处理的数据
def finishOperateAndSave(df_base, df_originFinal, df_ktFinal, df_cjbFinal, df_noUsePay, orderList_YC, orderList_CJB,
                         saveDir, analyzePath, orderCostPath):
    """
    :param df_base: 下单费用表中已处理df
    :param df_originFinal: 华为原厂df结果数据
    :param df_ktFinal: 鲲泰df结果数据
    :param df_cjbFinal: 超聚变df结果数据
    :param df_noUsePay: 下单费用表中的剩余付款df
    :param orderList_YC: 华为原厂数据的下单合同号列表
    :param orderList_CJB: 超聚变数据的下单合同号列表
    :param saveDir: 结果文件保存目录
    :param analyzePath: 毛利分析结果表路径
    :param orderCostPath: 下单费用表路径
    :return: 下单费用结果表保存路径
    """
    # 可能存在部分合同号，本次无下单数据但有剩余付款，需要将这部分添加到原数据中
    alreadyMatchOrder = orderList_YC + orderList_CJB
    df_noUsePay.drop(df_noUsePay[df_noUsePay["下单合同号"].isin(alreadyMatchOrder)].index, inplace=True)

    # 合并数据
    finialDf = pd.concat([df_base, df_originFinal, df_ktFinal, df_cjbFinal, df_noUsePay]).reset_index(drop=True)
    finialDf["使用激励金额"] = pd.to_numeric(finialDf["使用激励金额"], errors="coerce")

    # 将使用激励金额的数据筛选出来，同一个下单合同号只保留最新的使用激励金额（将第一次使用激励数据的金额进行更新）
    incentiveOrder = set(finialDf.loc[finialDf["使用激励金额"] > 0, "下单合同号"].tolist())
    for order in incentiveOrder:
        incentiveDf = finialDf.loc[(finialDf["使用激励金额"] > 0) & (finialDf["下单合同号"] == order)]
        if incentiveDf.shape[0] == 1:
            continue
        allIdx = incentiveDf.index
        firstIncentiveIdx = allIdx[0]
        lastIncentiveAmont = incentiveDf.tail(1).iloc[0]["使用激励金额"]
        for idx in allIdx:
            if idx == firstIncentiveIdx:
                finialDf.loc[idx, "使用激励金额"] = lastIncentiveAmont
            else:
                finialDf.loc[idx, "使用激励金额"] = np.nan
    # 将结果数据按照"下单合同号", "开单日期"排序， 并重新对列排序  --2024-07-11 增加一列“市场类型”--
    finialDf = finialDf.reset_index(drop=True).sort_values(by=["下单合同号", "开单日期"], ascending=[True, True])[
        resultColadd]

    # 将datafram类型的finialDf根据“下单合同号”，“收据编号”，“付款日期”，“付款金额”的值来去除重复的值
    # finialDf = finialDf.drop_duplicates(subset=["下单合同号", "收据编号", "付款日期", "付款金额"])
    # 先根据“下单合同号”，“收据编号”，“付款日期”，“付款金额”的值查出重复的行，再判断"供应商名称", "项目名称"等为空时去除重复，不为空就保留
    duplicated_rows = finialDf[finialDf.duplicated(subset=["下单合同号", "收据编号", "付款日期", "付款金额"], keep='first')]
    for index, row in duplicated_rows.iterrows():
        if row[["供应商名称", "项目名称", "客户名称", "销售员", "销售员编码", "事业部", "区域", "平台"]].isnull().all():
            finialDf.drop(index, inplace=True)

    # 将使用的毛利分析和下单费用基础表保存在标识文件中
    timestamp = int(datetime.now().timestamp())
    resultPathFlag = os.path.join(saveDir, f"下单费用{timestamp}表源文件说明.txt")
    with open(resultPathFlag, "w") as f:
        writeStr = f"使用的毛利分析结果表路径为-{analyzePath}\n使用的下单费用基础表路径为-{orderCostPath}"
        f.write(writeStr)

    # 将结果数据保存在下单费用结果路径
    resultPath = os.path.join(saveDir, f"下单费用{timestamp}.xlsx")
    finialDf.to_excel(resultPath, sheet_name="下单费用", index=False)

    # 清理内存
    gc.collect()
    return resultPath


# 设置下单费用结果表格式
def setStyle(finalPath):
    """
    :param finalPath: 下单费用结果表路径
    :return:
    """

    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finalPath)
    ws = wb.sheets["下单费用"]

    orderCol = get_column_letter(resultCol.index("下单合同号") + 1)
    personCol = get_column_letter(resultCol.index("销售员编码") + 1)
    receiptCol = get_column_letter(resultCol.index("收据编号") + 1)

    # 计算标志列
    colIndex1 = get_column_letter(resultCol.index("平台") + 1)
    colIndex2 = get_column_letter(resultCol.index("收据编号") + 1)
    colIndex3 = get_column_letter(resultCol.index("付款金额") + 1)
    colIndex4 = get_column_letter(resultCol.index("开单日期") + 1)
    colIndex5 = get_column_letter(resultCol.index("开单金额（含税）") + 1)
    colIndex6 = get_column_letter(resultCol.index("核查") + 1)
    colIndex7 = get_column_letter(len(resultCol))

    # 全表格式改成常规格式、表头字体为微软雅黑9号、内容字体为微软雅黑9号
    ws.range(f"A:{colIndex7}").number_format = "G/通用格式"
    ws.range(f"A:{colIndex7}").font.size = 9
    ws.range(f"A:{colIndex7}").font.name = "微软雅黑"
    # "下单合同号"、"销售员编码"、"收据编号"列为文本格式
    for col in [orderCol, personCol, receiptCol]:
        ws.range(f"{col}:{col}").number_format = "@"

    # plan A：已经用本文格式存储的数字，需要进行分列操作,但会导致部分列数据损失
    # colIndexList = get_column_interval(1, cols)
    # for indexCol in colIndexList:
    #     ws.range(indexCol+"1").expand("down").api.TextToColumns()

    # plan B:由于单元格格式已经为常规，将所有数据复制重新粘贴可去除文本格式显示的数字（表现为绿三角）
    ws.range("A1").value = ws.used_range.value

    # 进行冻结操作(F2)
    active_window = wb.app.api.ActiveWindow
    active_window.FreezePanes = False
    # wb.app.range("A2").select()  # 选"A2"冻结首行
    # active_window.SplitColumn = resultCol.index("客户名称") + 1  # 冻结至哪一列
    active_window.SplitRow = 1  # 冻结至哪一行
    active_window.FreezePanes = True

    # 设置标题背景色、字体颜色
    ws.range(f"A1:{colIndex1}1").color = (180, 198, 231)
    ws.range(f"{colIndex2}1:{colIndex3}1").color = (198, 224, 180)
    ws.range(f"{colIndex4}1:{colIndex5}1").color = (255, 192, 0)
    ws.range(f"{colIndex6}1:{colIndex7}1").color = (180, 198, 231)
    ws.range("A1:%s1" % colIndex7).font.color = (0, 0, 0)
    ws.range("A1:%s1" % colIndex7).font.bold = False

    # 新增数据设置格式, "实际税率"保留两位小数，其他数字列为千分位显示整数
    for col in ["付款日期", "付款金额", "开单日期", "开单金额", "实际税率", "开单金额（含税）", "核查", "使用激励金额",
                "贷款利率", "下单费用"]:
        index = get_column_letter(resultCol.index(col) + 1)
        if col in ["付款金额", "开单金额", "开单金额（含税）"]:
            # 千分位显示整数
            ws.range(f"{index}:{index}").number_format = "#,##0_ "
        elif col in ["贷款利率"]:
            # 百分比显示并保留两位小数
            ws.range(f"{index}:{index}").number_format = "0.00%"
        elif col in ["实际税率"]:
            # 保留两位小数
            ws.range(f"{index}:{index}").number_format = '0.00'
        elif col in ["核查", "下单费用", "使用激励金额"]:
            # 千分位显示整数，且负数显示为红色+括号
            ws.range(f"{index}:{index}").number_format = "#,##0_);[红色](#,##0)"
        elif col in ["付款日期", "开单日期"]:
            # 日期只显示年月日：yyyy/mm/dd
            ws.range(f"{index}:{index}").number_format = "yyyy/m/d"

    # 自适应宽度
    # ws.autofit()
    # 设置列宽，先将所有列列宽设置为10，"下单合同号","项目名称","客户名称"列宽为15
    ws.used_range.column_width = 10
    for colName in ["下单合同号", "项目名称", "客户名称"]:
        col = get_column_letter(resultCol.index(colName) + 1)
        ws.range(f"{col}:{col}").column_width = 15

    # "采购类型"、"最新付款日期"列隐藏
    buyTypeCol = get_column_letter(resultCol.index("采购类型") + 1)
    newlyDateCol = get_column_letter(resultCol.index("最新付款日期") + 1)
    ws.range(f"{newlyDateCol}:{newlyDateCol}").api.EntireColumn.Hidden = True

    wb.save(finalPath)
    wb.close()
    app.quit()


"""
companySimpleDict:供应商名称字典
filterCol:毛利分析结果表中需要筛选的列名列表
renameColDict: 毛利分析结果表中需要重命名的列名字典
payTableRenameDict：华为原厂数据的付款表中需要重命名的列名字典(同时key为需要操作的列名)
creditPayTableRenameDict：华为原厂数据的授信付款外挂表中需要重命名的列名字典(同时key为需要操作的列名)
payTableRenameDictCJB：超聚变付款外挂表中需要重命名的列名字典(同时key为需要操作的列名)
resultCol：下单费用结果表的列名列表
matchFlag：用于防止df.apply对第一条数据重复操作
logger：用于打印日志
"""
companySimpleDict = {"城投": "广州城投信息科技有限公司", "合神": "合肥神州数码有限公司", "北神": "北京神州数码有限公司"}
filterCol = ["备注", "下单合同号", "项目名称", "客户名称", "销售员", "销售员编码", "事业部", "区域", "平台",
             "出具发票日", "成本总价", "实际税率", "产品", "产品线"]
renameColDict = {"备注": "供应商名称", "出具发票日": "开单日期", "成本总价": "开单金额", "产品": "产品类别"}
payTableRenameDict = {"华为合同号": "下单合同号", "处理日期": "付款日期", "收据调整金额": "付款金额",
                      "对应收据": "收据编号"}
creditPayTableRenameDict = {"合同号": "下单合同号", "付款时间": "付款日期", "付款金额": "付款金额"}
payTableRenameDictCJB = {"华为合同号": "下单合同号", "付款时间": "付款日期", "付款金额": "付款金额"}
resultCol = ['供应商名称', '下单合同号', '项目名称', '客户名称', '销售员', '销售员编码', '事业部', '区域', '平台',
             '收据编号', '付款日期', '付款金额', '开单日期',
             '开单金额', '实际税率', '开单金额（含税）', '核查', '备注', '使用激励金额', '付款天数差', '贷款利率',
             '下单费用', '扣款时间', '扣款月份', '运输方式', '产品类别',
             '产品线', '特殊备注', '采购类型', "最新付款日期"]
resultColadd = resultCol + ["市场类型"]
matchFlag = False
logger = None

if __name__ == "__main__":
    # g_orderCostPath = r"E:\Uibot项目测试\下单费用流程\20230303下单费用.xlsx"
    # g_analyzePath = r"E:\Uibot项目测试\下单费用流程\毛利核算.xlsx"
    # g_dictGlobal = {"文件下载路径": r"E:\Uibot项目测试\下单费用流程",
    #                 "结果保存路径": r"E:\Uibot项目测试\下单费用流程",
    #                 "超聚变付款外挂表路径": r"E:\Uibot项目文件\配置表\超聚变付款表2023.xlsx",
    #                 "授信付款外挂表路径": r"E:\Uibot项目文件\配置表\授信付款外挂表.xlsx",
    #                 }
    g_orderCostPath = r"D:\xc_files\下单\20240903\结果文件\下单费用1724987780.xlsx"
    g_analyzePath = r"D:\xc_files\下单\20240903\结果文件\毛利核算.xlsx"
    g_dictGlobal = {"文件下载路径": r"D:\xc_files\下单\20240903",
                    "结果保存路径": r"D:\xc_files\下单",
                    "超聚变付款外挂表路径": r"D:\xc_files\下单\Uibot项目测试\配置表\超聚变付款表2023.xlsx",
                    "授信付款外挂表路径": r"D:\xc_files\下单\Uibot项目测试\配置表\授信付款外挂表.xlsx",
                    "鲲泰订单跟踪表路径": r"D:\xc_files\交接\zhangweno需求\华为下单费用流程\鲲泰订单跟踪表.xlsx",
                    }
    g_dictGlobal["incentiveRecordPath"] = getIncentiveRecordPath(g_dictGlobal["文件下载路径"])
    print(g_dictGlobal["incentiveRecordPath"])
    list_ = getSameFormatFile(g_dictGlobal["文件下载路径"] + "\\里程碑付款&调整台帐表", "里程碑付款&调整台帐表")
    orderFileDict = {}
    for i in list_:
        k = os.path.basename(i).split("_")[1]
        orderFileDict[k] = i
    print(len(orderFileDict))
    orderFileList = getSameFormatFile(g_dictGlobal["文件下载路径"], "订单表")
    print(len(orderFileList))

    g_analyzePath = initAnalyzeNoteText(orderFileList, g_analyzePath)

    # 读取下单费用表和毛利分析结果表，返回字典{"下单df": 下单数据df, "已处理df": 已处理的下单数据df, "剩余付款df": 未处理的付款数据df}
    initResultDict = getOriginOrderData(g_analyzePath, g_orderCostPath)

    # 初始化需要下载“里程碑付款 & 调整台帐”表的合同号（授信除外），返回字典{"华为原厂df": 华为原厂df, "需要下载的合同号Dict": {公司：[下单合同号]}}, "授信字典": {下单合同号: [订单激活时间, 订单变更前金额]}}
    initOrderDict = initDownLoadOrder(initResultDict["下单df"])

    g_incentiveDict = readIncentiveRecord(g_dictGlobal["incentiveRecordPath"])

    # 校验授信付款外挂表数据完整性，若存在授信的合同但外挂表中无对应数据，将这些数据保存并抛出异常, 返回：华为原厂授信数据df
    df_credit = validCreditData(g_incentiveDict, g_dictGlobal["授信付款外挂表路径"], g_dictGlobal["结果保存路径"])

    # 处理"华为原厂"数据，返回列表[华为原厂df结果数据，华为原厂数据的下单合同号列表]
    ycResultDict = calDataStep_YC(initOrderDict["华为原厂df"].copy(), df_credit, g_incentiveDict, orderFileDict,
                                  initResultDict["剩余付款df"], orderFileList, initResultDict["最新付款时间dict"])
    df_originFinal = ycResultDict[0]
    orderList_YC = ycResultDict[1]

    # 处理"鲲泰"数据，返回：鲲泰df结果数据
    df_ktFinal = calDataStep_KT(initResultDict["下单df"])
    df_ktFinal = transport_KT(df_ktFinal, g_dictGlobal["鲲泰订单跟踪表路径"])

    # 处理"超聚变"数据，返回列表[超聚变df结果数据，超聚变数据的下单合同号列表]
    cjbResultDict = calDataStep_CJB(initResultDict["下单df"], initResultDict["剩余付款df"],
                                    g_dictGlobal["超聚变付款外挂表路径"],
                                    initResultDict["最新付款时间dict"], g_analyzePath)
    df_cjbFinal = cjbResultDict[0]
    orderList_CJB = cjbResultDict[1]

    # 合并基础数据和处理的数据，返回"下单费用结果表"路径
    resultPath = finishOperateAndSave(initResultDict["已处理df"], df_originFinal, df_ktFinal, df_cjbFinal,
                                      initResultDict["剩余付款df"], orderList_YC, orderList_CJB,
                                      g_dictGlobal["结果保存路径"], g_analyzePath, g_orderCostPath)

    # 设置下单费用结果表格式
    setStyle(resultPath)
