#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import os
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import shutil



# ====== CRM目标表头 → 华为源数据列名 映射表 ======
SOURCE_FIELD_MAPPING = {
    "厂商PO号（必填）":         "华为订单号",
    "产品领域/产品线":           "产品线",
    "最终用户名称":             "最终客户",
    "备注":                   None,
    "纯软件订单标识":           "纯软件订单标识",
    "供应商签约主体":           "华为签约主体",
    "订单状态":                "订单状态",
    "回款日期":                None,
    "跟踪号":                  None,
    "服务金额":                "服务金额",
    "订单版本":                "订单版本",
    "批次":                   None,
    "项目毛利":                None,
    "采购责任人":              None,
    "运输方式":                "运输方式",
    "授信费用":                None,
    "批次数量":                "批次数量",
    "设备金额":                "设备金额",
    "代表处":                  "代表处",
    "开票状态":                "开票状态",
    "采购主体":                "签约经销商",
    "订单类型":                None,
    "下级经销商":              "二级经销商",
    "最晚预计备货完成时间":     "最晚预计备货完成时间",
    "订单出货总金额":           None,
    "物流状态":                "订单状态",
    "商务通知单":              None,
    "付款方式":                None,
    "45采购订单":              None,
    "订单总金额":              "订单总金额",
    "项目名称":                "订单名称",
    "激励使用金额":            "激励使用金额",
    "业务类型名称":            None,
    "地址待定标识":            "地址待定标识",
    "最早预计备货完成时间":     "最早预计备货完成时间",
    "付款状态":                "付款状态",
    "订单激活时间":            "订单激活时间",
    "负责人（必填）":           None,
    "人员-普通成员-只读":        None,
    "人员-普通成员-读写":        None,
    "部门-普通成员-只读":        None,
    "部门-普通成员-读写":        None,
    "用户组-普通成员-只读":      None,
    "用户组-普通成员-读写":      None,
    "角色-普通成员-只读":        None,
    "角色-普通成员-读写":        None,
}


def generateCRMImport(dataFilePaths, templatePath, saveDir):
    """
    合并两个华为数据文件，按模板映射生成CRM导入表。

    :param dataFilePaths: 两个源Excel文件路径列表，
                          例 [r"北京鲲泰_华为数据.xlsx", r"合肥信创_华为数据.xlsx"]
    :param templatePath:  模板文件路径，即"厂商PO号对象更新导入模板.xlsx"
                          （第1行为目标表头，后续行不影响逻辑）
    :param saveDir:       结果保存目录
    :return:              生成的文件完整路径
    """
    # ========== 步骤1：合并两个源数据 ==========
    dfList = []
    for fp in dataFilePaths:
        if not os.path.exists(fp):
            continue
        df = pd.read_excel(fp)
        dfList.append(df)

    if len(dfList) == 0:
        return ""

    mergedDf = pd.concat(dfList, ignore_index=True)

    # 建立源数据列名 → 列索引的快速查找表
    sourceColNameToIdx = {colName: idx for idx, colName in enumerate(mergedDf.columns)}

    # ========== 步骤2：读取模板第1行（目标表头） ==========
    wb = openpyxl.load_workbook(templatePath)

    if '厂商PO号导入模版' not in wb.sheetnames:
        wb.close()
        return ""

    ws = wb['厂商PO号导入模版']

    # 只读取第1行目标表头（生产模板只有此行，辅助模板以此行为准）
    targetHeaders = []
    for c in range(1, ws.max_column + 1):
        targetHeaders.append(ws.cell(1, c).value)

    wb.close()

    # ========== 步骤3：按映射提取数据 ==========
    resultData = {}
    emptyCount = 0
    mappedCount = 0

    for targetHeader in targetHeaders:
        sourceField = SOURCE_FIELD_MAPPING.get(targetHeader)

        if sourceField is None:
            # 无映射或映射表中未找到，填充空值
            resultData[targetHeader] = [None] * len(mergedDf)
            emptyCount += 1
            continue

        # 通过源数据列名匹配
        if sourceField in sourceColNameToIdx:
            colIdx = sourceColNameToIdx[sourceField]
            resultData[targetHeader] = mergedDf.iloc[:, colIdx].values
            mappedCount += 1
        else:
            resultData[targetHeader] = [None] * len(mergedDf)
            emptyCount += 1


    resultDf = pd.DataFrame(resultData, columns=targetHeaders)

    # ========== 步骤4：保存结果 ==========
    if not os.path.exists(saveDir):
        os.makedirs(saveDir)

    outputPath = os.path.join(saveDir, "CRM导入表.xlsx")
    resultDf.to_excel(outputPath, index=False)

    return outputPath



def process_crm_data(main_path, dl_path, fail_result_path):
    """
    模拟 UiBot 脚本的 CRM 数据处理流程
    :param main_path: 模板文件所在目录（通常包含“厂商PO号对象更新导入模板.xlsx”），crm导入流程
    :param dl_path: 华为数据及输出目录，钻石货期crm导入流程
    :param fail_result_path: 失败结果表完整路径（含文件名）
    """
    try:
        # ========== 1. 获取 CRM 数据 ==========
        # crm_template = os.path.join(main_path, "厂商PO号对象更新导入模板_政企.xlsx")
        main_path = main_path.replace("钻石货期CRM导入项目", "CRM导入项目")
        crm_template = os.path.join(main_path, "厂商PO号对象更新导入模板.xlsx")
        crm_df = pd.read_excel(crm_template, sheet_name="厂商PO号导入模版", header=0)
        print("读取政企CRM模板数据完毕")

        # ========== 2. 获取华为数据 ==========
        areas = ["北京鲲泰", "合肥信创"]
        hw_dfs = []
        for info in areas:
            hw_file = os.path.join(dl_path, f"{info}_华为数据.xlsx")
            hw_df = pd.read_excel(hw_file, sheet_name="Contract List Export", header=0)
            hw_dfs.append(hw_df)
            print(f"读取{info}数据完毕")
        hw_df_all = pd.concat(hw_dfs, ignore_index=True)
        print("读取华为数据完毕")

        # ========== 3. 合并数据（右连接） ==========
        merged_df = pd.merge(
            crm_df,
            hw_df_all,
            how='right',
            left_on='厂商PO号（必填）',
            right_on='华为订单号',
            suffixes=('_x', '_y')
        )
        print("CRM与华为数据合并完毕")

        # ========== 4. 处理表头并选择列 ==========
        array_columns = merged_df.columns.tolist()
        header_arr = []
        for col in array_columns:
            if col == "厂商PO号（必填）":
                header_arr.append("华为订单号")
            elif col == "厂商":
                header_arr.append("华为签约主体")
            elif col == "二级经销商（厂商）":
                header_arr.append("二级经销商_y")
            elif col == "二级经销商_x":
                header_arr.append("二级经销商_x")
            elif "_x" in col:          # 注意：二级经销商_x 已被上一分支处理，不会进入这里
                new_col = col.replace("_x", "_y")
                header_arr.append(new_col)
            elif "_y" in col or col in ("华为订单号", "华为签约主体"):
                continue               # 跳过这些列
            else:
                header_arr.append(col)

        print("表头读取完毕")
        # 选择列（若列不存在，pandas 会报错，但 header_arr 来自 merged_df，保证存在）
        selected_df = merged_df[header_arr]

        # ========== 5. 读取失败结果表 ==========
        has_fail_data = False
        po_arr = []
        try:
            if os.path.exists(fail_result_path):
                fail_df = pd.read_excel(fail_result_path, sheet_name="失败结果", header=0)
                if len(fail_df) > 0:
                    # 取 B 列（第二列），去除空值
                    if fail_df.shape[1] >= 2:
                        po_vals = fail_df.iloc[:, 1].dropna().astype(str).str.strip()
                        po_vals = po_vals[po_vals != ""]
                        po_arr = po_vals.tolist()
                        has_fail_data = True if po_arr else False
                    else:
                        has_fail_data = False
                else:
                    has_fail_data = False
            else:
                has_fail_data = False
        except Exception:
            print("失败结果Sheet不存在")
            has_fail_data = False

        # ========== 6. 根据失败PO号筛选数据 ==========
        if not has_fail_data:
            result_df = selected_df.head(0)   # 空表，只保留列名
        else:
            # 检查“厂商PO号（必填）”列是否存在（原代码中该列可能已被替换为“华为订单号”，此处按原逻辑）
            if "华为订单号" in selected_df.columns:
                result_df = selected_df[selected_df["华为订单号"].isin(po_arr)]
            else:
                # 列不存在时，按原逻辑无法筛选，返回空表
                result_df = selected_df.head(0)

        print("数据处理完毕")

        # ========== 7. 写入新文件 ==========
        output_dir = dl_path
        output_file = "CRM_导入失败数据.xlsx"
        output_path = os.path.join(output_dir, output_file)

        # 删除已存在的输出文件
        if os.path.exists(output_path):
            os.remove(output_path)

        # 复制模板文件到输出目录
        # template_path = os.path.join(main_path, "厂商PO号对象更新导入模板_政企.xlsx")
        template_path = os.path.join(main_path, "厂商PO号对象更新导入模板.xlsx")   # 增加使用”CRM导入流程“中的模板
        shutil.copy(template_path, output_dir)
        # 重命名
        copied = os.path.join(output_dir, "厂商PO号对象更新导入模板.xlsx")
        if os.path.exists(copied):
            os.rename(copied, output_path)

        # 用 openpyxl 写入数据（从 A2 开始，表头已存在）
        wb = load_workbook(output_path)
        ws = wb["厂商PO号导入模版"]
        # 将 pandas 的 NaT/NaN/NA 等特殊空值转为 None，否则 openpyxl 写入时会报错：
        # <class 'pandas._libs.tslibs.nattype.NaTType'>
        import numpy as np
        result_df = result_df.astype(object).where(pd.notnull(result_df), None)
        # 处理 numpy 类型的标量，确保 openpyxl 能识别
        def _normalize_val(v):
            if v is None:
                return None
            if isinstance(v, float) and np.isnan(v):
                return None
            if isinstance(v, (pd.Timestamp,)):
                return v.to_pydatetime()
            if v is pd.NaT:
                return None
            # numpy 标量转 Python 原生类型
            if isinstance(v, (np.integer,)):
                return int(v)
            if isinstance(v, (np.floating,)):
                return float(v) if not np.isnan(v) else None
            if isinstance(v, (np.bool_,)):
                return bool(v)
            return v

        data_rows = result_df.values.tolist()
        for i, row in enumerate(data_rows, start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=_normalize_val(val))
        wb.save(output_path)
        wb.close()

        print("写入到新导入文件完毕")
        return output_path

    except Exception as e:
        print("流程CRM数据处理失败")
        print(str(e))
        return str(e)



if __name__ == '__main__':
    # base = r'D:\xc_files\钻石货期CRM导入\项目数据'
    # dataFiles = [
    #     os.path.join(base, "北京鲲泰_华为数据.xlsx"),
    #     os.path.join(base, "合肥信创_华为数据.xlsx"),
    # ]
    # template = r"D:\xc_files\钻石货期CRM导入\厂商PO号对象更新导入模板.xlsx"
    # saveDir = base
    # result = generateCRMImport(dataFiles, template, saveDir)
    # if result:
    #     print(f"\n生成成功: {result}")


    main = r'D:\xc_files\钻石货期CRM导入\crm导入流程'
    dl = r'D:\xc_files\钻石货期CRM导入\项目数据\715'
    fail = r'D:\xc_files\钻石货期CRM导入\项目数据\715\厂商PO号对象导入结果_20260715.xlsx'
    aa = process_crm_data(main, dl, fail)

