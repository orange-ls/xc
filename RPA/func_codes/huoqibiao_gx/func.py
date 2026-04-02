#!/usr/bin/python3
# -*- coding: UTF-8 -*-


import glob
import logging.config
import os
import traceback
from datetime import datetime
from functools import wraps

import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter


# 初始化日志
def initWriteLog(rootDir):
    """
    :param rootDir: 日志写入目录
    :return:
    """
    global logger
    log_name = "销售明细"
    simple_format = '[%(levelname)s][%(asctime)s]%(message)s'
    logfile_dir = rootDir  # log文件的目录
    logfile_name = '%s.log' % log_name  # log文件名
    # 如果不存在定义的日志目录就创建一个
    if not os.path.exists(logfile_dir):
        os.makedirs(logfile_dir)

    # log文件的全路径
    logfile_path = os.path.join(logfile_dir, logfile_name)
    LOGGING_DIC = {
        'version': 1,
        'disable_existing_loggers': False,
        'formatters': {
            'simple': {
                'format': simple_format
            },
        },
        'filters': {},
        'handlers': {
            'default': {
                'level': 'INFO',
                'class': 'logging.handlers.RotatingFileHandler',
                'formatter': 'simple',
                'filename': logfile_path,
                'maxBytes': 1024 * 1024 * 5,
                'backupCount': 5,
                'encoding': 'utf-8',
            },
        },
        'loggers': {
            '': {
                'handlers': ['default'],
                'level': 'INFO',
                'propagate': True,
            },
        },
    }
    logging.config.dictConfig(LOGGING_DIC)
    logger = logging.getLogger(__name__)


# 定义一个日志装饰器，每次调用方法前后打印日志
# 注 uibot中无法调用，弃用
def logfun(func):
    @wraps(func)
    def logfunStep(*args, **kwargs):
        logger.info(f"{'-' * 25}{func.__name__}开始{'-' * 25}")
        result = func(*args, **kwargs)
        logger.info(f"{'-' * 25}{func.__name__}结束{'-' * 25}")
        return result

    return logfunStep


# 读取配置文件生成配置字典
def getConfigDict(baseConfPath):
    """
    :param baseConfPath：配置文件目录
    :return: 生成配置字典
    """
    dfDict = pd.read_excel(baseConfPath, dtype=str, sheet_name=None)
    resultDict = {}
    for sheetName, df in dfDict.items():
        df = df.fillna("")
        resultDict.update(dict(zip(df["配置名称"], df["配置内容"])))
    return resultDict


# 根据输入的文件根目录返回当日文件保存目录
def getSaveDir(rootDir):
    """
    :param rootDir：文件保存根目录，例E:\downDir
    :return: 返回根目录+当日日期的目录，如E:\downDir\2022年\6月\8日
    """
    today = datetime.now()
    timeDict = {}
    timeDict["year"], timeDict["month"], timeDict["day"], timeDict[
        "hour"] = today.year, today.month, today.day, today.hour
    for key, value in timeDict.items():
        timeDict[key] = str(value).zfill(2)
    finialDir = os.path.join(rootDir, "{year}年\\{month}月\\{day}日\\{hour}时".format(**timeDict))
    return finialDir


# 获取某文件夹下（包括子文件夹）中最新的包含某个文件名的文件列表
def getLatestFile(rootDir, searchKey):
    """
    :param rootDir: 查找文件夹
    :param searchKey: 文件名关键字
    :return: 查找到的文件列表
    """
    lastestDir = ""
    # 遍历文件夹及子文件夹，若文件夹日期最新且文件夹下有符合条件的文件名，则更新文件夹
    for root, dirs, files in os.walk(rootDir):
        if root > lastestDir and len(glob.glob(root + f"\\{searchKey}")) > 0:
            lastestDir = root
    # 将最终定位的文件夹下符合要求的文件进行返回
    filePathList = []
    for path in glob.glob(lastestDir + f"\\{searchKey}"):
        filePathList.append(path)
    return filePathList

# 更新授信付款外挂表
def updateCreditTable(CreditTable, updatetable):
    """
    :param CreditTable: 授信付款外挂表文件路径
    :param updatetable: 待更新的授信付款外挂表数据路径
    :return: 更新后的授信付款外挂表文件路径
    """
    # 读取授信外挂表CreditTable，并筛选出“付款方式”为“授信”的数据
    credit_table = pd.read_excel(CreditTable, dtype=str)
    credit_data = credit_table[credit_table['付款方式'] == '授信']
    # 读取待更新数据表中的付款信息工作表
    update_table = pd.read_excel(updatetable, sheet_name="付款信息", dtype=str)
    # 筛选出“付款时间”为“授信”的数据
    update_data = update_table[update_table['付款时间'] == '授信']
    # 选取合同号、付款时间、付款金额、付款方式变更日期
    update_data = update_data[["合同号", "付款方式变更日期", "付款金额", "付款时间"]]

    # 合并数据,如果合同号在CreditTable中存在，则更新付款时间，否则新增数据
    merged_data = pd.merge(credit_data, update_data, on='合同号', how='outer', suffixes=('_old', '_new'))

    # 判断谁的付款时间更晚，按照晚的更新CreditTable中的数据
    merged_data['付款时间'] = merged_data.apply(
        lambda row: row['付款时间_old'] if pd.isnull(row['付款方式变更日期']) else
        (row['付款方式变更日期'] if pd.isnull(row['付款金额_old']) else
        max(row['付款方式变更日期'], row['付款时间_old'])), axis=1)

    merged_data['付款金额'] = merged_data.apply(
        lambda row: row['付款金额_old'] if pd.isnull(row['付款方式变更日期']) else
        (row['付款金额_new'] if pd.isnull(row['付款金额_old']) else
        (row['付款金额_new'] if row['付款方式变更日期'] > row['付款时间_old'] else row['付款金额_old'])), axis=1)

    # 将merged_data['付款方式']更新为'授信'
    merged_data['付款方式'] = '授信'

    # 选择需要的列
    final_data = merged_data[['合同号', '付款时间', '付款金额', '付款方式']]
    final_data = final_data.drop_duplicates(subset=['合同号', '付款时间', '付款金额', '付款方式'], keep='last')

    # 保存更新后的授信付款外挂表
    final_data.to_excel(CreditTable, index=False)
    return CreditTable


# 生成货期表
def generateDeliveryTable(crmFile, huaweiFile, savePath, updateUrl, contractPaymentFile, CreditTable):
    """
    :param crmFile: CRM系统下载的厂商PO号表
    :param huaweiFile: CRM导入流程下载的华为订单表列表（一般有合神、北神、城投三个文件）
    :param savePath: 保存的货期表路径
    :param updateUrl: 更新货期表数据到DB的post地址
    :param contractPaymentFile: 合同回款表文件路径
    :param CreditTable: 授信付款外挂表文件路径
    :return: 生成按次post的数据列表
    """
    try:
        # 1.读取CRM下载的厂商PO号表，生成字典{PO号：[45采购订单，批次， 采购组织]}
        df_crm = pd.read_excel(crmFile, dtype=str)
        matchDict = dict(zip(df_crm["厂商PO号（必填）"], df_crm[["45采购订单", "批次", "采购组织"]].values))

        # 2.分别合并华为订单表数据和华为合同回款表数据
        # 2.1 合并华为订单表数据
        df_hw = pd.DataFrame()
        for path in huaweiFile:
            df_temp = pd.read_excel(path, dtype=str)
            df_hw = df_hw.append(df_temp)

        # 2.2 合并华为合同回款表数据
        df_pay = pd.DataFrame()
        for path in contractPaymentFile:
            df_temp = pd.read_excel(path, header=8,dtype=str)
            df_pay = df_pay.append(df_temp)
        df_pay[["华为合同号", "处理日期"]] = df_pay[["华为合同号", "处理日期"]].fillna(method="ffill")

        # --20240722-- 修改：在华为订单表中找出授信订单，将这些订单的付款时间按外挂表更新
        # 在df_pay中筛选出“是否用于归还授信”列为Y的行
        credit_pay_df = df_pay[df_pay["是否用于归还授信"] == "Y"]
        credit_pay_df = credit_pay_df[["华为合同号"]]

        # 读取 授信外挂表CreditTable
        credit_table = pd.read_excel(CreditTable, dtype=str)
        # 筛选出“付款方式”为“授信”的数据
        credit_data = credit_table[credit_table['付款方式'] == '授信']

        matchPayDateDict = dict(zip(df_pay["华为合同号"], df_pay["处理日期"]))
        # 在matchPayDateDict里找出在credit_pay_df中存在的合同号，将这些合同号的“处理日期”按照外挂表credit_data中的“付款时间”更新
        for key, value in matchPayDateDict.items():
            if key in credit_pay_df.values:
                if key in credit_data['合同号'].values:
                    matchPayDateDict[key] = credit_data[credit_data['合同号'] == key]['付款时间'].values[0]

        # 3.依据"华为订单号"匹配"45采购订单", "批次", "采购组织", "经销商付款日期"，列只取需要的列DeliveryTableCol
        df_hw[["45采购订单", "批次", "采购组织"]] = df_hw["华为订单号"].apply(lambda x: pd.Series(matchDict.get(x, ["", "", ""])))
        df_hw["经销商付款日期"] = df_hw["华为订单号"].apply(lambda x: matchPayDateDict.get(x, ""))

        df_hw = df_hw[DeliveryTableCol].fillna("")
        # 4.对所有的时间/日期列，只保留年月日（2023-06-16），且空值替换为"-"
        for col in DeliveryTableCol:
            if "时间" in col or "日期" in col:
                df_hw[col] = df_hw[col].str[:10]
                df_hw[col] = df_hw[col].apply(lambda x: "-" if x == "" else x)

        # 5. 新建文件，写入数据
        wb = xw.Book()
        ws = wb.sheets["Sheet1"]

        finalCol = get_column_letter(len(DeliveryTableCol))
        s_col = get_column_letter(DeliveryTableCol.index("45采购订单") + 1)
        e_col = get_column_letter(DeliveryTableCol.index("采购组织") + 1)

        # 写入表头
        ws.range("A1").value = DeliveryTableCol

        # 第一行内容水平居中
        ws.range(f'A1:{finalCol}1').api.HorizontalAlignment = -4108

        # 全表格式改成文本格式、字体为微软雅黑10号
        # ws.range(f"A:{colIndex3}").number_format = "G/通用格式"
        ws.range(f"A:{finalCol}").number_format = "@"
        ws.range(f"A:{finalCol}").font.size = 10
        ws.range(f"A:{finalCol}").font.name = "微软雅黑"

        # 进行冻结操作(F2)
        active_window = wb.app.api.ActiveWindow
        active_window.FreezePanes = False
        # wb.app.range("A2").select()  # 选"A2"冻结首行
        active_window.SplitColumn = 0  # 冻结至哪一列
        active_window.SplitRow = 1  # 冻结至哪一行
        active_window.FreezePanes = True

        # 据标题设置背景色、加粗
        ws.range("A1:%s1" % finalCol).font.bold = True
        ws.range("A1:%s1" % finalCol).color = "#BDD7EE"
        ws.range("%s1:%s1" % (s_col, e_col)).color = "#FFC000"

        # 设置格式, 时间以短日期显示，金额以会计专用显示(小数两位、货币符号无)
        for idx, col in enumerate(DeliveryTableCol):
            index = get_column_letter(idx + 1)
            if "时间" in col or "日期" in col:
                ws.range(f"{index}:{index}").number_format = "yyyy/m/d"
            if "金额" in col:
                ws.range(f"{index}:{index}").number_format = "_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * ""-""??_ ;_ @_ "

        # 写入数据
        ws.range("A2").value = df_hw.values

        # 自适应宽度
        # ws.autofit()
        # 设置列宽
        ws.used_range.column_width = 18

        # 保存并关闭工作簿
        wb.save(savePath)
        wb.close()

        # 调用接口上传DB
        # 6.对所有的时间/日期列，将"-"替换回None，金额、数量、次数列，将空替换为0
        for col in DeliveryTableCol:
            if "时间" in col or "日期" in col:
                df_hw[col] = df_hw[col].apply(lambda x: None if x == "-" else x)

            if "金额" in col or "数量" in col or "次数" in col:
                df_hw[col] = df_hw[col].apply(lambda x: 0 if x == "" else x)
        # 将列名替换为接口数据形式
        df_hw.columns = DeliveryTableColDB
        updateData = df_hw.to_dict(orient='records')

        # 依据post次数生成更新货期表数据
        times = (len(updateData) + maxtTransferData - 1) // maxtTransferData
        totalData = []
        for i in range(times):
            thisData = updateData[maxtTransferData * i: maxtTransferData * (i + 1)]
            totalData.append(thisData)
            # response = requests.post(updateUrl, json=thisData)
            # response_json = response.json()
            # print(response_json)
            # if not response_json["Result"]:
            #     raise Exception(response_json["Data"])
            # else:
            #     time.sleep(2)
        return totalData

    except Exception as e:
        raise Exception(traceback.format_exc(limit=3))


"""
DeliveryTableCol: 货期表表头
DeliveryTableColDB: 货期表上传DB时表头对应的字段
maxtTransferData: 货期表数据单次上传的数据量
logger：用于打印日志
"""
DeliveryTableCol = ["华为订单号", "订单名称", "45采购订单", "批次", "采购组织", "二级经销商", "订单激活时间", "订单状态", "运输方式", "最早预计备货完成时间",
                    "最晚预计备货完成时间", "付款状态", "订单总金额", "待付款金额", "是否全部发货", "客户签收时间", "订单版本", "纯软件订单标识", "签约经销商", "清洁订单标识",
                    "产品线", "是否商务受控", "最终客户", "地址待定标识", "开票状态", "代表处", "总代自提物流状态", "待激活原因", "项目负责人", "批次数量", "是否退货",
                    "退货状态", "实际发货时间", "详细收货地址", "加速次数", "授信使用金额", "未满足备发货条件", "预计到货日期", "实际备货完成日期", "到达收货城市时间", "设备金额",
                    "服务金额", "订单提交时间", "华为签约主体", "签收单上载时间", "订单已付金额", "最后一次付款时间", "激励使用金额", "经销商付款日期"]
DeliveryTableColDB = ['huawei_order_code', 'name', 'code_45', 'batch', 'purchasing_organization', 'second_level_dealer',
                      'order_activation_time', 'order_status', 'transportation_mode', 'ready_goods_at_the_earliest',
                      'ready_goods_at_latest', 'payment_status', 'order_total_amount', 'unpaid_amount',
                      'is_all_deliver_goods', 'customer_signing_time', 'order_versions', 'is_software_order',
                      'contracted_distributor', 'is_clear_order', 'product_line', 'is_business_control', 'end_customer',
                      'address_is_pending', 'billing_status', 'representative_office', 'self_pickup_logistics_status',
                      'unactivated_reason', 'project_leader', 'batch_quantity', 'is_return_goods',
                      'return_goods_status', 'actual_delivery_time', 'delivery_address', 'number_of_lifts',
                      'line_of_credit', 'reasons_for_non_delivery', 'estimated_date_of_delivery',
                      'actual_purchase_completion_date', 'time_of_arrival', 'amount_of_equipment', 'amount_of_service',
                      'order_submit_time', 'huawei_subject_of_contract', 'receipt_upload_time', 'order_amount_paid',
                      'last_payment_time', 'incentive_use_amount', "dealer_payment_date"]
maxtTransferData = 2000
logger = None

if __name__ == "__main__":
    print(getLatestFile(r"E:\Uibot项目\华为全流程\下载数据", "*_华为数据.xlsx"))
