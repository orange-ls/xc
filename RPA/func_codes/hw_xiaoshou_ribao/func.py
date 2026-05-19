#!/usr/bin/python3
# -*- coding: UTF-8 -*-


import calendar
import os
from datetime import datetime, timedelta

import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter


# 将MHTML文件转为xlsx
def changeMhtmlToXlsx(path):
    """
    :param path: MHTML文件路径
    :return: 转换后的xlsx文件路径
    """
    outPath = os.path.join(os.path.dirname(path), os.path.basename(path).replace(".MHTML", ".xlsx"))
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(path)
    # ws = wb.sheets["Sheet1"]
    # ws.autofit()
    wb.save(outPath)
    # 关闭工作簿
    wb.close()
    app.quit()

    # # 等待文件生成
    # startTIme = time.time()
    # while True:
    #     if time.time() - startTIme < 3600:
    #         if os.path.exists(outPath):
    #             break
    #     else:
    #         raise Exception("将文件：%s转为xlsx格式超时" % path)

    # 删除MHTML文件
    os.remove(path)
    return outPath


# 读取配置文件生成配置字典
def getConfigDict(baseConfPath):
    """
    :param baseConfPath：配置文件目录
    :return: 生成配置字典
    """
    df = pd.read_excel(baseConfPath, dtype=str, sheet_name="Sheet1").fillna("")
    resultDict = dict(zip(df["配置名称"], df["配置内容"]))
    return resultDict


# 根据输入的文件根目录返回当日文件保存目录
def getSaveDir(rootDir):
    """
    :param rootDir：文件保存根目录，例D:\downDir
    :return: 返回根目录+当日日期的目录，如D:\downDir\2022年\6月\8日
    """
    today = datetime.now()
    year, month, day = today.year, today.month, today.day
    finialDir = os.path.join(rootDir, "%s年\\%s月\\%s日" % (year, month, day))
    return finialDir


# 获取查询"销售日报"开始日期和结束日期及查询范围
def getQryTime():
    """
    开始日期：当前日期减一天所在月份的第一天（eg：5月1号时开始日期为：2022/04/01）
    结束日期：昨天
    """
    yesterday = datetime.now() - timedelta(days=1)
    lastMonthDay = datetime(yesterday.year, yesterday.month, 1)
    finishTime = yesterday.strftime("%Y/%m/%d")
    startTime = lastMonthDay.strftime("%Y/%m/%d")
    return [startTime, finishTime, f"{lastMonthDay.month}.{lastMonthDay.day}-{yesterday.month}.{yesterday.day}"]


# 获取查询"已销未提"开始日期和结束日期及查询范围
def getQryTime2():
    """
    开始日期：当前日期减一天所在月份的第一天（eg：5月1号时开始结束日期为：2022/04/01-2022/04/30）
    结束日期：当前日期减一天所在月份的最后一天
    """
    yesterday = datetime.now() - timedelta(days=1)
    firstMonthDay = datetime(yesterday.year, yesterday.month, 1)
    monthdays = calendar.monthrange(yesterday.year, yesterday.month)[1]
    LastMonthDay = datetime(yesterday.year, yesterday.month, monthdays)
    finishTime = LastMonthDay.strftime("%Y/%m/%d")
    startTime = firstMonthDay.strftime("%Y/%m/%d")
    return [startTime, finishTime, f"{firstMonthDay.month}.{firstMonthDay.day}-{LastMonthDay.month}.{LastMonthDay.day}"]


# 四舍五入
def new_round(_float, _len):
    """
    :param _float: 需要四舍五入的数
    :param _len: 保留小数点位数
    :return: 四舍五入结果
    """
    if isinstance(_float, float):
        if str(_float)[::-1].find('.') <= _len:
            return _float
        if str(_float)[-1] == '5':
            return round(float(str(_float)[:-1] + '6'), _len)
        else:
            return round(_float, _len)
    else:
        return round(_float, _len)


# 自定义解析时间数据（不对na和空值处理）
def myDateParser(x):
    """
    :param x: 传入的时间数据
    :return: 返回%Y/%m/%d形式的文本
    """
    if x and not pd.isna(x):
        return datetime.strftime(x, "%Y/%m/%d")
    else:
        return ""


# 生成销售日报(HV)文件路径
def get_HVPath(filedir):
    """
    :param filedir: 文件保存目录
    :return: 返回HV表文件路径
    """
    calTime = getQryTime()[1]
    HV_name = "FY%s销售明细(HV)_%s%s.xlsx" % (calTime[2:4], calTime[5:7], calTime[8:10])
    HV_path = os.path.join(filedir, HV_name)
    return HV_path


# 匹配销售日报(HV)结果表数据，并调整表格式
def match_HVdata(filepath, HV_path):
    """
    :param filepath: 销售日报表路径
    :param HV_path: 需要保存的销售明细HV结果表路径
    :return: 解析处理后的销售日报DataFrame
    """
    # 读取销售日报表
    df = pd.read_html(filepath, header=0)[0].fillna("").astype(str)

    # 筛选数据并写入 销售明细(HV)结果表
    # df_HV = df.query("分销渠道 == '15' and 产品组 == 'HV' and 客户编号 == ['200000041','200000201'] ")
# 20240620修改：客户编号的数据不用单独筛选
    df_HV = df.query("分销渠道 == '15' and 产品组 == 'HV'")
    df_HV.to_excel(HV_path, index=False)

    # 设置销售明细(HV)结果表格式
    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(HV_path)
    ws = wb.sheets["Sheet1"]
    cols = ws.used_range.shape[1]

    # 最后一列列名
    colIndex = get_column_letter(cols)

    # 全表格式改成常规格式、字体为微软雅黑9号
    ws.range(f"A:{colIndex}").number_format = "G/通用格式"
    ws.range(f"A:{colIndex}").font.size = 9
    ws.range(f"A:{colIndex}").font.name = "微软雅黑"
    # plan A：已经用本文格式存储的数字，需要进行分列操作,但会导致部分列数据损失
    # colIndexList = get_column_interval(1, cols)
    # for indexCol in colIndexList:
    #     ws.range(indexCol+"1").expand("down").api.TextToColumns()

    # plan B:由于单元格格式已经为常规，将所有数据复制重新粘贴可去除文本格式显示的数字（表现为绿三角）
    ws.range("A1").value = ws.used_range.value

    # 原有的数据标题设置背景色、加粗
    ws.range("A1:%s1" % colIndex).color = (211, 211, 211)
    ws.range("A1:%s1" % colIndex).font.bold = False

    # 自适应宽度
    ws.autofit()

    wb.save(HV_path)
    wb.close()
    app.quit()

    return df.query("分销渠道 != '15'").reset_index(drop=True), df.shape[1]


# 获取各币种对应的汇率
def getExchangeRate(exchangePath):
    """
    :param exchangePath: 汇率换算配置表路径
    :return: 返回汇率df
    """
    dfExchange = pd.read_excel(exchangePath, sheet_name="汇率换算", index_col=1)
    dfExchange.drop(columns="币种1", inplace=True)
    dfExchange = dfExchange.T
    dfExchange = dfExchange.sort_index()

    return dfExchange


# 计算合同含税金额
def calTaxAmount(series, dfExchange):
    """
    :param series: DataFrame行series
    :param dfExchange: 汇率df
    :return: 合同含税金额
    """
    # rateDict = dfExchange.asof(series["出具发票日"]).to_dict()  # asof：查找索引等于输入参数或输入参数前最近的一个非空行,uibot的pandas版本1.0.5，该方法报错，1.3.5可以
    rateDict = dfExchange.loc[dfExchange[dfExchange.index <= pd.to_datetime(series["出具发票日"])].index[-1]].to_dict()
    rate = rateDict.get(series["货币"], 1)
    calAmount = (float(series["税后金额"]) - float(series["返款折扣"])) * rate
    return calAmount


# 计算实际税率
def calRealRate(series, numrateDict):
    """
    :param series: DataFrame行series
    :param numrateDict: 订单号对应销项税字典
    :return: 实际税率
    """
    initValue = series["实际税率"]
    key = series.name
    return numrateDict.get(key, initValue)


# 计算返款折扣
def calrebate(series, dfExchange):
    """
    :param series: DataFrame行series
    :param dfExchange: 汇率df
    :return: 返款折扣
    """
    # rateDict = dfExchange.asof(series["出具发票日"]).to_dict()  # asof：查找索引等于输入参数或输入参数前最近的一个非空行,uibot的pandas版本1.0.5，该方法报错，1.3.5可以
    rateDict = dfExchange.loc[dfExchange[dfExchange.index <= pd.to_datetime(series["出具发票日"])].index[-1]].to_dict()
    rate = rateDict.get(series["货币"], 1)
    calAmount = float(series["返款折扣"]) * rate
    return calAmount


# 生成备注
def generateRemark(series: pd.Series, dfConf: pd.DataFrame):
    """
    :param series: DataFrame行series
    :param indexList: 已销未提表数据
    :return: 备注
    """

    indexList = dfConf.index.tolist()
    if "折让" in series["合同号（客户PO号）"] or "折让" in series["物料名称"]:
        return "返款抵欠款"
    elif "预开冲红" in series["合同号（客户PO号）"] or "预开冲红" in series["物料名称"]:
        return "已销未提"
    elif series.name in indexList:
        if series[["数量", "税前金额"]].tolist() == dfConf.loc[series.name, ["数量", "销售金额"]].tolist():
            return "已销未提"
        else:
            return "部分已销未提"
    else:
        return ""


# 生成月份（不对na和空值处理）
def calSaleMonth(x):
    """
    :param x: 传入的时间（例：2022/05/31）
    :return: 返回月份数据（例：5月）
    """
    if x and not pd.isna(x):
        return str(int(x[5:7])) + "月"
    else:
        return ""


# 生成销售分布（不对na和空值处理）
def calSaleWeek(x):
    """
    :param x: 传入的时间（例：2022/05/31）
    :return: 返回销售分布（例：5月第5周）
    """
    if x and not pd.isna(x):
        month = str(int(x[5:7])) + "月"
        day = int(x[8:])
        value = "%s月第%s周" % (month, (day - 1) // 7 + 1)
        return value
    else:
        return ""


# 计算销售员、销售员编码、合同含税金额、实际税率(初始)
def calDataStep1(df, dfExchange):
    """
    :param df:需要处理的df
    :param dfExchange:汇率df
    :return:处理后的df，透视表df（含实际税率），需要查询销项税的订单号列表
    """
    df["销售员"] = df["雇员姓名"]
    df["销售员编码"] = df["销售雇员"].map(lambda x: x.zfill(8))
    df["合同含税金额"] = df.apply(calTaxAmount, args=(dfExchange,), axis=1)
    currencyDictdict = dict(zip(df["销售订单号"], df["货币"]))
    df_pivot = df.astype({"税前金额": "float64", "税后金额": "float64"}).pivot_table(values=["税前金额", "税后金额"], index="销售订单号",
                                                                             aggfunc="sum")
    df_pivot["实际税率"] = (df_pivot["税后金额"] / df_pivot["税前金额"] - 1).map(lambda x: new_round(x, 2))
    df_pivot["货币"] = df_pivot.index.map(lambda x: currencyDictdict[x])
    df_needQry = df_pivot.query("货币 != 'RMB' or 实际税率 not in [0.13, 0.06]")
    return df, df_pivot, df_needQry.index.tolist()


# 计算实际税率(查询后)、合同不含税金额、折扣、折扣后合同金额、不含税收入、业绩划分，返回处理后的df
def calDataStep2(df, df_pivot, qryDict, dfExchange):
    """
    :param df: 需要处理的df
    :param df_pivot: 初始透视表数据（含实际税率）
    :param qryDict:从SAP查询到的销项税字典{订单号：税率}
    :param dfExchange:汇率df
    :return:处理后的df
    """
    qryDict = {i: new_round(float(qryDict[i]) / 100, 2) for i in qryDict}
    df_pivot["实际税率"] = df_pivot.apply(calRealRate, args=(qryDict,), axis=1)
    rateRealDict = dict(zip(df_pivot.index, df_pivot["实际税率"]))
    df["实际税率"] = df["销售订单号"].map(lambda x: rateRealDict.get(x))
    df["合同不含税金额"] = df["合同含税金额"] / (df["实际税率"] + 1)
    df["折扣"] = df.apply(calrebate, args=(dfExchange,), axis=1)
    df["折扣后合同金额"] = df["税后金额"]
    df["不含税收入"] = df["合同不含税金额"] / 10000
    year = getQryTime()[1][2:4]
    df["业绩划分"] = year + "年"
    return df


# 计算事业部、区域、平台、备注、月份、销售分布,返回生成处理后销售日报表路径,新表列名列表
def calDataStep3(df: pd.DataFrame, personConfigPath, filepath, HV_path, BO_path):
    """
    :param df: 需要处理的df
    :param personConfigPath: 销售员信息配置表
    :param filepath:已销未提表路径
    :param HV_path:销售日报（HV）路径，用于生成结果表路径
    :param BO_path: BusinessObjects下载的客户省份表
    :return: 处理结果表保存路径，新表列名列表
    """
    df_person = pd.read_excel(personConfigPath, dtype=str)
    for key in ["事业部", "区域", "平台"]:
        matchDict = dict(zip(df_person["人员代码"], df_person[key]))
        df[key] = df["销售员编码"].map(lambda x: matchDict.get(x))
    dfConf = pd.read_html(filepath, header=0)[0].fillna("").astype(str)
    # dfConf.set_index(keys=["销售凭证", "项目", "数量", "销售金额"], inplace=True)
    # df.set_index(keys=["销售订单号", "销售订单行项目", "数量", "税前金额"], drop=False, inplace=True)
    # df["备注"] = df.apply(generateRemark, args=(dfConf.index,), axis=1)
    for col in ["销售凭证", "项目"]:
        dfConf[col] = dfConf[col].str.lstrip("0")
    dfConf.set_index(keys=["销售凭证", "项目"], inplace=True)
    df.set_index(keys=["销售订单号", "销售订单行项目"], drop=False, inplace=True)
    df["备注"] = df.apply(generateRemark, args=(dfConf,), axis=1)
    df["月份"] = df["出具发票日"].map(calSaleMonth)
    df["销售分布"] = df["出具发票日"].map(calSaleWeek)

    # 读取BO下载的客户省份表，并在“客户名称”列后新增“客户省份”列
    df_BO = pd.read_excel(BO_path, dtype=str, header=3).fillna("")
    userProvinceDict = dict(zip(df_BO["客户代码"], df_BO["客户省份"]))
    addSeries = df["客户编号"].str.zfill(10).apply(lambda x: userProvinceDict.get(x, ""))
    df.insert(df.columns.tolist().index("客户名称") + 1, "客户省份", addSeries)

    finialPath = HV_path.replace("HV", str(int(HV_path[-9:-7])) + "月")
    df.to_excel(finialPath, index=False)
    return finialPath, df.columns.tolist()


# 设置销售日报表格式
def setStyle(finialPath, originalCols, allColumnsList):
    """
    :param finialPath: 结果表路径
    :param originalCols: 原销售日报列数，用于设置格式
    :param allColumnsList: 销售日报列名列表
    :return:
    """
    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finialPath)
    ws = wb.sheets["Sheet1"]
    cols = ws.used_range.shape[1]
    personCol = get_column_letter(allColumnsList.index("销售员编码") + 1)
    batchCol = get_column_letter(allColumnsList.index("批次") + 1)

    # 计算标志列（原表数据列终止位置、新增数据开始位置和结束位置）
    colIndex1 = get_column_letter(originalCols)
    colIndex2 = get_column_letter(originalCols + 1)
    colIndex3 = get_column_letter(cols)

    # 全表格式改成常规格式、字体为微软雅黑9号
    ws.range(f"A:{colIndex3}").number_format = "G/通用格式"
    ws.range(f"A:{colIndex3}").font.size = 9
    ws.range(f"A:{colIndex3}").font.name = "微软雅黑"
    ws.range(f"{personCol}:{personCol}").number_format = "@"  # "销售员编码"列为文本格式
    ws.range(f"{batchCol}:{batchCol}").number_format = "@"  # "批次"列为文本格式
    # plan A：已经用本文格式存储的数字，需要进行分列操作,但会导致部分列数据损失
    # colIndexList = get_column_interval(1, cols)
    # for indexCol in colIndexList:
    #     ws.range(indexCol+"1").expand("down").api.TextToColumns()

    # plan B:由于单元格格式已经为常规，将所有数据复制重新粘贴可去除文本格式显示的数字（表现为绿三角）
    ws.range("A1").value = ws.used_range.value

    # 进行冻结操作
    active_window = wb.app.api.ActiveWindow
    active_window.FreezePanes = False
    # wb.app.range("A2").select()  # 选"A2"冻结首行
    active_window.SplitColumn = 0  # 冻结至哪一列
    active_window.SplitRow = 1  # 冻结至哪一行
    active_window.FreezePanes = True

    # 原有的数据标题设置背景色、加粗
    ws.range("A1:%s1" % colIndex1).color = (211, 211, 211)
    ws.range("A1:%s1" % colIndex1).font.bold = False

    # 新增的数据标题设置背景色、字体颜色、加粗、数字格式
    ws.range("%s1:%s1" % (colIndex2, colIndex3)).color = (0, 112, 192)
    ws.range("%s1:%s1" % (colIndex2, colIndex3)).font.color = (255, 255, 255)
    ws.range("%s1:%s1" % (colIndex2, colIndex3)).font.bold = True

    # 新增数据设置格式, "实际税率"保留两位小数，其他数字列为千分位显示整数
    for col in ["实际税率", "合同含税金额", "合同不含税金额", "折扣", "折扣后合同金额", "不含税收入"]:
        index = get_column_letter(allColumnsList.index(col) + 1)
        if col == "实际税率":
            ws.range(f"{index}:{index}").number_format = '0.00'
        else:
            ws.range(f"{index}:{index}").number_format = "#,##0_ "

    # 自适应宽度
    ws.autofit()

    wb.save(finialPath)
    wb.close()
    app.quit()


if __name__ == "__main__":
    g_dictGlobal = {"销售日报": r"D:\xc_files\销售日报\1.1-1.30_销售日报export.MHTML",
                    "已销未提": r"D:\Uibot项目文件\test\下载\08日_已销未提export.xlsx",
                    "销售员大区对应表路径": r"D:\Uibot项目文件\配置表\外挂-华为SBU销售员大区对应表-21.xlsx",
                    "汇率换算表路径": r"D:\xc_files\销售日报\人员划分&汇率换算.xlsx",
                    "保存路径": r"D:\xc_files\销售日报\结果"}
    g_dictGlobal["销售日报HV"] = get_HVPath(g_dictGlobal["保存路径"])
    generateHVResult = match_HVdata(g_dictGlobal["销售日报"], g_dictGlobal["销售日报HV"])
    df = generateHVResult[0]
    originalCols = generateHVResult[1]
    dfExchange = getExchangeRate(g_dictGlobal["汇率换算表路径"])
    calResultMsg = calDataStep1(df, dfExchange)
    g_dictGlobal["df"] = calResultMsg[0]
    g_dictGlobal["df_pivot"] = calResultMsg[1]
    g_dictGlobal["税率查询订单列表"] = calResultMsg[2]
    numRateDict = {}
    df = calDataStep2(g_dictGlobal["df"], g_dictGlobal["df_pivot"], numRateDict, dfExchange)
    g_dictGlobal["销售日报"], colList = calDataStep3(df, g_dictGlobal["销售员大区对应表路径"], g_dictGlobal["已销未提"],
                                                 g_dictGlobal["销售日报HV"], "客户省份表路径")
    setStyle(g_dictGlobal["销售日报"], originalCols, colList)

    # colList = ['销售订单号', '销售订单行项目', '订单类型', '订单原因', '销售组织', '分销渠道', '产品组', '销售办公室', '销售组', '客户编号', '客户名称', '出具发票日',
    #            '预计还款日', '后勤处理组', '物料号', '物料名称', '订单标准价格', '数量', '单位', '折扣（含模拟返款折扣）', '返款折扣', '税后金额', '订单成本（公司代码货币）',
    #            '货币', '系统发票', '增值税发票号', '工厂', '库存地', '批次', '原始标准价格', '最低限价', '您的参考', '特价单号', '特单折扣', '厂商物料号', '销售雇员',
    #            '雇员姓名', '销售员比率', '销售地区', '地区描述', '合同号（客户PO号）', '送达方编号', '送达方名称', '送达方地址', '服务年限', '燎原计划', '回佣单价', '回佣额',
    #            '旧物料号', '销售订单版本', '订单标准价格×汇率', '折扣（含模拟返款折扣）×汇率', '返款折扣×汇率', '税后金额×汇率', '订单成本（利润中心货币）', '业务货币对公司代码货币汇率',
    #            '模拟返款', '行业', '行业文本', '实际销售员编号', '实际销售员名称', '行项目销售员', '行项目销售员姓名', '出具发票方', '出具发票方名称', 'ZTFF', 'ZTFD',
    #            '特价单号(抬头文本)', '内部订单号', '税前金额', '物料组', '物料组描述', '产品层次', '产品层次描述', '表头注释', '会计凭证号', '公司代码', '会计年度',
    #            '凭证抬头文本', '清帐凭证', '交货单号', '交货单行项目', '订货方名称', '现金点ZTEB', '折扣ZT6D', '商品税收分类编码', '发票类型', '项目注释',
    #            '合同号（送达方采购订单编号）', '行项目打包注释', '地区代码', '省', '城市', '发票行项目号', '是否取消', '定价单位', '单价', '单价*汇率', '系统发票创建日期',
    #            '销售订单创建日期', '税后金额*对RMB汇率', '售达方地址-省份', '销售员', '销售员编码', '合同含税金额', '实际税率', '合同不含税金额', '折扣', '折扣后合同金额',
    #            '不含税收入', '业绩划分', '事业部', '区域', '平台', '备注', '月份', '销售分布']
    # setStyle(g_dictGlobal["销售日报"], 102, colList)
