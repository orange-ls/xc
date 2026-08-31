#!/usr/bin/python3
# -*- coding: UTF-8 -*-


import gc
import glob
import logging.config
import os
import re
import shutil
from copy import deepcopy
from datetime import datetime, timedelta
from functools import wraps

import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter


# 初始化日志
def initWriteLog(rootDir):
    """
    :param rootDir: 日志写入目录
    :return:
    """
    global logger
    log_name = "销售明细"
    simple_format = '[%(levelname)s][%(asctime)s]%(message)s'
    logfile_dir = rootDir  # log文件的目录
    logfile_name = '%s.log' % log_name  # log文件名
    # 如果不存在定义的日志目录就创建一个
    if not os.path.exists(logfile_dir):
        os.makedirs(logfile_dir)

    # log文件的全路径
    logfile_path = os.path.join(logfile_dir, logfile_name)
    LOGGING_DIC = {
        'version': 1,
        'disable_existing_loggers': False,
        'formatters': {
            'simple': {
                'format': simple_format
            },
        },
        'filters': {},
        'handlers': {
            'default': {
                'level': 'INFO',
                'class': 'logging.handlers.RotatingFileHandler',
                'formatter': 'simple',
                'filename': logfile_path,
                'maxBytes': 1024 * 1024 * 5,
                'backupCount': 5,
                'encoding': 'utf-8',
            },
        },
        'loggers': {
            '': {
                'handlers': ['default'],
                'level': 'INFO',
                'propagate': True,
            },
        },
    }
    logging.config.dictConfig(LOGGING_DIC)
    logger = logging.getLogger(__name__)


# 初始化操作日期
def initOperateDate(selectPath):
    """
    :param selectPath:选择的销售日报路径
    :return:
    """

    global calYearMonth
    calYearMonthGroup = re.search(".*FY(\d{2}).*\((\d{1,2})月\).*", os.path.basename(selectPath))
    calYearMonth = f"20{calYearMonthGroup.group(1)}{calYearMonthGroup.group(2).zfill(2)}"


# 定义一个日志装饰器，每次调用方法前后打印日志
# 注 uibot中无法调用，弃用
def logfun(func):
    @wraps(func)
    def logfunStep(*args, **kwargs):
        logger.info(f"{'-' * 25}{func.__name__}开始{'-' * 25}")
        result = func(*args, **kwargs)
        logger.info(f"{'-' * 25}{func.__name__}结束{'-' * 25}")
        return result

    return logfunStep


# 将MHTML文件转为xlsx
def changeMhtmlToXlsx(path):
    """
    :param path: MHTML文件路径
    :return: 转换后的xlsx文件路径
    """
    outPath = os.path.join(os.path.dirname(path), os.path.basename(path).replace(".MHTML", ".xlsx"))
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(path)
    # ws = wb.sheets["Sheet1"]
    # ws.autofit()
    wb.save(outPath)
    # 关闭工作簿
    wb.close()
    app.quit()

    # # 等待文件生成
    # startTIme = time.time()
    # while True:
    #     if time.time() - startTIme < 3600:
    #         if os.path.exists(outPath):
    #             break
    #     else:
    #         raise Exception("将文件：%s转为xlsx格式超时" % path)

    # 删除MHTML文件
    os.remove(path)
    return outPath


# 读取配置文件生成配置字典
def getConfigDict(baseConfPath):
    """
    :param baseConfPath：配置文件目录
    :return: 生成配置字典
    """
    df = pd.read_excel(baseConfPath, dtype=str, sheet_name="Sheet1").fillna("")
    resultDict = dict(zip(df["配置名称"], df["配置内容"]))
    # 华为账户密码有多个，处理为列表格式
    for key in ["华为网站账号", "华为网站密码", "销售明细排除的销售员"]:
        resultDict[key] = resultDict[key].replace("；", ";").strip(";").split(";")
    return resultDict


# 根据输入的文件根目录返回当日文件保存目录
def getSaveDir(rootDir):
    """
    :param rootDir：文件保存根目录，例E:\downDir
    :return: 返回根目录+当日日期的目录，如E:\downDir\2022年\6月\8日
    """
    today = datetime.now()
    year, month, day = today.year, today.month, today.day
    finialDir = os.path.join(rootDir, "%s年\\%s月\\%s日" % (year, month, day))
    return finialDir


# 获取查询下载的日期范围
def getQryTimeRange(saveDir, keyWord, timeFmtStr, user=""):
    """
    :param saveDir: 汇总表保存的路径
    :param keyWord: 文件名关键词标识
    :param timeFmtStr: 时间日期格式转换类型，例 %Y/%m/%d
    :param user: 下载华为订单表和业绩表的账号
    :return: 查找的文件路径，本次查询开始日期，本次查询结束日期, 文件最新标志等
    """
    nowday = datetime.now()
    yesterday = nowday - timedelta(days=1)

    if keyWord == "物料移动明细汇总":
        """
        物料移动汇总：文件名:物料移动明细汇总_20210728.xlsx
        1.日期表示文件保存的截止日期，物料移动汇总会下载截止到昨天的数据，再次下载时无需下载重复日期的数据
        2.下载时的日期即为2021/07/29-昨天
        """
        fileList = glob.glob(f"{saveDir}\\*{keyWord}_*.xlsx")
        filePath = fileList[0]
# 20240618
        fileListadd = glob.glob(f"{saveDir}\\*{keyWord}(新增列)_*.xlsx")
        filePathadd = fileListadd[0]

        lastDateStr = re.search(".*_(\d{8}).xlsx", os.path.basename(filePath)).group(1)
        # lastDate： 下载的起始日期
        lastDate = datetime.strptime(lastDateStr, "%Y%m%d") + timedelta(days=1)
        dayDelta = (yesterday - lastDate).days
        if dayDelta <= -1:
            lastFlag = True
        else:
            lastFlag = False
        startDate = lastDate.strftime(timeFmtStr)
        endDate = yesterday.strftime(timeFmtStr)
        # return 物料移动汇总表路径，下载开始日期，下载结束日期， 文件最新标识
        return filePath, startDate, endDate, lastFlag, filePathadd
    elif keyWord == "预提表":
        """
        预提表：文件名:预提表_20210728.xlsx
        1.日期表示文件保存的截止日期，预提表会下载本年的数据，当年数据会重新下载
        2.下载时的日期即为2021/01/01-当天
        """
        fileList = glob.glob(f"{saveDir}\\*{keyWord}_*.xlsx")
        filePath = fileList[0]
        lastDateStr = re.search(".*_(\d{8}).xlsx", os.path.basename(filePath)).group(1)
        lastDate = datetime.strptime(lastDateStr, "%Y%m%d")
        startDate = datetime(lastDate.year, 1, 1)
        endDate = nowday
        dayDelta = (endDate - lastDate).days
        if dayDelta <= 0:
            lastFlag = True
        else:
            lastFlag = False
        startDate = startDate.strftime(timeFmtStr)
        endDate = nowday.strftime(timeFmtStr)
        # yearRange：下载数据的年份跨度，用于替换汇总表该年份数据
        yearRange = list(pd.period_range(startDate, endDate, freq="Y").year)
        # return 预提汇总表路径，下载开始日期，下载结束日期， 文件最新标识，下载日期范围年份跨度
        return filePath, startDate, endDate, lastFlag, yearRange
    elif keyWord == "订单表":  # todo：1.配置表中的每个账号必须已经有汇总数据，新账号先自行下载一部分数据按指定规则命名即可，否则不会下载改新账号数据 2.如需兼容新账号无数据的情况，此方法及合并表的方法需要修改
        # """
        # 华为订单表：文件名:账号_订单表_20210728.xlsx
        # 1.日期表示文件保存的截止日期，华为订单表会下载截止到昨天的数据，再次下载时无需下载重复日期的数据
        # 2.下载时的日期即为2021/07/29-昨天
        # """
        # fileList = glob.glob(f"{saveDir}\\{user}_{keyWord}_*.xlsx")
        # filePath = fileList[0]
        # matchObj = re.search("订单表_(\d{8}).xlsx", os.path.basename(filePath))
        #
        # lastDate = datetime.strptime(matchObj.group(1), "%Y%m%d") + timedelta(days=1)
        # dayDelta = (yesterday - lastDate).days
        # if dayDelta <= -1:
        #     lastFlag = True
        # else:
        #     lastFlag = False
        # startDate = lastDate.strftime(timeFmtStr)
        # endDate = yesterday.strftime(timeFmtStr)
        # # return 华为订单汇总表路径，下载开始日期，下载结束日期， 文件最新标识
        # return filePath, startDate, endDate, lastFlag

        """ 变更：订单表不再是不断汇总的形式，每次直接下载近1年的表"""
        # startDay = datetime(year=nowday.year - 1, month=1, day=1)
        startDay = datetime(year=nowday.year - 1, month=nowday.month, day=nowday.day)
        startDate = startDay.strftime(timeFmtStr)
        endDate = nowday.strftime(timeFmtStr)
        return startDate, endDate

    elif keyWord == "业绩表":
        """
        华为业绩表：文件名:账号_2020业绩表（20210203）.xlsx
        1.每个账号华为业绩表均会下载近三年的数据，重新下载的年份的数据均需要替换
        2.n年的数据需要在>n年下载才无需进行替换
        3.如有跨年情况，上次为20211228下载的2021年数据，本次需要下载到2022年数据，则2021年数据需要重新下载，在下次执行时即不需要
        """
        fileList = glob.glob(f"{saveDir}\\{user}_*{keyWord}（*）.xlsx")
        # startYear：当前年前两年，例如2026年则为2024年
        startYear = nowday.year - 2
        # totalYearList：从startYear年到当年的所有年
        totalYearList = list(pd.period_range(startYear, nowday.year, freq="Y").year)
        totalYearList = [str(i) for i in totalYearList]
        # 汇总表目录中的业绩表已经存在的数据年份
        existYear = []
        # 需要下载的年份
        downYearList = []
        # 数据非最新的文件，需要删除
        deleteFileList = []

        for filePath in fileList:
            matchObj = re.search("(\d{4})业绩表（(\d{8})）.xlsx", os.path.basename(filePath))
            existYear.append(matchObj.group(1))
            # 如果非当年数据不是由后几年下载的，则需要重新下载且删除
            if matchObj.group(1) != str(nowday.year):
                if matchObj.group(1) >= matchObj.group(2)[:4]:
                    downYearList.append(matchObj.group(1))
                    deleteFileList.append(filePath)
            else:  # 当年但非当天下载的数据需要重新下载且删除
                if matchObj.group(2) < nowday.strftime("%Y%m%d"):
                    downYearList.append(matchObj.group(1))
                    deleteFileList.append(filePath)
        # 汇总表目录中从2021-当年的数据，除去已经存在并校验的数据，缺失的年份需重新下载
        for i in existYear:
            try:
                totalYearList.remove(i)
            except ValueError:
                pass
        downYearList.extend(totalYearList)
        # return 下载的年份列表，需要删除的文件列表
        return downYearList, deleteFileList
    elif keyWord == "订单全字段报表":  # todo：1.配置表中的每个账号必须已经有汇总数据，新账号先自行下载一部分数据按指定规则命名即可，否则不会下载改新账号
        """
        华为订单全字段报表：文件名:账号_订单全字段报表_20210728.xlsx
        1.日期表示文件保存的截止日期，华为订单全字段报表会下载截止到当天01:00:00的数据，再次下载时无需下载重复日期的数据
        2.下载时的日期即为2021/07/28-当天凌晨1点
        """
        timeSuffix = " 01:00:00"
        fileList = glob.glob(f"{saveDir}\\{user}_{keyWord}_*.xlsx")
        filePath = fileList[0]
        matchObj = re.search("订单全字段报表_(\d{8}).xlsx", os.path.basename(filePath))

        lastDate = datetime.strptime(matchObj.group(1), "%Y%m%d")
        dayDelta = (nowday - lastDate).days
        if dayDelta <= 0:
            lastFlag = True
        else:
            lastFlag = False
        startDate = lastDate.strftime(timeFmtStr) + timeSuffix
        endDate = nowday.strftime(timeFmtStr) + timeSuffix
        # return 华为订单汇总表路径，下载开始日期，下载结束日期， 文件最新标识
        return filePath, startDate, endDate, lastFlag


# 四舍五入
def new_round(_float, _len):
    """
    :param _float: 需要四舍五入的数
    :param _len: 保留小数点位数
    :return: 四舍五入结果
    """
    if isinstance(_float, float):
        if str(_float)[::-1].find('.') <= _len:
            return _float
        if str(_float)[-1] == '5':
            return round(float(str(_float)[:-1] + '6'), _len)
        else:
            return round(_float, _len)
    else:
        return round(_float, _len)


# 自定义解析时间数据（不对na和空值处理）
def myDateParser(x):
    """
    :param x: 传入的时间数据
    :return: 返回%Y/%m/%d形式的文本
    """
    if x and not pd.isna(x):
        return datetime.strftime(x, "%Y/%m/%d")
    else:
        return ""


# 筛选销售日报有效数据
# @logfun
def match_validData(filepath, excludeCode: list, delPath):
    """
    :param filepath: 销售日报结果表路径
    :param excludeCode: 销售明细需要排除的销售员编码
    :param delPath: 分销销售名单表路径
    :return: 解析处理后的销售日报DataFrame，初始数据列数
    """

    # 读取销售日报表（不指定读取格式），获取所有时间日期列
    # dateColList = []
    # df_original = pd.read_excel(filepath, sheet_name="Sheet1")
    # for col in df_original.columns:
    #     if df_original[col].dtype == "datetime64[ns]":
    #         dateColList.append(col)

    # 读取销售日报表，将时间日期列重新解析为字符串
    df = pd.read_excel(filepath, sheet_name="Sheet1", dtype=str, parse_dates=dateColList,
                       date_parser=lambda x: myDateParser(x))
    df.rename(columns={"合同含税金额": "合同金额"}, inplace=True)

    # 筛选出"备注"列不为"已销未提"或"部分已销未提"的数据
    df = df.query("备注 not in ['已销未提', '部分已销未提']")

    # 筛选数据需要的列 --20250429 增加“项目注释”列，用于填充”价外费用“（后面删除这一列）
    df = df[usedCol+['项目注释']+['订单类型备注']]

    # 获取排除不参与计算的销售员编码后的数据
    df = df.query("销售员编码 not in @excludeCode")

    # 读取"分销销售名单",用于删除"产品组"为QJ中的人员
    dfConf = pd.read_excel(delPath, dtype=str)
    delList = dfConf["人员编号"].tolist()

    # 筛选出”产品组”列为PU、HI、HT、QJ且QJ不在"分销销售名单中"的数据
    df = df.query("产品组 in ['PU', 'HI', 'HT'] or (产品组 == 'QJ' and 销售员编码 not in @delList) or 订单类型备注 == '政企分销'")
    # 筛选出“事业部”不为“商业业务部”的数据
    # df = df.query("事业部 != '商业业务部'")

    # 清理内存
    gc.collect()

    return df, df.shape[1]


# 增加销售明细补充表数据
# @logfun
def addExtraData(df, extraPath):
    """
    :param df:筛选后的销售日报数据
    :param extraPath:销售明细补表路径
    :return 补充“销售明细补充”外挂表数据后的销售日报DataFrame, “销售明细补充”外挂表中人工新增的[下单合同号, 项目名称, 评审二代]的数据
    """

    """
    1.将“合同号（客户PO号）“列或”物料名称“列包含“折让”的数据复制
    2.将”销售订单号”改为”返款抵欠款”
    3.”数量”、” 合同金额”、” 合同不含税金额”均乘以-1（即将值修改为正数）并增加到原数据中
    """
    df_copy = df[df["合同号（客户PO号）"].str.contains("折让") | df["物料名称"].str.contains("折让")].copy()
    df_copy["销售订单号"] = "返款抵欠款"
    # print(df_copy[["数量", "合同金额", "合同不含税金额"]])
    df_copy[["数量", "合同金额", "合同不含税金额"]] = df_copy[["数量", "合同金额", "合同不含税金额"]].astype(
        "float") * -1
    # print(df_copy["数量"])
    df = df.append(df_copy)

    """
    将“销售明细补充”外挂表中的数据补充到数据集中
    """
    # # 读取“销售明细补充”外挂表（不指定读取格式），获取所有时间日期列
    # dateColList = []
    # df_original = pd.read_excel(extraPath)
    # for col in df_original.columns:
    #     if df_original[col].dtype == "datetime64[ns]":
    #         dateColList.append(col)
    # 读取“销售明细补充”外挂表，将时间日期列重新解析为字符串

    # “销售明细补充”外挂表中“销售订单号”列为extraOrderList中的值时为补充数据，其他数据为人工添加的matchCol未匹配到的数据
    df_extra = pd.read_excel(extraPath, dtype=str, parse_dates=dateColList,
                             date_parser=lambda x: myDateParser(x))
    # 删除表头为空的多余列（如“Unnamed: 38”），避免随append进入结果表
    df_extra = df_extra.loc[:, ~df_extra.columns.str.startswith("Unnamed")]
    # 删除因表头重复被pandas自动加".1/.2"后缀的重名列（如“市场类型.1”、“合并前批次.1”），避免随append进入结果表
    dupeSuffixCols = [c for c in df_extra.columns if re.fullmatch(r".+\.\d+", c)]
    if dupeSuffixCols:
        print(f"警告：销售明细补充表存在重复表头，已剔除列：{dupeSuffixCols}，请检查该表表头")
        df_extra = df_extra.drop(columns=dupeSuffixCols)
    df_extra["销售订单号"] = df_extra["销售订单号"].str.strip()
    validDf = df_extra.query("销售订单号.isin(@extraOrderList)")
    initMatchDf = df_extra.query("~销售订单号.isin(@extraOrderList)")
    df = df.append(validDf)

    # 删除df中的”服务产品线“和”服务产品类别“列
    cols_to_drop = [col for col in ["服务产品线", "服务产品类别"] if col in df.columns]
    if cols_to_drop:
        df.drop(columns=cols_to_drop, inplace=True)

    # 清理内存
    gc.collect()
    return df.reset_index(drop=True).fillna(""), initMatchDf


"""
将未处理的移动明细表(新下载的表)处理后合并到汇总表中
"""


# @logfun
# def handleMovementDetail(addfilePath, finalPath, matchDict, dateFlag):
# def handleMovementDetail(addfilePath, finalPath, dateFlag):
def handleMovementDetail(addfilePath, finalPath, dateFlag, finalPathAdd, BO_file):
    """
    :param addfilePath: 读取未处理的物料移动明细表（MHTML）
    :param finalPath: 移动明细表处理结果汇总表
    :param matchDict: BusinessObjects采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}
    :param dateFlag: 文件名的更新日期 %Y/%m/%d
    :param finalPathAdd: 新增物料移动明细表处理结果汇总表
    :param BO_file: BO采购信息表
    :return: 返回汇总表路径
    """
    # 读取物料移动明细表
    df_wl = pd.read_html(addfilePath, header=0)[0].fillna("").astype(str)
    # 当表格仅为表头（无数据行）或缺少“参照”列时，无需处理数据，仅重命名文件并返回
    if df_wl.empty or "参照" not in df_wl.columns:
        newFileName = "物料移动明细汇总_" + dateFlag.replace("/", "") + ".xlsx"
        newFileName_add = "物料移动明细汇总(新增列)_" + dateFlag.replace("/", "") + ".xlsx"
        newFilePath = os.path.join(os.path.dirname(finalPath), newFileName)
        newFilePathAdd = os.path.join(os.path.dirname(finalPathAdd), newFileName_add)
        os.rename(finalPath, newFilePath)
        os.rename(finalPathAdd, newFilePathAdd)
        gc.collect()
        return newFilePath
    # 删除后面的空记录行
    df_wl.drop(df_wl[df_wl["参照"] == ""].index, inplace=True)
    # 将“本位币金额列”转为float类型
    df_wl = df_wl.astype({"本位币金额": float})

    # 生成透视表
    df_pivot = df_wl.pivot_table(index=["参照", "记帐日期", "批次"], columns="MvT", values="本位币金额", aggfunc="sum")

    # 将批次变为列数据并复制
    df_pivot.reset_index(level="批次", inplace=True)
    df_pivot["批次2"] = df_pivot["批次"]
    # 重新排序列顺序
    df_pivot = df_pivot[["批次", "Z29", "批次2", "Z30"]]
    # 修改列名
    newColumns = ["Z29批次", "Z29本位币金额", "Z30批次", "Z30本位币金额"]
    df_pivot.columns = newColumns

# 2024/6/17
    # 生成新增字段透视表
    df_pivot_add = df_wl.pivot_table(index=["参照", "记帐日期", "批次", "物料", "物料号码"], columns="MvT", values="本位币金额", aggfunc="sum")

    # 将批次变为列数据并复制
    df_pivot_add.reset_index(level="批次", inplace=True)
    df_pivot_add["批次2"] = df_pivot_add["批次"]
    # 将物料变为列数据并复制
    df_pivot_add.reset_index(level="物料", inplace=True)
    df_pivot_add["物料2"] = df_pivot_add["物料"]
    # 将物料号码变为列数据并复制
    df_pivot_add.reset_index(level="物料号码", inplace=True)
    df_pivot_add["物料号码2"] = df_pivot_add["物料号码"]
    # 重新排序列顺序
    df_pivot_add = df_pivot_add[["物料", "物料号码", "批次", "Z29", "物料2", "物料号码2", "批次2", "Z30"]]
    # 修改列名
    newColumns_add = ["Z29-转入物料编号", "Z29-转入物料名称", "Z29批次", "Z29本位币金额", "Z30-转入物料编号", "Z30-转入物料名称", "Z30批次", "Z30本位币金额"]
    df_pivot_add.columns = newColumns_add

    # 将Z29/Z30中不属于该类型的批次号变为空
    for col in ["Z29", "Z30"]:
        df_pivot[f"{col}批次"] = df_pivot.apply(
            lambda series: "" if pd.isna(series[f"{col}本位币金额"]) else series[f"{col}批次"], axis=1)
        df_pivot_add[f"{col}批次"] = df_pivot_add.apply(
            lambda series: "" if pd.isna(series[f"{col}本位币金额"]) else series[f"{col}批次"], axis=1)

    # 将数据进行整合处理，每组数据顶部对齐
    newDf = pd.DataFrame()  # 最终的df
    newDf_add = pd.DataFrame()  # 最终的df
    completeList = []  # 由于存在相同的index名，将处理过的index存入list
    indexList = df_pivot.index
    indexList = sorted(indexList, key=lambda x: x[1], reverse=False)  # 按照日期重新排序
    for index_ in indexList:
        if index_ in completeList:
            continue
        # 获取该index的数据
        indexDf = df_pivot.loc[index_, :]
        nowIndex = indexDf.index
        # 重置索引，否则index相同无法依据index删除指定数据
        indexDf = indexDf.reset_index(drop=True)

        indexDf_add = df_pivot_add.loc[index_, :]
        nowIndex_add = indexDf_add.index
        # 重置索引，否则index相同无法依据index删除指定数据
        indexDf_add = indexDf_add.reset_index(drop=True)

        # 判断该index截取的数据是否为多条,需要重新整合格式
        if isinstance(indexDf, pd.core.frame.DataFrame):
            # 获取两个分别只包括Z29/Z30的数据,并合并
            df_temp_z29 = indexDf.drop(index=indexDf[indexDf["Z29批次"] == ""].index)[
                newColumns[:2]].reset_index(drop=True)
            df_temp_z30 = indexDf.drop(index=indexDf[indexDf["Z30批次"] == ""].index)[
                newColumns[2:]].reset_index(drop=True)
        if isinstance(indexDf_add, pd.core.frame.DataFrame):
            # 新增字段表的处理
            df_temp_z29_add = indexDf_add.drop(index=indexDf_add[indexDf_add["Z29批次"] == ""].index)[
                newColumns_add[:4]].reset_index(drop=True)
            df_temp_z30_add = indexDf_add.drop(index=indexDf_add[indexDf_add["Z30批次"] == ""].index)[
                newColumns_add[4:]].reset_index(drop=True)

            temp_df = df_temp_z29.join(df_temp_z30, how="outer")
            temp_df.index = nowIndex[:temp_df.shape[0]]
            newDf = newDf.append(temp_df)

            temp_df_add = df_temp_z29_add.join(df_temp_z30_add, how="outer")
            temp_df_add.index = nowIndex_add[:temp_df_add.shape[0]]
            newDf_add = newDf_add.append(temp_df_add)

        # index数据仅一条，直接合并到结果df
        else:
            newDf = newDf.append(indexDf)
            newDf_add = newDf_add.append(indexDf_add)
        completeList.append(index_)



    # 依据Z29批次新增[下单合同号, 项目名称, 评审二代]列（暂时不添加）
    # newDf_add[matchColadd] = newDf_add["Z29批次"].apply(
    #     lambda x: pd.Series(matchDict.get(x, ["采购信息表无该批次"] * 4)) if not pd.isna(x) else pd.Series(["", "", "", ""]))
    newDf = newDf.fillna("").reset_index()
    newDf_add = newDf_add.fillna("").reset_index()

    # 从BO采购信息表中匹配批次号为Z29对应的合同号、项目名称、评审二代、削价责任人
    df_BO = pd.read_excel(BO_file, dtype=str).fillna("")
    df_BO = df_BO[["批次", "华为_厂商PO号", "项目名称(查询 1 用 系统科技销售管理采购信息)", "签约客户名称", "销售员姓名"]]
    df_BO.columns = ["批次", "下单合同号", "项目名称", "评审二代", "削价责任人"]
    newDf_add = newDf_add.merge(df_BO, left_on="Z29批次", right_on="批次", how="left")
    # 去除表中的批次列
    newDf_add.drop(columns=["批次"], inplace=True)
    #当Z29批次为空时，将"下单合同号", "项目名称", "评审二代", "削价责任人"设置为""
    newDf_add.loc[newDf_add["Z29批次"] == "", matchColadd] = ""
    # 去除表中的重复数据
    newDf_add = newDf_add.drop_duplicates(subset=["Z29-转入物料编号", "Z29-转入物料名称", "Z29批次", "Z29本位币金额",
                                                  "Z30-转入物料编号", "Z30-转入物料名称", "Z30批次", "Z30本位币金额"], keep="last", ignore_index=True)

    # 打开汇总表
    app = xw.App(visible=True, add_book=False)
    app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finalPath)
    wb_add = app.books.open(finalPathAdd)
    ws = wb.sheets["Sheet1"]
    ws_add = wb_add.sheets["Sheet1"]
    # 将新数据补充道汇总表中
    lastRowNum = ws.used_range.shape[0]
    ws.range(f"A{lastRowNum + 1}").value = newDf.values
    lastRowNum = ws_add.used_range.shape[0]
    ws_add.range(f"A{lastRowNum + 1}").value = newDf_add.values

    # # 获取汇总表数据去重（已含表头数据）
    # allValueList = ws.used_range.value
    # totalDf = pd.DataFrame(data=allValueList).drop_duplicates(keep="first")
    #
    # # 情空数据并重新写入汇总表（含表头数据，从A1开始写入即可）
    # ws.used_range.clear_contents()
    # ws.range("A1").value = totalDf.values

    # ws.autofit()
    wb.save(finalPath)
    wb_add.save(finalPathAdd)
    # 关闭工作簿
    wb.close()
    wb_add.close()
    app.quit()

    # 更新文件名
    newFileName = "物料移动明细汇总_" + dateFlag.replace("/", "") + ".xlsx"
    newFileName_add = "物料移动明细汇总(新增列)_" + dateFlag.replace("/", "") + ".xlsx"
    newFilePath = os.path.join(os.path.dirname(finalPath), newFileName)
    newFilePathAdd = os.path.join(os.path.dirname(finalPathAdd), newFileName_add)
    os.rename(finalPath, newFilePath)
    os.rename(finalPathAdd, newFilePathAdd)

    # 清理内存
    gc.collect()
    return newFilePath


# 更新物料移动记录汇总表
# @logfun
def updateMovementDetail(recordDict, finalPath):
    """
    :param recordDict: 开单记录字典{(参照,记帐日期):{年月:金额}}
    :param finalPath: 移动明细表处理结果汇总表
    :return 移动明细表处理结果汇总表路径
    """
    # 转换字典格式 {(参照,记帐日期):{年月:金额}} -> {(参照,记帐日期):[str{年月:金额}]},用于转为DataFrame
    recordDict = {key: [str(value)] for key, value in recordDict.items()}
    recordDictDf = pd.DataFrame(data=recordDict).T.reset_index()
    # logger.info(f"新开单记录为：{recordDictDf}")

    # 打开汇总表
    app = xw.App(visible=True, add_book=False)
    app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finalPath)
    ws = wb.sheets["开单记录"]

    # 情空数据并重新写入汇总表
    ws.used_range.clear_contents()
    ws.range("A1").value = ["参照", "记帐日期", "开单记录"]
    ws.range("A2").value = recordDictDf.values

    # ws.autofit()
    wb.save(finalPath)
    # 关闭工作簿
    wb.close()
    app.quit()

    return finalPath


# 获取目录下相同文件格式的文件列表
def getSameFormatFile(rootDir, keyWord):
    """
    :param rootDir: 需要查找的文件目录
    :param keyWord: 查找的关键词
    :return:
    """
    fileList = glob.glob(f"{rootDir}\\*{keyWord}*.xlsx")
    return fileList


# 新建excel
def creatExcel(filePath):
    """
    :param filePath: 文件保存目录
    :return:
    """
    df = pd.DataFrame()
    df.to_excel(filePath)
    return filePath


# 合并OA预提表
# @logfun
def updateOAFile(addfilePath, finalPath, dateFlag, yearRange):
    """
    :param addfilePath: 读取未处理的OA预提表
    :param finalPath: OA预提汇总表
    :param dateFlag: 文件名的更新日期 %Y-%m-%d
    :param yearRange: 本次下载文件的日期所在年份列表
    :return: 返回OA预提汇总表路径
    """
    # OA预提表在指定日期内有数据（无数据不下载，addfilePath为None）
    if addfilePath:
        # 将年份列表数据转为字符串格式
        yearRange = [str(i) for i in yearRange]

        # 初始化App
        app = xw.App(visible=True, add_book=False)
        app.display_alerts = False
        app.screen_updating = True
        # 打开下载表
        wb_ = app.books.open(addfilePath)
        ws_ = wb_.sheets["Excel"]
        # 删除第一行和最后一行
        rows_ = ws_.used_range.shape[0]
        newCols = ws_.used_range.rows[0].value
        ws_.range(f"{rows_}:{rows_}").delete()
        ws_.range("1:1").delete()

        # 打开汇总表
        wb = app.books.open(finalPath)
        ws = wb.sheets["Excel"]
        oldCols = ws.used_range.rows[0].value
        if oldCols != newCols:
            raise Exception(f"汇总表{finalPath}和下载表{addfilePath}列不同")
        rows, cols = ws.used_range.shape
        # 获取"申请日期"所在的列
        finishCol = get_column_letter(cols)
        firstRowValue = ws.range(f"A1:{finishCol}1").value
        idx = firstRowValue.index("申请日期")
        dataCol = get_column_letter(idx + 1)
        # 预提表数据是以“申请日期”从新到旧排序，查找“申请日期”列，确认“申请日期”属于yearRange的数据量
        targetCol = 0
        dateValList = ws.range(f"{dataCol}2:{dataCol + str(rows)}").value
        for i in range(len(dateValList)):
            val = str(dateValList[i])
            if not val:
                targetCol = targetCol + 1
                continue
            if val[:4] in yearRange:
                targetCol = targetCol + 1
            else:
                break
        # 若存在“申请日期”属于yearRange的数据量，将其所在行删除
        if targetCol != 0:
            ws.range(f"2:{targetCol + 1}").delete()
        # 需要添加的数据行数共有rows-2，先插入空白行
        ws.range(f"2:{rows_ - 1}").insert()
        # 将需要添加的数据复制到汇总表中
        ws_.used_range.copy(destination=ws.range("A2"))

        # ws.autofit()
        wb.save(finalPath)
        # 关闭工作簿
        wb.close()
        wb_.close()
        app.quit()

    # 更新文件名
    newFileName = "预提表_" + dateFlag.replace("-", "") + ".xlsx"
    newFilePath = os.path.join(os.path.dirname(finalPath), newFileName)
    os.rename(finalPath, newFilePath)

    # 清理内存
    gc.collect()
    return newFilePath


# 合并华为订单表
# @logfun
def updateOrderFile(addfilePath, finalPath, dateFlag):
    """
    :param addfilePath: 读取未处理的华为订单表
    :param finalPath: 华为订单汇总表路径
    :param dateFlag: 文件名的更新日期 %Y-%m-%d
    :return: 返回华为订单汇总表路径
    """
    # 华为订单表在指定日期内有数据（无数据不下载，addfilePath为None）
    if addfilePath:
        # 初始化App
        app = xw.App(visible=True, add_book=False)
        app.display_alerts = False
        app.screen_updating = True
        # 打开下载表
        wb_ = app.books.open(addfilePath)
        ws_ = wb_.sheets[0]
        # 删除第一行标题行
        newCols = ws_.used_range.rows[0].value
        ws_.range("1:1").delete()

        # 打开汇总表
        wb = app.books.open(finalPath)
        ws = wb.sheets[0]
        oldCols = ws.used_range.rows[0].value
        if oldCols != newCols:
            raise Exception(f"汇总表{finalPath}和下载表{addfilePath}列不同")
        startRow = ws.used_range.shape[0] + 1

        # 将需要添加的数据复制到汇总表中
        ws_.used_range.copy(destination=ws.range(f"A{startRow}"))

        # ws.autofit()
        wb.save(finalPath)
        # 关闭工作簿
        wb.close()
        wb_.close()
        app.quit()

    # 更新文件名
    fileUser = os.path.basename(finalPath).split("_")[0]
    newFileName = f"{fileUser}_订单表_" + dateFlag.replace("-", "") + ".xlsx"
    newFilePath = os.path.join(os.path.dirname(finalPath), newFileName)
    os.rename(finalPath, newFilePath)

    # 清理内存
    gc.collect()
    return newFilePath


# 依据人工补充的[下单合同号, 项目名称, 评审二代]对数据初始化(暂不使用)
def matchProjectInit(series, initMatchDf):
    """
    :param series: DataFrame行series
    :param initMatchDf:“销售明细补充”外挂表中人工新增的[下单合同号, 项目名称, 评审二代]的数据
    :return: 返回"下单合同号"、"项目名称"、"评审二代"
    """
    orderNum, oneProject = series[["销售订单号", "销售订单行项目"]]
    matchDf = initMatchDf.query("销售订单号 == @orderNum and 销售订单行项目 == @oneProject")
    if matchDf.empty:
        return pd.Series(data=["", "", ""], index=matchCol)
    else:
        matchData = matchDf.iloc[0][matchCol]
        return matchData


# 依据"物料名称"、"批次"匹配"下单合同号"、"项目名称"、"评审二代"
def matchProject1(series):
    """
    :param series: DataFrame行series
    :return: 返回"下单合同号"、"项目名称"、"评审二代"
    """

    # 优先判断折让
    if "折让" in series["合同号（客户PO号）"] or "折让" in series["物料名称"]:
        return pd.Series(data=["折让", "", ""], index=matchCol)

    if "价外费用" in series["物料名称"]:
        if series["物料号"] in ["600-072584", "600-072588", "600-072592", "600-072596", "600-217668"]:
            return pd.Series(data=["罚息", "", ""], index=matchCol)
        elif series["物料号"] in ["600-072585", "600-072589", "600-072593", "600-072597"]:
            return pd.Series(data=["物流费", "", ""], index=matchCol)
        elif series["物料号"] in ["600-072586", "600-072590", "600-072594", "600-072598"]:
            return pd.Series(data=["诉讼费&律师费", "", ""], index=matchCol)
        elif series["物料号"] in ["600-072587", "600-072591", "600-072595", "600-072599"]:
            return pd.Series(data=["折旧费", "", ""], index=matchCol)
        else:
            return pd.Series(data=["价外费用", "", ""], index=matchCol)  # 后面将"价外费用"替换成BO采购信息中的"下单合同号"
    # elif series["批次"].startswith("WG"):
    #     return pd.Series(data=["外购", "", ""], index=matchCol)
    elif series["批次"].startswith("D") and re.match("[A-Z]{3}.*", series["批次"], re.I):
        return pd.Series(data=["样机借转销", "", ""], index=matchCol)
    # elif "NCS" in series["批次"] or "ECAS" in series["批次"]:
    #     return pd.Series(data=["外购", "", ""], index=matchCol)
    elif series["物料号"][:2] in ["80", "81"] and series["产品组"] == "HT":
        return pd.Series(data=["自有服务", "", ""], index=matchCol)
    elif series["批次"] == "":
        return pd.Series(data=["未匹配原因：批次号为空"] * 3, index=matchCol)
    else:
        return pd.Series(data=["", "", ""], index=matchCol)


# 依据"批次"匹配"下单合同号"、"项目名称"、"评审二代"
def matchProject2(key, matchDict, colList):
    """
    :param key: 批次数据
    :param matchDict: BusinessObjects采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}
    :param colList: 数据列名
    :return: 返回"下单合同号"、"项目名称"、"评审二代"
    """

    searchValue = matchDict.get(key, ["", "", ""])
    return pd.Series(data=searchValue, index=colList)


# 依据"批次"和物料移动明细匹配"下单合同号"、"项目名称"、"评审二代"
def matchProject3(series, df_Z30List, df_Move, matchDict):
    """
    :param series: DataFrame行series
    :param df_Z30List: 物料移动明细汇总表总Z30批次列表
    :param df_Move: 物料移动明细汇总df
    :param matchDict: BusinessObjects采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}
    :return: 返回"下单合同号"、"项目名称"、"评审二代"
    """
    # DataFrame.apply第一行数据会操作两次，导致重复开单，加入matchFlag忽略第一次操作
    global matchFlag
    if matchFlag == False:
        matchFlag = True
        return pd.Series(data=["无效数据（忽略）"] * 4, index=addCol)

    z30bacthNum = series["批次"]
    custonName = series["客户名称"]
    count = df_Z30List.count(z30bacthNum)
    """
    获取需要操作的合并后批次号在df_Z30List（物料移动明细汇总）中出现的次数，依据次数进行操作：
    1.次数=0，说明为匹配到改批次
    2.次数=1，说明物料移动明细汇总表中仅有一组数据与其匹配，返回需要处理的index
    3.次数>1，物料移动明细汇总表有多组数据匹配，需要筛选后确定index
    """
    if count == 0:
        return pd.Series(data=["未匹配原因：未匹配到批次"] * 3 + [""], index=addCol)
    elif count == 1:
        index = df_Move[df_Move["Z30批次"] == z30bacthNum].index[0]
    else:
        """
        多组数据匹配时，查看各组数据对应的“评审二代”，与原数据的“客户名称”匹配时说明有效
        """
        indexList = df_Move[df_Move["Z30批次"] == z30bacthNum].index
        matchCustomNum = 0
        finalIdx = 0

        for idx in indexList:
            tempDf = df_Move.loc[idx].copy()
            tempDf["评审二代"] = tempDf["Z29批次"].apply(lambda x: matchDict.get(x, [""] * 3)[2])
            if isinstance(tempDf, pd.core.frame.DataFrame):
                customList = tempDf["评审二代"].tolist()
            else:
                customList = tempDf.to_frame().T["评审二代"].tolist()
            if custonName in customList:
                finalIdx = idx
                matchCustomNum += 1
        if matchCustomNum == 0:
            return pd.Series(data=["未匹配原因：匹配到多个批次但均无法匹配客户名称"] * 3 + [""], index=addCol)
        elif matchCustomNum == 1:
            index = finalIdx
        else:
            return pd.Series(data=[f"未匹配原因：匹配到多个批次且有{matchCustomNum}组均匹配客户名称"] * 3 + [""], index=addCol)

    # 获取到index后，依据移动明细数据、开单记录匹配到"下单合同号"、"项目名称"、"评审二代"
    return matchingMovementTable(series, index, df_Move, matchDict)


# 依据移动明细数据、开单记录、本次开单金额匹配"下单合同号"、"项目名称"、"评审二代"
def matchingMovementTable(series, index_, df_Move, matchDict):
    """
    :param series: DataFrame行series
    :param index_: 物料移动明细Df需要操作的index
    :param df_Move: 物料移动明细汇总df
    :param matchDict: BusinessObjects采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}
    :return: 返回"下单合同号"、"项目名称"、"评审二代"
    """

    # 获取剩余额度
    def initOriginalData(usedAmount, originalDict):
        """
        :param usedAmount: 已使用的额度
        :param originalDict: 总额度字典{批次：总金额}
        :return: 剩余额度列表[[批次,剩余金额]]
        """
        initDict = deepcopy(originalDict)
        resetOrder = sorted(initDict.keys(), reverse=False)
        initList = [[i, initDict[i]] for i in resetOrder]  # 按照批次号正向排序
        for valList in initList:
            if usedAmount > 0:
                amount = valList[1]
                if amount <= usedAmount:
                    valList[1] = 0
                    usedAmount = new_round(usedAmount - amount, 2)
                else:
                    valList[1] = new_round(amount - usedAmount, 2)
                    usedAmount = 0
            else:
                break
        return initList

    # 获取本次使用额度明细(额度>=0)
    def getData(initList, useAmount):
        """
        :param initList: 额度列表[[批次,剩余金额]]
        :param useAmount: 需要使用的额度
        :return: resultDict: 本次使用的字典{使用批次：使用金额}
        """
        # resetOrder = sorted(initDict.keys(), reverse=False)
        # initDict = {i: initDict[i] for i in resetOrder}  # 按照批次号正向排序
        lastBatchNum = ""
        resultDict = {}
        for key, v in initList:
            lastBatchNum = key
            if v == 0:
                continue
            if useAmount > 0:
                if v <= useAmount:
                    resultDict[key] = v
                    useAmount = new_round(useAmount - v, 2)

                else:
                    resultDict[key] = useAmount
                    useAmount = 0
            else:
                break
        # 处理useAmount为0的情况，优先分配有额度的批次
        if resultDict == {}:
            resultDict[lastBatchNum] = 0
        # 处理额度不足的情况，将超出的金额附加在最后一个批次上
        if useAmount > 0:
            resultDict[lastBatchNum] = resultDict.get(lastBatchNum, 0) + useAmount
        return resultDict

    # 获取本次使用额度明细(额度<0)
    def getDataReverse(initList, useAmount, originalDict):
        """
       :param initList: 额度列表[[批次,剩余金额]]
       :param useAmount: 需要退回的的额度
       :param: originalDict: 原始的额度字典，用于获取每个批次的最大额度
       :return: resultDict: 本次使用的字典{使用批次：使用金额}
        """
        # resetOrder = sorted(initDict.keys(), reverse=True)
        # initDict = {i: initDict[i] for i in resetOrder}  # 按照批次号反向排序
        initList = initList[::-1]
        lastBatchNum = ""
        resultDict = {}
        useAmount = -useAmount
        for key, value in initList:
            lastBatchNum = key
            if useAmount > 0:
                diff = originalDict[key] - value
                if diff == 0:
                    continue
                if diff > useAmount:
                    resultDict[key] = -useAmount
                    useAmount = 0
                else:
                    resultDict[key] = -diff
                    useAmount = new_round(useAmount - diff, 2)
            else:
                break

        # 处理回填额度时超过原有最大值的情况，将超出的金额附加在最后处理的批次上
        if useAmount > 0:
            resultDict[lastBatchNum] = resultDict.get(lastBatchNum, 0) - useAmount
        return resultDict

    global addDataFrame, recordDict
    # 获取需要处理批次的df
    needCalDf = df_Move.loc[index_]
    if isinstance(needCalDf, pd.core.series.Series):
        needCalDf = needCalDf.to_frame().T
    # 重置索引，防止后续依据index删除空行时删除全部数据
    needCalDf = needCalDf.reset_index()
    dropIndex = needCalDf[needCalDf["Z29批次"].isna()].index.tolist()
    if dropIndex:
        needCalDf = needCalDf.drop(dropIndex)
    if needCalDf.empty:
        return pd.Series(data=["未匹配原因：无对应合并前批次"] * 3 + [""], index=addCol)

    # 获取到移动明细中的额度字典{批次：总金额}
    originalDict = dict(zip(needCalDf["Z29批次"], needCalDf["Z29本位币金额"].apply(lambda x: new_round(x, 2))))
    # 获取开单记录中该参照的已使用额度字典{批次：使用额度}
    usedDict = recordDict.get(index_, {})
    usedAmount = sum(usedDict.values())

    # 获取到各批次剩余额度 #
    initList = initOriginalData(usedAmount, originalDict)
    # 本次需要开单金额
    useAmount = new_round(float(series[calAmountCol]), 2)

    # 获取开单明细字典
    if useAmount >= 0:
        resultDict = getData(initList, useAmount)
    else:
        resultDict = getDataReverse(initList, useAmount, originalDict)

    # 更新批次使用额度字典
    sumValue = usedDict.get(calYearMonth, 0) + useAmount
    usedDict[calYearMonth] = new_round(sumValue, 2)
    recordDict[index_] = usedDict

    """
    判断使用额度字典中批次数量，大于1需要拆行，然后匹配"下单合同号"、"项目名称"、"评审二代"
    """
    if len(resultDict) == 1:
        num = list(resultDict.keys())[0]
        # usedData = needCalDf.query("Z29批次 == @num").iloc[0]  # series
        # return usedData[matchCol]  # series
        matchData = matchDict.get(num, ["未匹配原因：采购信息表无该批次"] * 3) + [num]

        return pd.Series(matchData, index=addCol)
    else:  # len>1
        """
        拆行处理：
        1.将"合同金额", "合同不含税金额"按照开单金额百分比计算，匹配批次对应的"下单合同号"、"项目名称"、"评审二代"
        2."返款折扣×汇率"列第一行数据保持不变，后续拆行的数据为空
        3.拆行的数据保存到addDataFrame中，原数据标记为"已拆行处理"
        """
        flag = 0
        for num, amount in resultDict.items():
            # usedData = needCalDf.query("Z29批次 == @num").iloc[0]  # series
            percent = amount / useAmount
            tempSeries = series.copy()
            # 增加"下单合同号"、"项目名称"、"评审二代"列
            matchData = matchDict.get(num, ["未匹配原因：采购信息表无该批次"] * 3) + [num]
            tempSeries[addCol] = pd.Series(matchData, index=addCol)
            # 对金额相关行按金额比例划分
            tempSeries[calAmountCol] = amount
            tempSeries[["合同金额", "合同不含税金额"]] = tempSeries[["合同金额", "合同不含税金额"]].astype(
                "float") * percent
            if flag != 0:  # 第一组数据金额不变，拆分后新增行的金额清空
                tempSeries[["返款折扣×汇率"]] = ""
            addDataFrame = addDataFrame.append(tempSeries)
            flag += 1

        else:
            return pd.Series(data=["已拆行处理"] * 4, index=addCol)


# 匹配“采购类型”
# 匹配"采购类型"
def matchCGLX(series):
    """
    :param series: DataFrame行series
    :return: 返回"采购类型"
    """
    series = series.copy().fillna("")
    if series["下单合同号"] == "折让":
        return "折让"
    elif "价外费用" in series["物料名称"]:
        return "价外费用"
    elif (series["下单合同号"] == "外购" or series["批次"].startswith("WG") or "NCS" in series["批次"] or "ECAS" in series["批次"]
          or (series["批次"]==''and (not series["下单合同号"].startswith("CY") or not series["下单合同号"].startswith("1Y")))):
        return "外购"
    elif series["下单合同号"] == "样机借转销":
        return "渠道分销"
    elif (series["下单合同号"] and not series["下单合同号"].startswith("CY") and not series["下单合同号"].startswith("1Y") and
          not '\u4e00' <= series["下单合同号"][0] <= '\u9fa5' and not series["采购类型"] == "外购"):
        return "内部采购"
    elif series["下单合同号"].startswith("CY"):
        return "内部采购"
    else:
        return "原厂下单"

# 匹配"采购类型-二级分类"
def matchCGLX_2(series):
    """
    :param series: DataFrame行series
    :return: 返回"采购类型"
    """
    series = series.copy().fillna("")
    if "价外费用" in series["物料名称"]:
        if series["物料号"] in ["600-072584", "600-072588", "600-072592", "600-072596", "600-217668"]:
            return "罚息"
        elif series["物料号"] in ["600-072585", "600-072589", "600-072593", "600-072597"]:
            return "物流费"
        elif series["物料号"] in ["600-072586", "600-072590", "600-072594", "600-072598"]:
            return "诉讼费&律师费"
        elif series["物料号"] in ["600-072587", "600-072591", "600-072595", "600-072599"]:
            return "折旧费"
        else:
            return "价外费用"
    elif (series["下单合同号"] and not series["下单合同号"].startswith("CY") and not series["下单合同号"].startswith("1Y") and
          not '\u4e00' <= series["下单合同号"][0] <= '\u9fa5' and not series["采购类型"] == "外购"):
        return "鲲泰"
    elif series["下单合同号"].startswith("CY"):
        return "超聚变"
    else:
        return ""


# 匹配“运输方式”
def matchYSFS(series, BoTransDict, HWOrderDict):
    """
    :param series: DataFrame行series
    :param BoTransDict: BO下单合同号对应的运输方式字典{下单合同号：运输方式}
    :param HWOrderDict: 华为订单表字典{下单合同号：运输方式}
    :param KTOrderDict: 鲲泰外挂表字典{下单合同号：运输方式}
    :return: 返回"运输方式"
    """
    purchaseType = series["采购类型"]
    orderNum = series["下单合同号"]
    if purchaseType == "内部采购" or purchaseType == "外购":
        # return KTOrderDict.get(orderNum, "鲲泰外挂表未匹配到")
        # --20250429 参考"库存地"字段最后两个字符判断
        stock_last2 = series["库存地"][-2:]
        if stock_last2 == "99":
            return "汽运"
        elif stock_last2 == "01" or stock_last2 == "XN" or stock_last2 == '':
            return "自提"
    elif purchaseType in ["公有云", "服务", "服务预提", "渠道分销", "价外费用", "折让"]:
        return "自提"
    # if purchaseType == "超聚变":
    #     return BoTransDict.get(orderNum, "自提")
    else:  # todo:华为订单表未匹配到的默认值为"自提"
        return HWOrderDict.get(orderNum, "华为订单表未匹配到")


# 计算“市场类型”列
def calMarketType(series):
    """
    :param series: DataFrame行series
    :return: 返回”市场类型”
    """

    if series["业务部"] in ["安平系统部", "数字政府系统部"]:
        return "政府安平"
    elif series["客户分类"] in ["核心NA", "战略NA", "卓越NA"]:
        return "NA"
    elif series["客户分类"] in ["价值NA", "商业市场客户"]:
        return "商业市场"
    else:
        return ""

# 计算“市场类型”列
def calMarketType1(series):
    """
    :param series: DataFrame行series
    :return: 返回”市场类型”
    """
    if series["业绩业务类型"] == "中国区政企" and series["客户分类"] in ["核心NA", "战略NA", "卓越NA"]:
        return "NA"
    elif series["业绩业务类型"] == "中国区政企" and series["客户分类"] == "价值NA":
        return "VNA"
    elif series["业绩业务类型"] == "中国区政企" and series["客户分类"] in ["商业市场客户", ""]:
        return "SMA"
    else:
        return ""

# 匹配”产品”、”产品线”、“市场类型”
def matchProduct(series, confDict):
    """
    :param series: DataFrame行series
    :param confDict: 字典{下单合同号:[产品，产品线，市场类型]}
    :return: 返回”产品”、”产品线”、“市场类型”
    """
    purchaseType = series["采购类型"]
    contractNum = series["下单合同号"]
    productGroup = series["产品组"]
    # 显式指定返回Series的index，避免依赖pandas版本差异下的按位置对齐行为
    resultIndex = ["产品", "产品线", "市场类型"]

    if purchaseType == "鲲泰":
        return pd.Series(data=["鲲泰", "鲲泰", ""], index=resultIndex)
    elif productGroup == "QJ":
        return pd.Series(data=["超聚变", "超聚变", ""], index=resultIndex)
    else:
        return pd.Series(data=confDict.get(contractNum, ["不计入业绩", "不计入业绩", ""]), index=resultIndex)


# 生成月份（不对na和空值处理）
def calMonth(x):
    """
    :param x: 传入的时间（例：2022/05/31）
    :return: 返回月份数据（例：5月）
    """
    if x and not pd.isna(x):
        return str(int(x[5:7])) + "月"
    else:
        return ""


# @logfun
# 依据"物料名称"、"批次"的内容初步匹配"下单合同号"、"项目名称"、"评审二代"
def calDataStep1(df: pd.DataFrame):
    """
    :param df:需要处理的df
    :return:处理后的df
    """
    # 依据"物料名称"、"批次"的内容初步匹配"下单合同号"、"项目名称"、"评审二代"
    df.loc[df["下单合同号"] == '', matchCol] = df.loc[df["下单合同号"] == ''].apply(matchProject1, axis=1)
    # # 判断是否有数据未匹配到
    # flag = df.query("下单合同号 == ''").empty

    return df


# @logfun
# 通过批次匹配BO采购信息表中的"下单合同号"、"项目名称"、"评审二代"
def calDataStep2(df, BO_file):
    """
    :param df: 需要处理的df
    :param BO_file: BusinessObjects采购信息下载路径
    :return:处理后的df，BO采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}，BO下单合同号对应的运输方式字典{下单合同号：运输方式}
    """
        # 最小新增：折让数据隔离，不参与BO匹配避免覆盖
    mask_zhe = df["下单合同号"] == "折让"
    df_zhe = df[mask_zhe].copy()
    df = df[~mask_zhe]
    
    # reg = "|".join(ignoreCharsList)
    reg = "".join(ignoreCharsList)
    # 读取BO下载的采购信息，忽略“华为_厂商PO号”部分特殊字符，将长度过低的数据去除
    df_BO = pd.read_excel(BO_file, dtype=str).fillna("")
    # df_BO["华为_厂商PO号"] = df_BO["华为_厂商PO号"].str.replace(reg, "")
    df_BO["华为_厂商PO号"] = df_BO["华为_厂商PO号"].str.strip(reg)
    dfBo_copy = df_BO[["华为_厂商PO号", "发货方式"]].copy()  # 用于生成{下单合同号：运输方式}字典
    df_BO = df_BO.query("华为_厂商PO号.str.len() >= 4 and 批次 != '' ")
    print(f"BO采购信息表大小为：{df_BO.shape}")

    """生成{批次：[下单合同号, 项目名称, 评审二代]}字典"""
    # 去重（保留最后一条数据）并匹配到销售明细df中
    df_BO_pc = df_BO.drop_duplicates("批次", keep="last", ignore_index=True)
    matchDict = dict(zip(df_BO_pc["批次"], df_BO_pc[boMatchCol].values.tolist()))
    # # --20250813
    # df_BO_po = df_BO.drop_duplicates("华为_厂商PO号", keep="last", ignore_index=True)
    # po_matchDict = dict(zip(df_BO_po["华为_厂商PO号"], df_BO_po[["项目名称(查询 1 用 系统科技销售管理采购信息)", "签约客户名称"]].values.tolist()))
    #
    # # logger.info(matchDict)
    # # df.loc[df["下单合同号"] == '', matchCol] = df['批次'].apply(matchProject2, args=(matchDict, matchCol,))
    # df.loc[df["下单合同号"] == '', matchCol] = df.loc[df["下单合同号"] == '', "批次"].apply(lambda x: pd.Series(data=matchDict.get(x, ["", "", ""]), index=matchCol))
    #
    # # --20250807 下单合同号 新增批次和区域的匹配
    # df.loc[df["批次"] == '', ["下单合同号"]] = df.loc[df["批次"] == '', "项目注释"].apply(lambda x: str(x).split(';')[0] if pd.notna(x) else "")
    # df.loc[df["区域"] == '创新业务', ["下单合同号"]] = df.loc[df["区域"] == '创新业务', "项目注释"].apply(lambda x: str(x).split(';')[0] if pd.notna(x) else "")
    # # --20250429 新增”价外费用“的填充，然后替换”价外费用“，并删除”项目注释“列
    # # df.loc[df["下单合同号"] == '价外费用', ["项目名称", "评审二代"]] = df.loc[df["下单合同号"] == '价外费用', "批次"].apply(lambda x: pd.Series(data=matchDict.get(x, ["", "", ""])[1:], index=["项目名称", "评审二代"]))
    # # df.loc[df["下单合同号"] == '价外费用', ["下单合同号"]] = df.loc[df["下单合同号"] == '价外费用', "项目注释"].apply(lambda x: str(x).split(';')[0] if pd.notna(x) else "")
    # mask = df["下单合同号"] == '价外费用'
    # if mask.any():
    #     df.loc[mask, ["项目名称", "评审二代"]] = df.loc[mask, "批次"].apply(lambda x: pd.Series(data=matchDict.get(x, ["", "", ""])[1:], index=["项目名称", "评审二代"]))
    #     df.loc[mask, ["下单合同号"]] = df.loc[mask, "项目注释"].apply(lambda x: str(x).split(';')[0] if pd.notna(x) else "")
    # df.loc[df["项目名称"] == '', ["项目名称", "评审二代"]] = df.loc[df["项目名称"] == '', "下单合同号"].apply(lambda x: pd.Series(data=po_matchDict.get(x, ["", ""]), index=["项目名称", "评审二代"]))
    # df.drop(columns=["项目注释"], errors='ignore', inplace=True)

    """生成{下单合同号：运输方式}字典"""
    dfBo_copy["发货方式"] = dfBo_copy["发货方式"].map(lambda x: BoDeliveryDict.get(x, ""))
    dfBo_copy = dfBo_copy.drop_duplicates("华为_厂商PO号", keep="last", ignore_index=True)
    BoTransDict = dict(zip(dfBo_copy["华为_厂商PO号"], dfBo_copy["发货方式"]))

    # # 判断是否有数据未匹配到
    # flag = df.query("下单合同号 == ''").empty
    # 合并折让数据
    df = pd.concat([df, df_zhe], ignore_index=True)

    # 清理内存
    gc.collect()
    return df, matchDict, BoTransDict

def calDataStep2_crm(df, crm_file, matchDict):
    """
    :param df: 需要处理的df
    :param crm_file: CRM外挂表路径，包含"商机名称（必填）"和"关联的厂商PO号"两列
    :param matchDict: BO采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}
    :return: 处理后的df
    """
    mask_zhe = df["下单合同号"] == "折让"
    df_zhe = df[mask_zhe].copy()
    df = df[~mask_zhe]

    # 读取CRM外挂表，拆分"关联的厂商PO号"中用";"隔开的多个PO号
    df_crm = pd.read_excel(crm_file, dtype=str).fillna("")
    df_crm["关联的厂商PO号"] = df_crm["关联的厂商PO号"].str.replace("；", ";")  # 统一中英文分号
    df_crm = df_crm.assign(**{"关联的厂商PO号": df_crm["关联的厂商PO号"].str.split(";")}).explode("关联的厂商PO号")
    df_crm["关联的厂商PO号"] = df_crm["关联的厂商PO号"].str.strip()
    df_crm = df_crm[df_crm["关联的厂商PO号"] != ""]
    # 生成{PO号: 商机名称}字典，去重保留最后一条
    df_crm = df_crm.drop_duplicates("关联的厂商PO号", keep="last", ignore_index=True)
    crm_dict = dict(zip(df_crm["关联的厂商PO号"], df_crm["商机名称（必填）"]))
    print(f"CRM外挂表拆分后大小为：{df_crm.shape}")

    # "下单合同号"取值：用"项目注释"列填充
    # 条件：物料名称包含价外费用 OR 批次空白 OR 批次WG开头 OR 批次含NCS OR 批次含ECAS OR 批次KT开头
    fill_mask = (
        df["物料名称"].str.contains("价外费用", na=False) |
        (df["批次"] == '') |
        df["批次"].str.startswith("WG", na=False) |
        df["批次"].str.contains("NCS", na=False) |
        df["批次"].str.contains("ECAS", na=False) |
        df["批次"].str.startswith("KT", na=False)
    )
    df.loc[fill_mask, ["下单合同号"]] = df.loc[fill_mask, "项目注释"].apply(
        lambda x: str(x).split(';')[0] if pd.notna(x) else "")

    # 下单合同号还是空时，按BO表的批次来匹配
    # df.loc[df["下单合同号"] == '', "下单合同号"] = df.loc[df["下单合同号"] == '', "批次"].apply(lambda x: matchDict.get(x, [""])[0])
    df.loc[df["下单合同号"] == '', matchCol] = df.loc[df["下单合同号"] == '', "批次"].apply(lambda x: pd.Series(data=matchDict.get(x, ["", "", ""]), index=matchCol))

    # "项目名称"和"评审二代"仅在fill_mask范围内取值
    # "项目名称"：用"下单合同号"关联CRM的"关联的厂商PO号"匹配"商机名称（必填）"
    df.loc[fill_mask, "项目名称"] = df.loc[fill_mask, "下单合同号"].map(crm_dict).fillna("")
    # "评审二代"：直接取df中的"客户名称"
    df.loc[fill_mask, "评审二代"] = df.loc[fill_mask, "客户名称"]

    df = pd.concat([df, df_zhe], ignore_index=True)

    # 合并折让数据后再删除“项目注释”列，避免df_zhe将该列重新引入
    df.drop(columns=["项目注释"], errors='ignore', inplace=True)

    # 清理内存
    gc.collect()
    return df


# @logfun
# 通过物料移动明细表匹配"下单合同号"、"项目名称"、"评审二代"
def calDataStep3(df: pd.DataFrame, moveFilePath, matchDict):
    """
    :param df: 需要处理的df
    :param moveFilePath:物料移动明细汇总表
    :param matchDict: BusinessObjects采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}
    :return: 处理后的df
    """

    mask_zhe = df["下单合同号"] == "折让"
    df_zhe = df[mask_zhe].copy()
    df = df[~mask_zhe]

    global addDataFrame, recordDict, df_Move

    addDataFrame = pd.DataFrame()  # 记录切割后的数据（需要插入到数据集中）、

    # 读取的物流移动记录汇总表，包含移动汇总数据和开单明细
    moveAllDf = pd.read_excel(moveFilePath, sheet_name=None, dtype={"参照": str, "Z29批次": str, "Z30批次": str})
    df_Move = moveAllDf["Sheet1"].set_index(keys=["参照", "记帐日期"]).sort_index()
    moveRecordDf = moveAllDf["开单记录"].set_index(keys=["参照", "记帐日期"])

    # 物料移动汇总中合并后批次列表
    df_Z30List = df_Move["Z30批次"].tolist()
    # 物料开单记录字典{(参照,记帐日期):{年月:金额}}
    recordDict = dict(zip(moveRecordDf.index, moveRecordDf["开单记录"].fillna("{}")))
    # 将物料开单记录字典中本月数据清空，防止重复开单
    for key, value in recordDict.items():
        value_ = eval(value)
        if calYearMonth in value_.keys():
            value_[calYearMonth] = 0
        recordDict[key] = value_

    # 依据"批次"和物料移动明细匹配"下单合同号"、"项目名称"、"评审二代"
    global matchFlag
    matchFlag = False
    df[addCol[-1]] = ""
    df.loc[df["下单合同号"] == '', addCol] = df.loc[df["下单合同号"] == ''].apply(matchProject3, args=(df_Z30List, df_Move, matchDict), axis=1)

    # 将标记为"已拆行处理"的原数据删除并将拆行后的数据集addDataFrame添加到原数据中
    dropIndex = df[df["下单合同号"] == "已拆行处理"].index.tolist()
    if dropIndex:
        df.drop(dropIndex, inplace=True)
    df = df.append(addDataFrame)
    """
    1.由于拆行后数据的index和原本相同，将index转为列进行排序会自动保持原有数据顺序
    2.对"返款折扣×汇率"再次排序是保证拆行后的数据顺序不变，同组的第一条数据的"返款折扣×汇率"不为空
    3.重新排序后删除添加的"index"行
    """
    df.reset_index(inplace=True)
    df.sort_values(by=["index", "返款折扣×汇率"], ascending=[True, False], inplace=True)
    # df.sort_values(by=["index","下单合同号"],ascending=[True, True],inplace=True)
    df.drop(columns="index", inplace=True)
    # df.to_excel(r"C:\Users\11598\Desktop\out.xlsx", index=False)

    # 更新开单记录
    updateMovementDetail(recordDict, moveFilePath)
    # # 判断是否有数据未匹配到
    # flag = df.query("下单合同号 == ''").empty

    df = pd.concat([df, df_zhe], ignore_index=True)

    # 清理内存
    gc.collect()

    return df


# 将人工补充的[下单合同号, 项目名称, 评审二代]填入销售明细数据中
def matchExtraData(df: pd.DataFrame, initMatchDf: pd.DataFrame):
    """
    :param df: 需要处理的df
    :param initMatchDf:“销售明细补充”外挂表中人工新增的[下单合同号, 项目名称, 评审二代]的数据
    :return: 处理后的df
    """

    mask_zhe = df["下单合同号"] == "折让"
    df_zhe = df[mask_zhe].copy()
    df = df[~mask_zhe]

    # 将数据集及人工匹配数据的calAmountCol列转为数字类型
    df[calAmountCol] = pd.to_numeric(df[calAmountCol], errors='coerce')
    initMatchDf[calAmountCol] = pd.to_numeric(initMatchDf[calAmountCol], errors='coerce')

    """
    1.为防止以["销售订单号", "销售订单行项目", calAmountCol]为准按行处理数据时，出现1对多、多对一、多对多的情况，因此将上述三列作为基准去匹配
    2.先判断相同基准下是否有两边数据量不对等的情况，有的话直接抛异常
    3.在2的基础上，将人工补充的[下单合同号, 项目名称, 评审二代]数据替换到数据集中
    """
    dataList = initMatchDf[["销售订单号", "销售订单行项目", calAmountCol]].values.tolist()
    dataSet = set([tuple(i) for i in dataList])
    for orderNum, oneProject, amount in dataSet:
        df_Match = initMatchDf.loc[(initMatchDf["销售订单号"] == orderNum) & (initMatchDf["销售订单行项目"] == oneProject) & (initMatchDf[calAmountCol] == amount)]
        target_df = df.loc[(df["销售订单号"] == orderNum) & (df["销售订单行项目"] == oneProject) & (df[calAmountCol] == amount)]
        if df_Match.shape[0] != target_df.shape[0]:
            df_Match_ = initMatchDf.loc[(initMatchDf["销售订单号"] == orderNum) & (initMatchDf["销售订单行项目"] == oneProject)][["销售订单号", "销售订单行项目", calAmountCol]]
            target_df_ = df.loc[(df["销售订单号"] == orderNum) & (df["销售订单行项目"] == oneProject)][["销售订单号", "销售订单行项目", calAmountCol]]
            raise Exception(
                f"销售订单号为{orderNum}，销售订单行项目为{oneProject}，{calAmountCol}为{amount}的数据补充表匹配异常\n"
                f"补充表数据为{df_Match_.values.tolist()},匹配到的数据为{df_Match[['销售订单号', '销售订单行项目', calAmountCol]].values.tolist()}\n"
                f"销售明细表对应数据为{target_df_.values.tolist()}，匹配到的数据为{target_df[['销售订单号', '销售订单行项目', calAmountCol]].values.tolist()}")
        df.loc[(df["销售订单号"] == orderNum) & (df["销售订单行项目"] == oneProject) & (df[calAmountCol] == amount), matchCol] = df_Match[matchCol].values
    df = pd.concat([df, df_zhe], ignore_index=True)
    return df


# 匹配“采购类型”、“销售类型”、“运输方式”、“是否直发”
def calDataStep4(df: pd.DataFrame, BoTransDict, HWOrderPathList, KTconfigPath):
    """
    :param df: 需要处理的df
    :param BoTransDict: BO下单合同号对应的运输方式字典{下单合同号：运输方式}
    :param HWOrderPathList: 华为订单表路径列表
    :param KTconfigPath: 鲲泰跟踪表路径，用于匹配鲲泰运输方式 (弃用)
    :return: 处理后的df
    """
    # 匹配采购类型
    # df["采购类型"] = df.apply(matchCGLX, axis=1)
    df.loc[df["采购类型"] == '', "采购类型"] = df.loc[df["采购类型"] == ''].apply(matchCGLX, axis=1)
    # 在"采购类型"后增加一列"采购类型-二级分类"，仅对该列为空的行取值
    if "采购类型-二级分类" not in df.columns:
        col_position = df.columns.get_loc("采购类型") + 1
        df.insert(col_position, "采购类型-二级分类", "")
    df.loc[df["采购类型-二级分类"] == '', "采购类型-二级分类"] = df.loc[df["采购类型-二级分类"] == ''].apply(matchCGLX_2, axis=1)

    '''
    # 匹配销售类型
    """
    Step1.合同号包含”折让”，销售类型=折让
    Step2.筛选合同金额<0的所有合同号：
        a. 单一下单合同号下，合同金额=0且出具发票日相同，销售类型标“冲红”
        b. 单一下单合同号下，合同金额=0但出具发票日不同，合同金额>0的填写“正常销售”，合同金额<0的填写“退货”
        c. 单一下单合同号下，合同金额<0，销售类型标“退货”
        d. 单一下单合同号下，合同金额>0，销售类型标“正常销售”
    Step3.其余数据标记为“正常销售”
    """
    # Step1
    df["销售类型"] = df["下单合同号"].apply(lambda x: "折让" if "折让" in x else "")
    # Step2
    df["合同金额"] = df["合同金额"].astype(float)
    contractNum = set(df.query("合同金额 < 0 and 销售类型=='' ")["下单合同号"].tolist())
    for num in contractNum:
        df_temp = df.query("下单合同号 == @num")
        totalAmount = df_temp["合同金额"].sum()
        if abs(totalAmount) <= 0.1:
            dataSet = set(df_temp["出具发票日"].tolist())  # 出具发票日总条数
            if len(dataSet) == 1:
                df.loc[df["下单合同号"] == num, "销售类型"] = "冲红"
            else:
                df.loc[(df["下单合同号"] == num) & (df["合同金额"] > 0), "销售类型"] = "正常销售"
                df.loc[(df["下单合同号"] == num) & (df["合同金额"] < 0), "销售类型"] = "退货"
        elif totalAmount < 0:
            df.loc[df["下单合同号"] == num, "销售类型"] = "退货"
        else:
            df.loc[df["下单合同号"] == num, "销售类型"] = "正常销售"
    # Step3
    df.loc[df["销售类型"] == "", "销售类型"] = "正常销售"
    '''

    '''
    # 匹配销售类型
    """
    Step1.合同号包含”折让”，销售类型=折让
    Step2.判断各合同号下各出具发票日下的“合同金额”总和
        1. 若“合同金额”总和在-0.1-0.1之间，销售类型标“冲红”
        2. 若“合同金额”总和 < -0.1，销售类型标“退货”
        3. 若“合同金额”总和 > 0.1，销售类型标“正常销售”
    """
    # Step1
    df["销售类型"] = df["下单合同号"].apply(lambda x: "折让" if "折让" in x else "")
    # Step2
    df["合同金额"] = df["合同金额"].astype(float)
    contractNum = set(df.query("销售类型=='' ")["下单合同号"].tolist())
    for num in contractNum:
        df_ = df.query("下单合同号 == @num")
        allDate = set(df_["出具发票日"].tolist())
        for date in allDate:
            df_temp = df_.query("出具发票日 == @date")
            totalAmount = df_temp["合同金额"].sum()
            if abs(totalAmount) <= 0.1:
                df.loc[(df["下单合同号"] == num) & (df["出具发票日"] == date), "销售类型"] = "冲红"
            elif totalAmount < 0:
                df.loc[(df["下单合同号"] == num) & (df["出具发票日"] == date), "销售类型"] = "退货"
            else:
                df.loc[(df["下单合同号"] == num) & (df["出具发票日"] == date), "销售类型"] = "正常销售"
    # # 新增一列“销售类型-二级分类”，值复制"销售类型"列
    # col_pos = df.columns.get_loc("销售类型") + 1
    # # 在指定位置插入新列
    # df.insert(col_pos, "销售类型-二级分类", df["销售类型"])
    '''

    # '''
    # 匹配销售类型
    """
    Step1.合同号包含”折让”，销售类型=折让
    Step2.筛选合同金额<0的合同号，按合同号+客户+销售员维度处理：
        1. 若“合同金额”总和在-0.1-0.1之间，销售类型标“冲红”
        2. 若“合同金额”总和 < -0.1，销售类型标“退货”
        3. 若“合同金额”总和 > 0.1，销售类型标“正常销售”
    Step3.其他情况标正常销售
    """
    # Step1：标记折让类型
    df["销售类型"] = df["下单合同号"].apply(lambda x: "折让" if "折让" in x else "")
    # Step2：处理非折让订单
    # 筛选条件：销售类型为空且合同金额<0的合同号
    df["合同金额"] = df["合同金额"].astype(float)
    df.loc[df["销售类型"].isnull(), "销售类型"] = ""
    contract_groups = df[(df["销售类型"] == '')].groupby(['下单合同号', '客户名称', '销售员', '出具发票日'])
    for (contract_num, customer, seller, date), group in contract_groups:
        # 计算金额总和
        total_amount = group['合同金额'].sum()
        if abs(total_amount) <= 0.1:
            df.loc[group.index[group['出具发票日'] == date], '销售类型'] = "冲红"
        elif total_amount < -0.1:
            df.loc[group.index[group['出具发票日'] == date], '销售类型'] = "退货"
        else:
            df.loc[group.index[group['出具发票日'] == date], '销售类型'] = "正常销售"
    # Step3：未处理数据兜底
    df.loc[df["销售类型"] == '', "销售类型"] = "正常销售"
    # '''

    # 匹配运输方式
    df_HWOrder = pd.DataFrame()
    for path in HWOrderPathList:
        orderTempDf = pd.read_excel(path, dtype=str).fillna("")
        df_HWOrder = df_HWOrder.append(orderTempDf)
    HWOrderDict = dict(zip(df_HWOrder["华为订单号"], df_HWOrder["运输方式"]))
    # df_KT = pd.read_excel(KTconfigPath, dtype=str)
    # df_KT.rename(columns={"供货方编号": "下单合同号"}, inplace=True)
    # KTOrderDict = dict(zip(df_KT["下单合同号"].fillna(""), df_KT["运输方式"].fillna("鲲泰外挂表运输方式为空")))
    # df["运输方式"] = df.apply(matchYSFS, args=(BoTransDict, HWOrderDict, KTOrderDict), axis=1)
    df["运输方式"] = df.apply(matchYSFS, args=(BoTransDict, HWOrderDict), axis=1)

    # 匹配是否直发
    df["是否直发"] = df["运输方式"].apply(lambda x: "否" if x == "自提" else ("是" if x in ["汽运", "空运"] else ""))

    # 清理内存
    gc.collect()

    return df


# 匹配”成本总价”、“月份”
def calDataStep5(df: pd.DataFrame, fwyrPath):
    """
    :param df: 需要处理的df
    :param fwyrPath: 服务预提表路径
    :param : 华为订单表字典{下单合同号：运输方式}
    :return: 处理后的df
    """
    # 匹配”成本总价”
    df["成本总价"] = df["订单成本（利润中心货币）"]
    totalOrderList = df["销售订单号"].tolist()
    # Step1:读取服务预提表，合并销售订单号相同的金额
    df_fwyt = pd.read_excel(fwyrPath, dtype={"销售订单号": str})
    df_fwyt.drop(df_fwyt[df_fwyt["销售订单号"].isna()].index, inplace=True)
    yt_pivot = df_fwyt.pivot_table(index=["销售订单号"], values="技服预提金额(RMB)", aggfunc="sum")
    addFWYTdf = pd.DataFrame()
    # Step2:遍历预提数据，判断订单号是否在销售明细表中出现，出现取销售明细该订单号的第一条数据，
    #       将金额相关列变为空，成本总价依据订单号个数取”技服预提金额(RMB)“平均值
    for i_, series_ in yt_pivot.iterrows():
        orderList = series_.name.replace("，", ",").strip(",").split(",")
        for num in orderList:
            if num in totalOrderList:
                tempSeries = df.query("销售订单号 == @num").iloc[0]
                tempSeries[["返款折扣×汇率", "订单成本（利润中心货币）", "合同金额", "合同不含税金额"]] = ""
                tempSeries["成本总价"] = series_["技服预提金额(RMB)"] / len(orderList)
                addFWYTdf = addFWYTdf.append(tempSeries)
    else:
        # Step3:将所有在预提表中出现的数据按Step2复制处理后添加到原数据集中
        df = df.append(addFWYTdf)

    # 匹配”月份”
    df["月份"] = df["出具发票日"].map(calMonth)

    # 清理内存
    gc.collect()

    return df


# 匹配”产品”、”产品线”:
def calDataStep6(df: pd.DataFrame, HWYJPathList, productConfPath, saveDir):
    """
    :param df: 需要处理的df
    :param HWYJPathList: 华为业绩表路径列表
    :param productConfPath: 产品线外挂表路径
    :param saveDir: 结果文件保存路径
    :return: 销售明细结果文件路径
    """

    # 匹配”产品”、”产品线”:
    df_conf = pd.read_excel(productConfPath, sheet_name="产品线", dtype={"产品": str, "产品线": str}).fillna("")
    productDict = dict(zip(df_conf["产品"], df_conf["产品线"]))

    # 读取华为业绩表
    df_YJ = pd.DataFrame()
    for path in HWYJPathList:
        yjTempDf = pd.read_excel(path, dtype=str).fillna("")
        df_YJ = df_YJ.append(yjTempDf)
    df_YJ["产品"] = df_YJ["产品子类"] + df_YJ["产品分类标签"]
    # 由于“经销商业绩”列有空值，将该列转为数字类型，不能转换的默认为NaN
    df_YJ["经销商业绩"] = pd.to_numeric(df_YJ['经销商业绩'], errors='coerce')
    # 将数据透视，对于同一"华为合同号"的不同"产品"，取"经销商业绩"最高的一项
    # yj_pivot = df_YJ.pivot_table(index=["华为合同号", "产品"], values=["经销商业绩", "业务部", "客户分类"],
    #                              aggfunc={"经销商业绩": "sum", "业务部": "first", "客户分类": "first"})
    yj_pivot = df_YJ.pivot_table(index=["华为合同号", "产品"], values=["经销商业绩", "业绩业务类型", "客户分类"], aggfunc={"经销商业绩": "sum", "业绩业务类型": "first", "客户分类": "first"})
    yj_pivot.reset_index(inplace=True)
    yj_pivot.sort_values(by="经销商业绩", inplace=True, ascending=False)
    yj_pivot.drop_duplicates("华为合同号", keep="first", inplace=True, ignore_index=True)
    yj_pivot["产品线"] = yj_pivot["产品"].apply(lambda x: productDict.get(x, "产品线外挂表无对应产品线"))
    yj_pivot["市场类型"] = yj_pivot.apply(calMarketType1, axis=1)
    confDict = dict(zip(yj_pivot["华为合同号"], yj_pivot[["产品", "产品线", "市场类型"]].values.tolist()))

    df[["产品", "产品线", "市场类型"]] = df.apply(matchProduct, args=(confDict,), axis=1)

    # 数据整理：对于“下单合同号”已匹配到的数据，“合并前批次”列数据变为""
    df.loc[df[addCol[-1]] != "", addCol[-1]] = df.loc[df[addCol[-1]] != ""].apply(
        lambda x: x[addCol[-1]] if x["下单合同号"] == "未匹配原因：采购信息表无该批次" else "", axis=1)
    # 数据写入文件
    # 兜底删除多余列（表头为空的Unnamed列、“项目注释”列）
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df = df.drop(columns=["项目注释"], errors="ignore")
    # 兜底删除重复列（如外部表带入的“市场类型.1”、“合并前批次.1”），仅保留第一列
    if df.columns.duplicated().any():
        dupCols = df.columns[df.columns.duplicated()].tolist()
        print(f"警告：结果数据存在重复列，已剔除：{dupCols}")
        df = df.loc[:, ~df.columns.duplicated()]
    # 将“订单类型备注”移到最后一列
    df = df[[c for c in df.columns if c != "订单类型备注"] + ["订单类型备注"]]
    finalPath = os.path.join(saveDir, "销售明细.xlsx")
    df.to_excel(finalPath, sheet_name="销售明细", index=False)

    # 清理内存
    gc.collect()

    return finalPath


# 设置结果表格式
def setStyle(finalPath, originalCols):
    """
    :param finalPath: 结果表路径
    :param originalCols: 原销售日报列数，用于设置格式
    :return:
    """
    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finalPath)
    ws = wb.sheets["销售明细"]
    cols = ws.used_range.shape[1]

    # 表头列表
    allColumnsList = ws.used_range.rows[0].value

    personCol = get_column_letter(allColumnsList.index("销售员编码") + 1)
    batchCol = get_column_letter(allColumnsList.index("批次") + 1)
    contractCol = get_column_letter(allColumnsList.index("下单合同号") + 1)

    # 计算标志列（原表数据列终止位置-3、新增数据开始位置-3和结束位置）（"事业部"、"区域"、"平台"作为新增数据列进行颜色处理）
    colIndex1 = get_column_letter(originalCols - 3)
    colIndex2 = get_column_letter(originalCols - 2)
    colIndex3 = get_column_letter(cols)

    # 全表格式改成常规格式、表头字体为微软雅黑9号、内容字体为微软雅黑9号
    ws.range(f"A:{colIndex3}").number_format = "G/通用格式"
    ws.range(f"A:{colIndex3}").font.size = 9
    ws.range(f"A:{colIndex3}").font.name = "微软雅黑"
    ws.range(f"{personCol}:{personCol}").number_format = "@"  # "销售员编码"列为文本格式
    ws.range(f"{batchCol}:{batchCol}").number_format = "@"  # "批次"列为文本格式
    ws.range(f"{contractCol}:{contractCol}").number_format = "@"  # "下单合同号"列为文本格式

    # plan A：已经用本文格式存储的数字，需要进行分列操作,但会导致部分列数据损失
    # colIndexList = get_column_interval(1, cols)
    # for indexCol in colIndexList:
    #     ws.range(indexCol+"1").expand("down").api.TextToColumns()

    # plan B:由于单元格格式已经为常规，将所有数据复制重新粘贴可去除文本格式显示的数字（表现为绿三角）
    ws.range("A1").value = ws.used_range.value

    # 进行冻结操作(F2)
    active_window = wb.app.api.ActiveWindow
    active_window.FreezePanes = False
    # wb.app.range("A2").select()  # 选"A2"冻结首行
    active_window.SplitColumn = allColumnsList.index("客户名称") + 1  # 冻结至哪一列
    active_window.SplitRow = 1  # 冻结至哪一行
    active_window.FreezePanes = True

    # 原有的数据标题设置背景色、加粗
    ws.range("A1:%s1" % colIndex1).color = (0, 112, 192)
    ws.range("A1:%s1" % colIndex1).font.color = (255, 255, 255)
    ws.range("A1:%s1" % colIndex1).font.bold = False

    # 新增的数据标题设置背景色、字体颜色、加粗、数字格式
    ws.range("%s1:%s1" % (colIndex2, colIndex3)).color = (255, 192, 0)
    ws.range("%s1:%s1" % (colIndex2, colIndex3)).font.color = (0, 0, 0)
    ws.range("%s1:%s1" % (colIndex2, colIndex3)).font.bold = True
    # "合并前批次"列背景色变为红色（按列名定位，避免“订单类型备注”移到末列后被误标红）
    mergeCol = get_column_letter(allColumnsList.index("合并前批次") + 1)
    ws.range(f"{mergeCol}1").color = (255, 0, 0)

    # 新增数据设置格式, "实际税率"保留两位小数，其他数字列为千分位显示整数
    for col in ["订单成本（利润中心货币）", "实际税率", "合同金额", "合同不含税金额", "成本总价"]:
        index = get_column_letter(allColumnsList.index(col) + 1)
        if col == "实际税率":
            ws.range(f"{index}:{index}").number_format = '0.00'
        else:
            ws.range(f"{index}:{index}").number_format = "#,##0_ "

    # 自适应宽度
    # ws.autofit()
    # 设置列宽，先将所有列列宽设置为10，"下单合同号","项目名称","评审二代"列宽为15
    ws.used_range.column_width = 10
    for colName in ["下单合同号", "项目名称", "评审二代"]:
        col = get_column_letter(allColumnsList.index(colName) + 1)
        ws.range(f"{col}:{col}").column_width = 15

    wb.save(finalPath)
    wb.close()
    app.quit()

# 遍历文件夹，获取最新的FY26销售明细(*月)_****.xlsx 文件
def findLatestSalesDetailFile(baseDir, maxBackDays=90):
    """
    从今天日期的文件夹开始，向前回溯查找销售明细Excel文件。
    文件夹结构：baseDir\2026年\6月\18日
    文件名格式：FY**销售明细(*月)_****.xlsx（如 FY26销售明细(6月)_0617.xlsx）

    :param baseDir: 基础目录路径，如 F:/Uibot项目文件/result
    :param maxBackDays: 最大回溯天数，默认90天，超过此范围仍未找到则返回None
    :return: 找到的文件完整路径，若未找到返回None
    """
    today = datetime.today()

    for i in range(maxBackDays):
        checkDate = today - timedelta(days=i)
        yearStr = str(checkDate.year)
        monthStr = str(checkDate.month)
        dayStr = str(checkDate.day)

        targetDir = os.path.join(baseDir, f'{yearStr}年', f'{monthStr}月', f'{dayStr}日')

        if not os.path.exists(targetDir):
            continue

        prevDate = checkDate - timedelta(days=1)
        prevMonthStr = str(prevDate.month).zfill(2)
        prevDayStr = str(prevDate.day).zfill(2)
        file_name = f'FY{yearStr[2:4]}销售明细({int(prevMonthStr)}月)_{prevMonthStr}{prevDayStr}.xlsx'
        filePath = os.path.join(targetDir, file_name)

        if os.path.exists(filePath):
            return filePath

    return ''


# 将模板文件复制为目标文件
def copyFile(templatePath, targetPath):
    """
    :param templatePath: 模板文件路径
    :param targetPath: 目标文件路径
    :return: 目标文件路径
    """
    shutil.copyfile(templatePath, targetPath)
    return targetPath


"""
extraOrderList:销售明细补充表中增加数据“销售订单号”值的列表
BoDeliveryDict:BO采购信息表发货方式字典
dateColList: 读取销售日报结果表或销售明细补充表需要处理的时间列
matchCol：匹配的列
addCol：匹配的三列+"合并前批次"（数据对应的合并前批次）
boMatchCol：BO表格中对应matchCol的匹配列
usedCol: 销售明细表需要的列
calAmountCol：匹配订单金额列
ignoreCharsList：BO采购信息表中“华为_厂商PO号”忽略的字符列表
matchFlag：用于防止df.apply对第一条数据重复操作
logger：用于打印日志
calYearMonth：操作的年月 yyyymm
"""
extraOrderList = ["激励抵罚息", "转储"]
BoDeliveryDict = {"直发": "汽运", "非直发": "自提", "": "自提"}
dateColList = ["出具发票日", "预计还款日"]
matchCol = ["下单合同号", "项目名称", "评审二代"]
matchColadd = matchCol + ["削价责任人"]
addCol = matchCol + ["合并前批次"]
boMatchCol = ["华为_厂商PO号", "项目名称(查询 1 用 系统科技销售管理采购信息)", "签约客户名称"]
usedCol = ["销售订单号", "销售订单行项目", "订单类型", "产品组", "客户编号", "客户名称",
           "出具发票日", "预计还款日", "物料号", "物料名称", "数量",
           "工厂", "库存地", "批次", "合同号（客户PO号）", "返款折扣×汇率",
           "订单成本（利润中心货币）", "会计凭证号", "销售员", "销售员编码", "合同金额",
           "实际税率", "合同不含税金额", "事业部", "区域", "平台"]
calAmountCol = "订单成本（利润中心货币）"
ignoreCharsList = [";", "；", " "]
matchFlag = False
logger = None
calYearMonth = ""

if __name__ == "__main__":
    # aa = getQryTimeRange(r'D:\xc_files\毛利分析\汇总表', "订单全字段报表", "%Y-%m-%d", 'kuntai_pm')
    # bb = findLatestSalesDetailFile(r'F:\Uibot项目文件\result')

    addfilePath = r"D:\xc_files\毛利分析\0831\物料移动明细（20260830-20260830）.MHTML"
    # addfilePath = r"D:\xc_files\毛利分析\物料移动明细（20240604-20240604）.MHTML"
    finalPath = r"D:\xc_files\毛利分析\0831\物料移动明细汇总_20260829.xlsx"
    dateFlag = "2026/08/30"
    finalPathAdd = r"D:\xc_files\毛利分析\0831\物料移动明细汇总(新增列)_20260829.xlsx"
    BO_file = r"D:\xc_files\毛利分析\BO采购信息表.xls"
    cc = handleMovementDetail(addfilePath, finalPath, dateFlag, finalPathAdd, BO_file)
    print(cc)

    # g_dictGlobal = {"销售日报": r"C:\Users\11598\Desktop\测试文件\FY23销售明细(2月)_0205.xlsx",
    #                 "分销销售名单": r"C:\Users\11598\Desktop\测试文件\分销销售名单.xlsx",
    #                 "BO下载路径": r"C:\Users\11598\Desktop\测试文件\BO采购信息表.xls",
    #                 "销售明细补充表路径": r"C:\Users\11598\Desktop\测试文件\销售明细补充.xlsx",
    #                 "物料移动明细": r"C:\Users\11598\Desktop\测试文件\物料移动明细汇总_20230205.xlsx",
    #                 "OA预提表路径": r"C:\Users\11598\Desktop\测试文件\预提表_20230206.xlsx",
    #                 "产品线": r"C:\Users\11598\Desktop\测试文件\产品线-22年.xlsx",
    #                 "结果保存路径": r"C:\Users\11598\Desktop\测试文件",
    #                 }
    # g_dictGlobal = {"销售日报": r"D:\xc_files\销售明细\731\FY26销售明细(8月)_0803.xlsx",
    #                 "分销销售名单": r"D:\xc_files\销售明细\分销销售名单.xlsx",
    #                 "BO下载路径": r"D:\xc_files\销售明细\BO采购信息表.xls",
    #                 "CRM外挂表路径": r"D:\xc_files\销售明细\CRM外挂表.xlsx",
    #                 "销售明细补充表路径": r"D:\xc_files\销售明细\销售明细补充.xlsx",
    #                 "物料移动明细": r"D:\xc_files\销售明细\物料移动明细汇总_20260803.xlsx",
    #                 "OA预提表路径": r"D:\xc_files\销售明细\预提表_20260804.xlsx",
    #                 "产品线": r"D:\xc_files\销售明细\产品线-22年.xlsx",
    #                 "结果保存路径": r"D:\xc_files\销售明细\result",
    #                 }
    # g_selectPath = g_dictGlobal["销售日报"]
    # BO_filePath = g_dictGlobal["BO下载路径"]
    # crm_file = g_dictGlobal["CRM外挂表路径"]
    # ytPath = g_dictGlobal["OA预提表路径"]
    # moveFilePath = g_dictGlobal["物料移动明细"]
    # orderFileList = getSameFormatFile(g_dictGlobal["结果保存路径"], "订单表")
    # yjFileList = getSameFormatFile(g_dictGlobal["结果保存路径"], "业绩表")
    #
    # # initWriteLog(r"C:\Users\11598\Desktop\test\log(exe_cmd)\202207")
    #
    # # 初始化func中的全局变量calYearMonth（处理日期: yyyymm）
    # initOperateDate(g_selectPath)
    #
    # # 筛选销售日报有效数据, 返回df：筛选后的的df；colNum：原始数据列数
    # df, colNum = match_validData(g_selectPath, ["00049539", "00074453"], g_dictGlobal["分销销售名单"])
    #
    # # 增加销售明细补充表数据
    # # initMatchDf: “销售明细补充”外挂表中人工新增的[下单合同号, 项目名称, 评审二代]的数据
    # df, initMatchDf = addExtraData(df, g_dictGlobal["销售明细补充表路径"])
    #
    # # ==============================================
    # # 【唯一修改：提前标记折让，不影响任何原有逻辑】
    # # ==============================================
    # mask_zhe = (
    #     df["合同号（客户PO号）"].str.contains("折让", na=False) |
    #     df["物料名称"].str.contains("折让", na=False) |
    #     (df["销售订单号"] == "返款抵欠款")
    # )
    # df.loc[mask_zhe, matchCol] = ["折让", "", ""]
    #
    # # 依据"物料名称"、"批次"的内容初步匹配"下单合同号"、"项目名称"、"评审二代"
    # df = calDataStep1(df)
    #
    # # 通过批次匹配BO采购信息表的"下单合同号"、"项目名称"、"评审二代"
    # # matchDict：BusinessObjects采购信息字典{批次：[下单合同号, 项目名称, 评审二代]}
    # # BoTransDict: BO下单合同号对应的运输方式字典{下单合同号：运输方式}
    # df, matchDict, BoTransDict = calDataStep2(df, BO_filePath)
    #
    # # 通过”项目注释“匹配"下单合同号"、通过CRM外挂表匹配"项目名称"、"评审二代"
    # df = calDataStep2_crm(df, crm_file, matchDict)
    #
    # # 通过物料移动明细表匹配"下单合同号"、"项目名称"、"评审二代"
    # df = calDataStep3(df, moveFilePath, matchDict)
    # # "下单合同号"、"项目名称"、"评审二代"初步匹配完成后，再写入人工补充的数据
    # df = matchExtraData(df, initMatchDf)
    #
    # allOrderList = glob.glob(r"E:\Uibot项目文件\汇总表\*订单表*.xlsx")
    # # 匹配“采购类型”、“销售类型”、“运输方式”、“是否直发”
    # df = calDataStep4(df, BoTransDict, orderFileList, KTconfigPath='')
    #
    # # 匹配”成本总价”、“月份”
    # df = calDataStep5(df, ytPath)
    #
    # HWYJPathList = glob.glob(r"E:\Uibot项目文件\汇总表\*业绩表*.xlsx")
    # # 匹配”产品”、”产品线”,返回finalPath：销售明细结果文件路径
    # finalPath = calDataStep6(df, yjFileList, g_dictGlobal["产品线"], g_dictGlobal["结果保存路径"])
    #
    # # 设置格式
    # setStyle(finalPath, colNum)