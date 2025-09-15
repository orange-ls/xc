#!/usr/bin/python3
# -*- coding: UTF-8 -*-


import gc
import glob
import os
import re
from copy import deepcopy
from datetime import datetime

import pandas as pd
import xlwings as xw
from openpyxl.utils import get_column_letter


# 四舍五入
def new_round(_float, _len):
    """
    :param _float: 需要四舍五入的数
    :param _len: 保留小数点位数
    :return: 四舍五入结果
    """
    if isinstance(_float, float):
        if str(_float)[::-1].find('.') <= _len:
            return _float
        if str(_float)[-1] == '5':
            return round(float(str(_float)[:-1] + '6'), _len)
        else:
            return round(_float, _len)
    else:
        return round(_float, _len)


# 自定义解析时间数据（不对na和空值处理）
def myDateParser(x):
    """
    :param x: 传入的时间数据
    :return: 返回%Y/%m/%d形式的文本
    """
    if x and not pd.isna(x):
        return datetime.strftime(x, "%Y/%m/%d")
    else:
        return ""


# 初始化操作日期
def initOperateDate(selectPath):
    """
    :param selectPath:选择的销售日报路径
    :return:
    """

    global calYearMonth
    calYearMonthGroup = re.search(".*FY(\d{2}).*\((\d{1,2})月\).*", os.path.basename(selectPath))
    calYearMonth = f"20{calYearMonthGroup.group(1)}{calYearMonthGroup.group(2).zfill(2)}"


# 将两张表的数据格式进行统一，如表1列为['a', 'b', 'c']，表2列为['c', 'b1', 'b2', 'a']，
# 则表1和表2均统一为['a', 'b', 'c', 'b1', 'b2']的形式
def fitExcel(wb_original, wb_new):
    """
    :param wb_original: 原始表的workbook
    :param wb_new: 新表的workbook
    :return: 适配后的workbook
    """

    # step1：获取原表和新表的列名列表
    ws1 = wb_original.sheets[0]
    originalCol = ws1.used_range.rows[0].value
    ws2 = wb_new.sheets[0]
    newCol = ws2.used_range.rows[0].value

    # step2：遍历新表的列，与原表相比新增的列需要写入原表的位置
    insertNum = 0  # 当前新增的列数
    for idx, col in enumerate(newCol):
        if col not in originalCol:
            insertNum += 1
            lastCol = get_column_letter(len(originalCol) + insertNum)
            ws2.range(f"{get_column_letter(idx + 1)}1").copy(destination=ws1.range(f"{lastCol}1"))

    # step3：在新表中新增”Fit“的sheet，用于保存适配后的数据
    sheetNames = [sheet.name for sheet in wb_new.sheets]
    if "Fit" not in sheetNames:
        wb_new.sheets.add("Fit", after=sheetNames[0])
    ws_ = wb_new.sheets["Fit"]
    ws_.clear()

    """ step4：遍历原表的列（经过step2，列名已包含新增的列），与原表相比：
            1.新表含有该列，将新表该列数据复制到sheet“Fit”中
            2.新表不含该列，将原表该列名所在的单元格复制到新表的sheet“Fit”中
    """
    finalCol = ws1.used_range.rows[0].value
    copyNum = 1
    for idx, col in enumerate(finalCol):
        # 新sheet“Fit”中当前操作列数的列标识
        desLetter = get_column_letter(copyNum)
        if col in newCol:
            colLetter = get_column_letter(newCol.index(col) + 1)
            ws2.range(f"{colLetter}:{colLetter}").copy(destination=ws_.range(f"{desLetter}1"))
        else:
            # insetLetter = get_column_letter(copyNum + 1)
            # ws_.range(f"{insetLetter}:{insetLetter}").insert()
            ws1.range(f"{get_column_letter(idx + 1)}1").copy(destination=ws_.range(f"{desLetter}1"))
        copyNum += 1
    wb_new.save()
    return ws_


# 合并华为订单全字段报表
# @logfun
def updateAllFieldFile(addfilePath, finalPath, dateFlag):
    """
    :param addfilePath: 读取未处理的华为订单全字段报表
    :param finalPath: 华为订单全字段报表汇总表路径
    :param dateFlag: 文件名的更新日期 %Y-%m-%d %H:%M:%S
    :return: 返回华为订单全字段报表汇总表路径
    """
    # 华为订单全字段报表在指定日期内有数据（无数据不下载，addfilePath为None）
    if addfilePath:
        # 初始化App
        app = xw.App(visible=True, add_book=False)
        app.display_alerts = False
        app.screen_updating = True
        # 打开下载表
        wb_ = app.books.open(addfilePath)
        ws_ = wb_.sheets[0]
        # 删除第一行标题行
        newCols = ws_.used_range.rows[0].value
        ws_.range("1:1").delete()

        # 打开汇总表
        wb = app.books.open(finalPath)
        ws = wb.sheets[0]
        oldCols = ws.used_range.rows[0].value
        if oldCols != newCols:
            raise Exception(f"汇总表{finalPath}和下载表{addfilePath}列不同")
        startRow = ws.used_range.shape[0] + 1

        # 将需要添加的数据复制到汇总表中
        ws_.used_range.copy(destination=ws.range(f"A{startRow}"))

        # ws.autofit()
        wb.save(finalPath)
        # 关闭工作簿
        wb.close()
        wb_.close()
        app.quit()

    # 更新文件名
    fileUser = os.path.basename(finalPath).split("_")[0]
    newFileName = f"{fileUser}_订单全字段报表_" + dateFlag.replace("-", "")[:8] + ".xlsx"
    newFilePath = os.path.join(os.path.dirname(finalPath), newFileName)
    os.rename(finalPath, newFilePath)

    # 清理内存
    gc.collect()
    return newFilePath


# 更新"运费"、"现金返点"记录文件
# @logfun
def updateRecordDetail(recordPath, recordDict, recordDictCJB):
    """
    :param recordPath:"运费"、"现金返点"记录文件路径
    :param recordDict: 通过"华为订单全字段报表"获取"运费"、"现金返点"的"华为合同号"记录字典
    :param recordDictCJB: 通过"超聚表"外挂表获取"运费"、"现金返点"的"华为合同号"记录字典
    :return "运费"、"现金返点"记录文件路径
    """

    # 打开汇总表
    app = xw.App(visible=True, add_book=False)
    app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(recordPath)

    targetDict = {"全字段报表": recordDict, "超聚变": recordDictCJB}
    for key, recordDict_ in targetDict.items():
        # 转换字典格式 {华为合同号:{年月:{运费:金额, 现金返点:金额}}}} -> {华为合同号:[str{年月:{运费:金额, 现金返点:金额}}]},用于转为DataFrame
        dfDict = {key: [str(value)] for key, value in recordDict_.items()}
        recordDf = pd.DataFrame(data=dfDict).T.reset_index()

        # 清空数据并重新写入汇总表
        ws = wb.sheets[key]
        ws.used_range.clear_contents()
        ws.range("A1").value = ["华为合同号", "记录"]
        ws.range("A2").value = recordDf.values

    # ws.autofit()
    wb.save(recordPath)
    # 关闭工作簿
    wb.close()
    app.quit()

    return recordPath


# 对销售明细透视并补充“成本调整”外挂表初始化数据
def init_PivotData(filepath, costAdjustPath):
    """
    :param filepath: 销售明细结果表
    :param costAdjustPath: “成本调整”外挂表
    :return: 初始化数据后的df
    """
    df = pd.read_excel(filepath, sheet_name="销售明细", dtype=str, keep_default_na=False, parse_dates=dateColList,
                       date_parser=lambda x: myDateParser(x)).fillna("")
    for col in pivotVal + ["实际税率"]:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    # 筛选出销售类型不为冲红的数据进行透视(pivotVal列为求和，产品、产品线、市场类型列取第一个值)
    # df = df.query("销售类型 != '冲红'")
    df = df.query(
        "销售类型 != '冲红'"
        " and 采购类型 != '折让'"  # 剔除采购类型为折让
        " and (采购类型 != '价外费用' or `采购类型-二级分类` == '折旧费')"  # 保留价外费用的折旧费
        # " and not (采购类型 == '价外费用' and `采购类型-二级分类` in ['罚息','物流费','诉讼费&律师费'])"  # 剔除价外费用的特定分类
    )
    funDict = {col: "sum" for col in pivotVal}
    funDict.update({col: "first" for col in ["销售类型", "产品", "产品线", "市场类型"]})
    df = df.pivot_table(index=pivotIndex, values=pivotVal + ["销售类型", "产品", "产品线", "市场类型"], aggfunc=funDict)
    # 重置索引，将index变为列
    df.reset_index(inplace=True)

    # 读取“成本调整”外挂表并将数据补充到数据集中
    df_add = pd.read_excel(costAdjustPath, sheet_name=0, dtype=str, parse_dates=["出具发票日"],
                           date_parser=lambda x: myDateParser(x)).fillna("")
    for col in pivotVal + ["实际税率"]:
        df_add[col] = pd.to_numeric(df_add[col], errors='coerce')
    df = df.append(df_add)
    df.reset_index(drop=True, inplace=True)

    # “采购类型”为“折让”，不参与核算；或“采购类型”为“价外费用“，合同号为“罚息”、“物流费“、“诉讼费&律师费”、“价外费用”的数据，不参与核算
    dropIndex = df.loc[
        (df["采购类型"] == "折让") | ((df["采购类型"] == "价外费用") & (df["下单合同号"].isin(["罚息", "物流费", "诉讼费&律师费", "价外费用"])))].index
    df = df.drop(index=dropIndex)
    df.reset_index(drop=True, inplace=True)

    return df


# 计算"运费", "现金返点"
def matchProject1(series, originalDict, originalDictCJB, hisRecordDict):
    """
    :param series: DataFrame行series
    :param originalDict: "华为全字段报表"获取的"运费"、"现金返点"字典   {华为合同号:["运费", "现金返点"]}
    :param originalDictCJB: "超聚表"外挂表获取的"运费"、"现金返点"字典    {华为合同号:["运费", "现金返点"]}
    :param hisRecordDict: 人工处理历史记录字典{华为合同号:[运费, 现金折扣金额]}
    :return: 返回"运费", "现金返点"
    """

    # 计算“华为合同号”各月的运费(现金返点)记录值总和
    def calSum(dict_, key):
        """
        :param dict_: “华为合同号”的"运费", "现金返点"记录字典 {年月：{运费：金额，现金返点：金额}}
        :param key: 计算的指标(运费/现金返点)
        :return: 返回运费(现金返点)的记录值总和
        """
        sumVal = 0
        for valueDict in dict_.values():  # valueDict {运费：金额，现金返点：金额}
            sumVal += valueDict[key]
        return sumVal

    # 获取本次使用的金额
    def getData(orderNum_, originalDict_, orderReDict_, hisReDict_, tax_):
        """
        :param orderNum_: 下单合同号
        :param originalDict_: "华为全字段报表"或"超聚变"外挂表获取的"运费"、"现金返点"字典  {华为合同号:["运费", "现金返点"]}
        :param orderReDict_: “华为合同号”的"运费", "现金返点"记录字典 {年月：{运费：金额，现金返点：金额}}
        :param hisReDict_: 人工处理历史记录字典{华为合同号:[运费, 现金折扣金额]}
        :param tax_: 实际税率（全字段报表中的金额为含税，实际计算需要不含税的数据）
        :return: 返回本次"运费", "现金返点"的使用金额
        """
        # 获取运费、现金返点总值
        originalShip = originalDict_[orderNum_][0] / (1 + tax_)
        originalCash = originalDict_[orderNum_][1] / (1 + tax_)
        # 获取运费、现金返点人工历史处理记录值
        historyShipRe = hisReDict_.get(orderNum_, [0, 0])[0]
        historyCashRe = hisReDict_.get(orderNum_, [0, 0])[1]
        # 计算运费、现金返点的使用金额
        shipRecord = calSum(orderReDict_, "运费")
        cashRecord = calSum(orderReDict_, "现金返点")
        # 计算运费、现金返点各自差值
        shipDiff = new_round(originalShip - shipRecord - historyShipRe, 2)
        cashDiff = new_round(originalCash - cashRecord - historyCashRe, 2)

        return [shipDiff, cashDiff]

    # DataFrame.apply第一行数据会操作两次，导致重复开单，加入matchFlag忽略第一次操作
    global matchFlag, recordDict, recordDictCJB, ctOrderList, hsOrderList, bsOrderList
    if matchFlag == False:
        matchFlag = True
        return pd.Series(data=["无效数据（忽略）"] * 3, index=matchCol + ["备注"])

    buyType = series["采购类型"]
    orderNum = series["下单合同号"]
    tax = series["实际税率"]
    # 判断是否是‘城投’账号的数据，如是‘备注’列写入‘城投’
    remarkText = "城投" if orderNum in ctOrderList else "合神" if orderNum in hsOrderList else "北神" if orderNum in bsOrderList else ""

    if buyType in type1Col:
        return pd.Series(data=[""] * 3, index=matchCol + ["备注"])
    elif series["事业部"] == "服务事业部" and ("保卡" in series["项目名称"] or "考试券" in series["项目名称"]):
        return pd.Series(data=[0, series["成本总价"] * 0.015, remarkText], index=matchCol + ["备注"])
    elif buyType in type2Col:
        # todo:判断实际数据中，业务提供合同号有记录但全字段报表中没有
        if orderNum in originalDict.keys():
            # 获取“华为合同号”的"运费", "现金返点"记录字典 {年月：{运费：金额，现金返点：金额}}
            orderReDict = deepcopy(recordDict.get(orderNum, {}))
            resultVal = getData(orderNum, originalDict, orderReDict, hisRecordDict, tax)
            # 更新记录字典
            if resultVal[0] != 0 or resultVal[1] != 0:
                shipRecord = orderReDict.get(calYearMonth, {}).get("运费", 0)
                cashRecord = orderReDict.get(calYearMonth, {}).get("现金返点", 0)
                orderReDict[calYearMonth] = {"运费": new_round(shipRecord + resultVal[0], 2),
                                             "现金返点": new_round(cashRecord + resultVal[1], 2)}
                recordDict[orderNum] = orderReDict

            return pd.Series(data=resultVal + [remarkText], index=matchCol + ["备注"])
        else:
            return pd.Series(data=[""] * 2 + [remarkText], index=matchCol + ["备注"])
    elif buyType in type3Col:
        if orderNum in originalDictCJB.keys():
            # 获取“华为合同号”的"运费", "现金返点"记录字典 {年月：{运费：金额，现金返点：金额}}
            orderReDict = deepcopy(recordDictCJB.get(orderNum, {}))
            resultVal = getData(orderNum, originalDictCJB, orderReDict, hisRecordDict, tax)
            # 更新记录字典
            if resultVal[0] != 0 or resultVal[1] != 0:
                shipRecord = orderReDict.get(calYearMonth, {}).get("运费", 0)
                cashRecord = orderReDict.get(calYearMonth, {}).get("现金返点", 0)
                orderReDict[calYearMonth] = {"运费": new_round(shipRecord + resultVal[0], 2),
                                             "现金返点": new_round(cashRecord + resultVal[1], 2)}
                recordDictCJB[orderNum] = orderReDict
            return pd.Series(data=resultVal + [""], index=matchCol + ["备注"])
        else:
            return pd.Series(data=[""] * 3, index=matchCol + ["备注"])


# @logfun
# 计算"运费", "现金返点"并更新记录表
def calDataStep1(df: pd.DataFrame, allFieldfpList, cjbFp, recordfp, historyRecordfp):
    """
    :param df: 需要处理的df数据
    :param allFieldfpList: “华为订单全字段报表”路径列表
    :param cjbFp: “超聚表”外挂表
    :param recordfp: "华为全字段报表"和 “超聚表”外挂表的“运费”、“现金返点”记录文件路径
    :param historyRecordfp: 手工记录的"运费"、"现金返点"历史记录表
    :return: 处理后df
    """
    global recordDict, recordDictCJB, ctOrderList, hsOrderList, bsOrderList
    # ctOrderList: ‘城投’账号‘订单全字段报表’中的“华为合同号”列表
    # hsOrderList: ‘合神’账号‘订单全字段报表’中的“华为合同号”列表
    # bsOrderList: ‘北神’账号‘订单全字段报表’中的“华为合同号”列表
    ctOrderList, hsOrderList, bsOrderList = [], [], []

    # 读取并合并所有的“华为订单全字段报表”
    df_AF = pd.DataFrame()
    for path in allFieldfpList:
        fieldTempDf = pd.read_excel(path, dtype=str)
        userName = os.path.basename(path).split("_")[0]
        if userName == "13544480167":
            ctOrderList = fieldTempDf["华为合同号"].tolist()
        elif userName == "hfszsm":
            hsOrderList = fieldTempDf["华为合同号"].tolist()
        elif userName == "szshbj":
            bsOrderList = fieldTempDf["华为合同号"].tolist()

        df_AF = df_AF.append(fieldTempDf)
    df_AF = df_AF.query("草案状态 == '决策通过'")
    # 读取“超聚表”外挂表
    df_CJB = pd.read_excel(cjbFp, dtype=str)
    # 将“合同编号”去除前后空格，并将(1)、(2)...及（1）、（2）... 变为空
    df_CJB["合同编号"] = df_CJB["合同编号"].str.strip().str.replace(r"\(\d+\)|（\d+）", "")
    if "订单版本" not in df_CJB.columns:
        df_CJB["订单版本"] = 1

    # 将"订单版本", "现金折扣金额", "运费"列数据变为数字类型，且运费列为负值的需*-1，正值的作为0处理
    # todo:看超聚变记录的运费是否有正值，负值需要*-1吗
    for df_ in [df_AF, df_CJB]:
        df_.rename(columns={'现金折扣金额': '现金返点'}, inplace=True)
        for col in ["订单版本", "现金返点", "运费"]:
            df_[col] = pd.to_numeric(df_[col], errors='coerce')
            if col == "运费" and df_.equals(df_AF):
                df_[col] = df_[col].apply(lambda x: new_round(-1 * x, 2) if not pd.isna(x) and x < 0 else 0)

    # 按照“订单版本”排序，取最新“订单版本”的数据
    df_AF.sort_values(by="订单版本", inplace=True, ascending=False)
    df_AF.drop_duplicates("华为合同号", keep="first", inplace=True, ignore_index=True)
    df_CJB.sort_values(by="订单版本", inplace=True, ascending=False)
    df_CJB.drop_duplicates("合同编号", keep="first", inplace=True, ignore_index=True)
    #  获取字典{华为合同号:[运费, 现金折扣金额]}
    originalDict = dict(zip(df_AF["华为合同号"], df_AF[matchCol].fillna(0).values.tolist()))
    originalDictCJB = dict(zip(df_CJB["合同编号"], df_CJB[matchCol].fillna(0).values.tolist()))

    # todo：华为全字段报表记录和超聚变记录是否存入一个sheet
    # 读取"运费"和"现金返点"的记录文件，将“全字段报表”和"超聚变"sheet中当月的记录清空
    recordTotalDf = pd.read_excel(recordfp, sheet_name=None, dtype={"华为合同号": str})
    recordDf = recordTotalDf["全字段报表"]
    recordDfCJB = recordTotalDf["超聚变"]
    # "运费"和"现金返点"的记录字典{华为合同号:{年月:{运费: 金额, 现金返点: 金额}}}
    recordDict = dict(zip(recordDf["华为合同号"], recordDf["记录"].fillna("{}")))
    recordDictCJB = dict(zip(recordDfCJB["华为合同号"], recordDfCJB["记录"].fillna("{}")))
    # 清空本月记录数据，以防重复计算
    for dict_ in [recordDict, recordDictCJB]:
        for key, value in dict_.items():
            value_ = eval(value)
            if calYearMonth in value_.keys():
                value_[calYearMonth] = {"运费": 0, "现金返点": 0}
            dict_[key] = value_

    # 读取手工记录的"运费"、"现金返点"历史记录表，计算时需要扣除已使用额度
    df_his = pd.read_excel(historyRecordfp, sheet_name="记录", dtype={"下单合同号": str})
    df_hisP = df_his.pivot_table(index="下单合同号", values=matchCol, aggfunc="sum")
    # 获取人工处理历史记录字典{华为合同号:[运费, 现金折扣金额]}
    hisRecordDict = dict(zip(df_hisP.index, df_hisP[matchCol].values.tolist()))

    # 匹配 "运费", "现金返点"
    global matchFlag
    matchFlag = False
    df[matchCol + ["备注"]] = df.apply(matchProject1,
                                     args=(originalDict, originalDictCJB, hisRecordDict),
                                     axis=1)

    # 更新记录
    updateRecordDetail(recordfp, recordDict, recordDictCJB)

    for col_ in outCol:
        if col_ not in df.columns:
            df[col_] = ""
            print(f"列【{col_}】没有")
    df = df[outCol]

    # 清理内存
    gc.collect()

    return df


# 向df中写入公式
def calDataStep2(df):
    """
    :param df: 未写入公式的df
    :return: 写入公式后的df
    """
    # 获取各指标在excel中的列
    letter1 = get_column_letter(outCol.index("合同不含税金额") + 1)
    letter2 = get_column_letter(outCol.index("成本总价") + 1)
    letter3 = get_column_letter(outCol.index("运费") + 1)
    letter4 = get_column_letter(outCol.index("现金返点") + 1)
    letter5 = get_column_letter(outCol.index("实际总成本") + 1)
    letter6 = get_column_letter(outCol.index("实际毛利") + 1)

    # 向df写入"实际总成本"、"实际毛利"、"实际毛利率"的公式
    for idx in df.index:
        row = idx + 2
        # 备注“城投”的数据，"实际总成本"公式与普通情况不同
        if df.loc[idx, "备注"] == "城投":
            df.loc[idx, "实际总成本"] = f"={letter2}{row}/1.016+{letter3}{row}+{letter4}{row}"
        else:
            df.loc[idx, "实际总成本"] = f"={letter2}{row}+{letter3}{row}+{letter4}{row}"
        df.loc[idx, "实际毛利"] = f"={letter1}{row}-{letter5}{row}"
        df.loc[idx, "实际毛利率"] = f"={letter6}{row}/{letter1}{row}"
    return df


# 设置毛利分析表格式
def setStyleAndValue(df, finalPath):
    """
    :param df: df数据集
    :param finalPath: 销售明细结果表路径
    :return:
    """
    # 打开结果表
    app = xw.App(visible=True, add_book=False)
    # app.display_alerts = False
    app.screen_updating = True
    wb = app.books.open(finalPath)

    # 判断是否有结果sheet，没有创建
    sheetNames = [sheet.name for sheet in wb.sheets]
    if "账面毛利分析" not in sheetNames:
        wb.sheets.add("账面毛利分析")
    ws = wb.sheets["账面毛利分析"]
    ws.clear()

    # 计算标志列, "销售员编码"、"下单合同号"列格式为文本
    personCol = get_column_letter(outCol.index("销售员编码") + 1)
    contractCol = get_column_letter(outCol.index("下单合同号") + 1)

    # 计算标志列,分别为每种背景色的开始列和截止列
    col1 = get_column_letter(outCol.index(colType1[-1]) + 1)
    col2_s = get_column_letter(outCol.index(colType2[0]) + 1)
    col2_e = get_column_letter(outCol.index(colType2[-1]) + 1)
    col3_s = get_column_letter(outCol.index(colType3[0]) + 1)
    col3_e = get_column_letter(outCol.index(colType3[-1]) + 1)
    col4_s = get_column_letter(outCol.index(colType4[0]) + 1)
    col4_e = get_column_letter(outCol.index(colType4[-1]) + 1)

    # 全表格式改成常规格式、字体为微软雅黑9号
    ws.range(f"A:{col4_e}").number_format = "G/通用格式"
    ws.range(f"A:{col4_e}").font.size = 9
    ws.range(f"A:{col4_e}").font.name = "微软雅黑"
    ws.range(f"{personCol}:{personCol}").number_format = "@"  # "销售员编码"列为文本格式
    ws.range(f"{contractCol}:{contractCol}").number_format = "@"  # "下单合同号"列为文本格式

    # 进行冻结操作
    active_window = wb.app.api.ActiveWindow
    active_window.FreezePanes = False
    # wb.app.range("A2").select()  # 选"A2"冻结首行
    active_window.SplitColumn = 0  # 冻结至哪一列
    active_window.SplitRow = 1  # 冻结至哪一行
    active_window.FreezePanes = True

    # 设置背景色
    ws.range(f"A1:{col1}1").color = (141, 180, 226)
    ws.range(f"{col2_s}1:{col2_e}1").color = (252, 213, 180)
    ws.range(f"{col3_s}1:{col3_e}1").color = (196, 215, 155)
    ws.range(f"{col4_s}1:{col4_e}1").color = (183, 222, 232)

    # 金额数据设置格式, "实际税率"保留两位小数，"实际毛利率"百分比显示，其他数字列为千分位显示整数
    for col in ["实际税率"] + pivotVal + matchCol + colType3[:-1]:
        index = get_column_letter(outCol.index(col) + 1)
        if col == "实际税率":
            ws.range(f"{index}:{index}").number_format = '0.00'
        elif col == "实际毛利率":
            ws.range(f"{index}:{index}").number_format = "0.00%"
        else:
            ws.range(f"{index}:{index}").number_format = "#,##0_ "

    # 写入表头和数据
    ws.range("A1").value = outCol
    ws.range("A2").value = df.values
    # 自适应宽度
    # ws.autofit()
    # 设置列宽，默认所有列列宽为10，G列列宽25，H、I列列宽15
    ws.range(f"A:{col4_e}").column_width = 10
    ws.range("G:G").column_width = 25
    ws.range("H:H, I:I").column_width = 15

    wb.save(finalPath.replace("销售明细", "毛利核算"))
    wb.close()
    app.quit()


"""
dateColList: 读取销售明细表需要处理的时间列
pivotIndex：销售明细表透视需要的行
pivotVal：销售明细表透视需要的值
type1Col：无"运费"、"现金返点"的采购类型
type2Col："运费"、"现金返点"需要从华为全字段报表中获取的采购类型
type3Col："运费"、"现金返点"需要从超聚变外挂表中获取的采购类型
matchCol：需要获取并记录的指标
colType1：标题背景色为第一种的列
colType2：标题背景色为第二种的列
colType3：标题背景色为第三种的列
colType4：标题背景色为第四种的列
outCol：毛利分析结果表标题列表
matchFlag：用于防止df.apply对第一条数据重复操作
logger：用于打印日志
calYearMonth：操作的年月 yyyymm
"""
dateColList = ["出具发票日", "预计还款日"]
pivotIndex = ["出具发票日", "销售员", "销售员编码", "事业部", "区域", "平台", "客户名称", "下单合同号", "项目名称", "是否直发", "实际税率", "采购类型"]
pivotVal = ["返款折扣×汇率", "合同金额", "合同不含税金额", "成本总价"]
type1Col = ["公有云", "价外费用", "鲲泰", "渠道分销", "外购", "折让", "其他", "补充", "创新业务"]
type2Col = ["服务", "原厂下单"]
type3Col = ["超聚变"]
matchCol = ["运费", "现金返点"]
colType1 = pivotIndex[:-1] + pivotVal
colType2 = matchCol
colType3 = ["实际总成本", "实际毛利", "实际毛利率", "备注"]
colType4 = ["销售类型", "采购类型", "产品", "产品线", "市场类型"]
outCol = colType1 + colType2 + colType3 + colType4
matchFlag = False
logger = None
calYearMonth = ""

if __name__ == "__main__":
    g_dictGlobal = {"销售日报": r"E:\Uibot项目文件\result\2022年\7月\1日\FY22销售明细(7月)_0731.xlsx",
                    "分销销售名单": r"E:\Uibot项目文件\配置表\分销销售名单.xlsx",
                    "BO下载路径": r"E:\Uibot项目文件\download\2022年\6月\15日\BO采购信息表.xls",
                    "销售明细补充表路径": r"E:\Uibot项目文件\配置表\销售明细补充.xlsx",
                    "物料移动明细": r"E:\Uibot项目文件\汇总表\物料移动明细汇总_20220807.xlsx",
                    }
    calYearMonth = "202207"
    filePath = r"C:\Users\user\Desktop\销售明细.xlsx"
    recordPath = r"E:\Uibot项目文件\配置表\运费、现金返点记录.xlsx"
    allFieldfpList = glob.glob(r"E:\Uibot项目文件\汇总表\*订单全字段报表*.xlsx")

    df = init_PivotData(filePath, r"E:\Uibot项目文件\配置表\成本调整外挂表.xlsx")
    df = calDataStep1(df, allFieldfpList, r"E:\Uibot项目文件\配置表\超聚变外挂表.xlsx", recordPath,
                      r"E:\Uibot项目文件\配置表\运费、现金返点人工记录.xlsx")
    df = calDataStep2(df)
    setStyleAndValue(df, filePath)
